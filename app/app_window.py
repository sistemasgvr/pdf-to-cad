"""
marcar_utilidades.py — Digitalizar planos y marcar utilidades.

App de escritorio (PySide6) para ingeniería civil: abre el PDF de un plano,
marca utilidades, coloca Leaders (flechas), escribe texto libre,
borra zonas y exporta un único DXF (base digitalizada + tu marcado) en las
mismas coordenadas. Ver menú Ayuda → Manual de usuario.
"""
import sys, os, copy, math, json, zipfile
import fitz
import numpy as np
import ezdxf
from PySide6 import QtCore, QtGui, QtWidgets

import config as C
import vector_pipeline as VP
import geometry as G
import dxf_export
import excel_import
from geo import georef as georef_mod
from geometry import point_in_poly, qimage_to_gray
from model import (VERSION, TIPOS, ACI_RGB, LEADER_TEXT_FT, LEADER_ORIENT,
                   Z_PDF, Z_ERASE, Z_MARK, Z_HANDLE, GRAVITY_LAYERS,
                   TAB_PIPE, TAB_ML, TAB_LEADER, TAB_TEXT, TAB_REGION, TAB_BZ,
                   WORK_UNITS, DEFAULT_WORK_UNIT, is_valid_work_unit, CHANGELOG,
                   PIPE_DIAMETERS_IN, PIPE_MATERIALS, DEFAULT_PIPE_MATERIAL,
                   nearest_pipe_diameter)

DOWNLOADS = os.path.join(os.path.expanduser("~"), "Downloads")
BTN_ON = "background:#2e9e4f;color:white;font-weight:bold;padding:8px;border-radius:4px;"
BTN_OFF = "background:#3c5a99;color:white;padding:8px;border-radius:4px;"


def aci_qcolor(a): return QtGui.QColor(*ACI_RGB.get(a, (235, 235, 235)))
def layer_qcolor(l): return aci_qcolor(C.OUTPUT_LAYERS.get(l, 7))


def _extract_diam_from_size(size_str):
    """Extrae el primer número de un tamaño del catálogo, p.ej. '24 in' → 24.0,
    '12 in x 8 in' → 12.0. Retorna 0.0 si no encuentra número."""
    import re as _re
    if not size_str: return 0.0
    m = _re.match(r"\s*(\d+(?:\.\d+)?)", str(size_str))
    return float(m.group(1)) if m else 0.0


def swatch_icon(color, size=14):
    pm = QtGui.QPixmap(size, size); pm.fill(QtCore.Qt.transparent)
    p = QtGui.QPainter(pm); p.setBrush(color); p.setPen(QtGui.QPen(QtGui.QColor(70, 70, 70)))
    p.drawRect(0, 0, size - 1, size - 1); p.end(); return QtGui.QIcon(pm)


class InlineEdit(QtWidgets.QTextEdit):
    """Editor embebido. Enter = aplicar; Ctrl+Shift+Enter o Shift+Enter = salto de
    línea. El commit se difiere con un timer para no destruir el widget dentro de
    su propio evento (eso provocaba cierres inesperados)."""
    committed = QtCore.Signal(str)

    def __init__(self, text):
        super().__init__(); self.setPlainText(text); self._done = False
        self.setStyleSheet("background:#111;color:#7f7;border:1px solid #7f7;")

    def _commit(self):
        if not self._done:
            self._done = True; txt = self.toPlainText()
            QtCore.QTimer.singleShot(0, lambda: self.committed.emit(txt))

    def keyPressEvent(self, e):
        if e.key() in (QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter):
            m = e.modifiers()
            if (m & QtCore.Qt.ShiftModifier):     # Shift(+Ctrl)+Enter → salto de línea
                self.insertPlainText("\n"); return
            self._commit(); return
        super().keyPressEvent(e)

    def focusOutEvent(self, e):
        super().focusOutEvent(e); self._commit()


class PipelineWorker(QtCore.QThread):
    done = QtCore.Signal(str, str)

    def __init__(self, pdf, tmp): super().__init__(); self.pdf, self.tmp = pdf, tmp

    def run(self):
        try:
            import digitize
            digitize.main(self.pdf, self.tmp, verbose=False); self.done.emit(self.tmp, "")
        except Exception as e:
            import traceback; self.done.emit("", f"{e}\n\n{traceback.format_exc()}")


class Canvas(QtWidgets.QGraphicsView):
    clicked = QtCore.Signal(float, float, object)
    dbl = QtCore.Signal(float, float)
    moved = QtCore.Signal(float, float)

    def __init__(self, win):
        super().__init__(); self.win = win
        self.setScene(QtWidgets.QGraphicsScene(self))
        self.setRenderHints(QtGui.QPainter.Antialiasing | QtGui.QPainter.SmoothPixmapTransform)
        self.setTransformationAnchor(QtWidgets.QGraphicsView.AnchorUnderMouse)
        self.setBackgroundBrush(QtGui.QColor(13, 19, 33)); self.setAcceptDrops(True)
        self.setFocusPolicy(QtCore.Qt.StrongFocus)
        self.setMouseTracking(True); self.viewport().setMouseTracking(True)
        self.pixmap_item = None; self._pan = False; self._pan0 = None; self._moving = False
        self.pdf_opacity = 1.0

    def set_image(self, qimg):
        self.scene().clear()
        self.pixmap_item = self.scene().addPixmap(QtGui.QPixmap.fromImage(qimg))
        self.pixmap_item.setZValue(Z_PDF)
        self.pixmap_item.setOpacity(self.pdf_opacity)
        self.setSceneRect(self.pixmap_item.boundingRect())
        self.resetTransform(); self.fitInView(self.pixmap_item, QtCore.Qt.KeepAspectRatio)

    def set_pdf_opacity(self, val):
        self.pdf_opacity = max(0.1, min(1.0, val))
        if self.pixmap_item: self.pixmap_item.setOpacity(self.pdf_opacity)

    def keyPressEvent(self, e):
        if e.key() in (QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter):
            self.win._on_enter(); e.accept(); return
        if e.key() == QtCore.Qt.Key_Escape:
            self.win._on_escape(); e.accept(); return
        super().keyPressEvent(e)

    def wheelEvent(self, e):
        if self.pixmap_item:
            f = 1.25 if e.angleDelta().y() > 0 else 0.8; self.scale(f, f)

    def mousePressEvent(self, e):
        if e.button() == QtCore.Qt.MiddleButton:
            self._pan = True; self._pan0 = e.position(); self.setCursor(QtCore.Qt.ClosedHandCursor); return
        if not self.pixmap_item: return super().mousePressEvent(e)
        self.setFocus(QtCore.Qt.MouseFocusReason)
        sp = self.mapToScene(e.position().toPoint())
        if self.win.mode == "move" and e.button() == QtCore.Qt.LeftButton:
            self._moving = True; self.win.begin_move(sp.x(), sp.y()); return
        if e.button() in (QtCore.Qt.LeftButton, QtCore.Qt.RightButton):
            self.clicked.emit(sp.x(), sp.y(), e.button()); return
        super().mousePressEvent(e)

    def mouseMoveEvent(self, e):
        if self.pixmap_item:
            sp = self.mapToScene(e.position().toPoint()); self.moved.emit(sp.x(), sp.y())
        if self._pan:
            d = e.position() - self._pan0; self._pan0 = e.position()
            self.horizontalScrollBar().setValue(int(self.horizontalScrollBar().value() - d.x()))
            self.verticalScrollBar().setValue(int(self.verticalScrollBar().value() - d.y())); return
        if self._moving:
            sp = self.mapToScene(e.position().toPoint()); self.win.do_move(sp.x(), sp.y()); return
        super().mouseMoveEvent(e)

    def mouseReleaseEvent(self, e):
        if e.button() == QtCore.Qt.MiddleButton:
            self._pan = False; self.setCursor(QtCore.Qt.ArrowCursor); return
        if self._moving and e.button() == QtCore.Qt.LeftButton:
            self._moving = False; self.win.end_move(); return
        super().mouseReleaseEvent(e)

    def mouseDoubleClickEvent(self, e):
        if self.win.mode == "pipe": self.win.finish_pipe(); return
        sp = self.mapToScene(e.position().toPoint()); self.dbl.emit(sp.x(), sp.y())

    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls(): e.acceptProposedAction()

    def dropEvent(self, e):
        for u in e.mimeData().urls():
            self.win.open_path(u.toLocalFile()); break


class Main(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"Digitalizador de planos — utilidades  (v{VERSION})"); self.resize(1480, 940)
        self.setAcceptDrops(True)
        self.canvas = Canvas(self); self.canvas.clicked.connect(self.on_click)
        self.canvas.dbl.connect(self.on_dblclick); self.setCentralWidget(self.canvas)
        self.zoom = 3.5; self.scale = 20 / 72.0; self.rot = 0; self.W = 0; self.H = 0
        self.derot = fitz.Matrix(1, 0, 0, 1, 0, 0); self.gray = None; self.page_idx = 0; self.pageH_px = 0
        self.pdf_path = None; self.doc = None; self.project_path = None; self.leader_hpx = 40
        self.cur_pts = []; self.pipes = []; self.leaders = []; self.text_marks = []
        self.erase_regions = []; self._erase_pts = []; self.structures = []
        self.mode = "idle"; self._pending = None
        self.snap = False; self.snap_r = 14
        self.sel_pipe = -1; self.sel_leader = -1; self.sel_region = -1; self.sel_text = -1; self.sel_bz = -1
        self._no_center = False
        self._move0 = None; self._drag_vertex = None; self._edit_pts = None; self._edit_closed = False; self._edit_leader = None
        self._move_kind = None; self._moved = False; self._press_xy = None; self._last_xy = None
        self._extending = False; self._ext_layer = None; self._ext_pipe = None; self._ext_at = None
        self._editor = None; self._undo, self._redo, self._overlay = [], [], []
        self._dirty = False; self._style_guard = False; self._prop_guard = False; self._clip = None
        self.georef = georef_mod.Georef()          # georreferenciación (píxel→UTM); inactiva por defecto
        self.work_unit = DEFAULT_WORK_UNIT          # unidad de trabajo del proyecto: 'ft' o 'in' (obligatoria)
        self.show_bz_labels = False                # ¿dibujar el código del buzón al lado del círculo?
        # Detectar versiones de Civil 3D instaladas para escanear su catálogo imperial.
        # civil_year = None si no hay ninguna (se usa igual sin dropdown de familias).
        import civil_catalog as _cc
        _vs = _cc.installed_versions()
        self.civil_year = _vs[-1] if _vs else None
        self._build_ui(); self._apply_style(); self._shortcuts(); self._update_ui()

    # ─────────────────────────── UI ───────────────────────────
    def _build_ui(self):
        mb = self.menuBar()
        mfile = mb.addMenu("&Archivo")
        self._menu_act(mfile, "Abrir PDF…", self.open_pdf)
        self._menu_act(mfile, "Abrir Excel…", self.open_excel)
        mfile.addSeparator()
        self._menu_act(mfile, "Abrir proyecto…", self.open_project)
        self._menu_act(mfile, "Guardar proyecto", self.save_project, "Ctrl+S")
        self._menu_act(mfile, "Guardar proyecto como…", self.save_project_as, "Ctrl+Shift+S")
        mfile.addSeparator()
        self._menu_act(mfile, "Cerrar proyecto", self.close_project, "Ctrl+W")
        medit = mb.addMenu("&Edición")
        self._menu_act(medit, "Deshacer", self.undo, "Ctrl+Z")
        self._menu_act(medit, "Rehacer", self.redo, "Ctrl+Shift+Z")
        mtools = mb.addMenu("&Herramientas")
        self._menu_act(mtools, "Insertar buzón en línea…", self.insert_manhole)
        self._menu_act(mtools, "Importar Excel de red…", self.import_network_excel)
        mtools.addSeparator()
        self._menu_act(mtools, "Editor de catálogo Civil 3D…", self.open_catalog_editor)
        mtools.addSeparator()
        self._menu_act(mtools, "Georreferenciar…", self.open_georef)
        self._menu_act(mtools, "Quitar georreferencia", self.clear_georef)
        mhelp = mb.addMenu("A&yuda")
        self._menu_act(mhelp, "Acerca de…", self.show_about)
        self._menu_act(mhelp, "Manual de usuario", self.show_manual)
        self._menu_act(mhelp, "Atajos de teclado", self.show_shortcuts)

        # ── Barra de acción superior: zoom · deshacer/rehacer · imán · exportar ──
        tb = self.addToolBar("Acciones"); tb.setMovable(False)
        def tact(txt, tip, fn):
            a = QtGui.QAction(txt, self); a.setToolTip(tip); a.triggered.connect(fn); tb.addAction(a); return a
        tact("🔍＋", "Acercar", self._zoom_in); tact("🔍－", "Alejar", self._zoom_out)
        tb.addSeparator(); tact("↶", "Deshacer (Ctrl+Z)", self.undo); tact("↷", "Rehacer (Ctrl+Shift+Z)", self.redo)
        tb.addSeparator()
        self.chk_snap = QtWidgets.QCheckBox("Imán al trazo"); self.chk_snap.toggled.connect(self._toggle_snap); tb.addWidget(self.chk_snap)
        spacer = QtWidgets.QWidget(); spacer.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Preferred); tb.addWidget(spacer)
        # Sin selector de unidad: TODO va en pies por campo (cotas/coordenadas),
        # salvo los diámetros que van siempre en pulgadas (lista fija del catálogo).
        tb.addSeparator()
        tb.addWidget(QtWidgets.QLabel("Civil 3D:"))
        self.cmb_civil = QtWidgets.QComboBox()
        import civil_catalog as _cc
        _all_years = list(_cc.SUPPORTED_YEARS); _inst = set(_cc.installed_versions())
        for y in _all_years:
            self.cmb_civil.addItem(f"{y}{'' if y in _inst else '  (no instalado)'}", y)
        if self.civil_year is not None:
            i = _all_years.index(self.civil_year); self.cmb_civil.setCurrentIndex(i)
        self.cmb_civil.currentIndexChanged.connect(
            lambda i: setattr(self, "civil_year", self.cmb_civil.itemData(i)))
        self.cmb_civil.setToolTip("Versión de Civil 3D. El catálogo imperial se busca en\n"
                                  "C:\\ProgramData\\Autodesk\\C3D <año>\\<idioma>\\Pipes Catalog\\US Imperial Structures")
        tb.addWidget(self.cmb_civil)
        tb.addSeparator()
        self.btn_export = QtWidgets.QPushButton("⭳  Exportar DXF")
        self.btn_export.setStyleSheet("QPushButton{background:#4d8eff;color:#00285d;font-weight:bold;padding:5px 14px;border-radius:4px;} QPushButton:hover{background:#66a3ff;}")
        self.btn_export.clicked.connect(lambda: self.run_pipeline("todo"))
        tb.addWidget(self.btn_export)
        # Capas de referencia reales de LA (checkables ocultas: se activan por menú Herramientas si hace falta).
        self.act_la_ref = QtGui.QAction("Incluir calles reales de LA", self); self.act_la_ref.setCheckable(True)
        self.act_la_parcels = QtGui.QAction("Incluir parcelas de LA", self); self.act_la_parcels.setCheckable(True)

        # ─────────────────────────── DOCK IZQUIERDO ───────────────────────────
        # Aquí construimos el panel lateral izquierdo como un ACORDEÓN de secciones
        # (QToolBox). Cada acción del usuario (Dibujar utilidad, Multileader, Leader,
        # Texto libre, Borrar zona, Georreferenciar, Cotas/red 3D, OCR/ICR) es su
        # propia sección. Solo UNA sección está abierta a la vez, así el usuario
        # ve ÚNICAMENTE las opciones relevantes al paso en el que está.
        #
        # QToolBox = "acordeón" en Qt: contenedor con un botón-cabecera por página.
        # Al hacer clic en una cabecera, esa página se despliega y las demás se
        # colapsan. Es el mismo widget que usamos para el historial de "Acerca de".
        ldock = QtWidgets.QDockWidget("Herramientas", self)
        ldock.setFeatures(QtWidgets.QDockWidget.NoDockWidgetFeatures)     # no se puede sacar/flotar
        left = QtWidgets.QWidget(); lv = QtWidgets.QVBoxLayout(left); lv.setContentsMargins(0, 0, 0, 0)
        self.toolbox = QtWidgets.QToolBox()
        # Guardamos por NOMBRE el índice de cada sección para poder abrirla desde código.
        self._sec_idx = {}

        # ── Helper de creación: crea una página del toolbox con su layout vertical ──
        def _page(title, key):
            page = QtWidgets.QWidget()
            lay = QtWidgets.QVBoxLayout(page); lay.setSpacing(6)
            self.toolbox.addItem(page, title)
            self._sec_idx[key] = self.toolbox.count() - 1
            return page, lay

        # ═══════════════════════════════════════════════════════════════════════
        # PRIMERO: creamos TODOS los widgets (una sola vez) — luego los repartimos
        # en las secciones. Los que se comparten entre secciones (orient_combo,
        # gtxt) los REPARENTAMOS al abrir cada sección (ver _on_toolbox_change).
        # ═══════════════════════════════════════════════════════════════════════

        # Fila de navegación de páginas del PDF
        self.gp = QtWidgets.QWidget(); lp = QtWidgets.QHBoxLayout(self.gp); lp.setContentsMargins(0, 0, 0, 0)
        self.btn_prev = QtWidgets.QPushButton("◀"); self.btn_prev.setFixedWidth(34); self.btn_prev.clicked.connect(self._prev_page)
        self.btn_next = QtWidgets.QPushButton("▶"); self.btn_next.setFixedWidth(34); self.btn_next.clicked.connect(self._next_page)
        self.page_edit = QtWidgets.QLineEdit(); self.page_edit.setAlignment(QtCore.Qt.AlignCenter)
        self.page_edit.setToolTip("Escribe un número de página y pulsa Enter")
        # returnPressed = Enter en un QLineEdit; editingFinished = perdió el foco también
        self.page_edit.returnPressed.connect(self._goto_page_edit)
        self.page_edit.editingFinished.connect(self._goto_page_edit)
        self.lbl_page = QtWidgets.QLabel(" / — ")
        lp.addWidget(self.btn_prev); lp.addWidget(self.page_edit, 1); lp.addWidget(self.lbl_page); lp.addWidget(self.btn_next)

        # Fila de transparencia del PDF de fondo (para ver mejor el marcado encima)
        self.gtr = QtWidgets.QWidget(); ltr = QtWidgets.QHBoxLayout(self.gtr); ltr.setContentsMargins(0, 0, 0, 0)
        tb_l = QtWidgets.QPushButton("−"); tb_l.setFixedWidth(30); tb_l.setToolTip("Más translúcido"); tb_l.clicked.connect(lambda: self._bump_opacity(-10))
        self.lbl_opacity = QtWidgets.QLabel("100%"); self.lbl_opacity.setAlignment(QtCore.Qt.AlignCenter)
        tb_r = QtWidgets.QPushButton("+"); tb_r.setFixedWidth(30); tb_r.setToolTip("Más opaco"); tb_r.clicked.connect(lambda: self._bump_opacity(10))
        ltr.addWidget(tb_l); ltr.addWidget(self.lbl_opacity, 1); ltr.addWidget(tb_r)

        # Botones de acción (uno por sección; el color verde/azul lo pone _update_ui)
        self.btn_pipe = QtWidgets.QPushButton("✏  Dibujar utilidad"); self.btn_pipe.clicked.connect(self.toggle_pipe)
        self.btn_leader_simple = QtWidgets.QPushButton("↘  Colocar Leader"); self.btn_leader_simple.clicked.connect(lambda: self.start_leader(True))
        self.btn_text = QtWidgets.QPushButton("T  Texto libre"); self.btn_text.clicked.connect(self.toggle_text_mode)
        self.btn_erase = QtWidgets.QPushButton("▭  Borrar zona"); self.btn_erase.clicked.connect(self.toggle_erase)

        # Grupo "Tipo de utilidad" (usado al DIBUJAR una utilidad)
        # QComboBox es la lista desplegable clásica. addItem(icono, texto, dato) le
        # asocia a cada opción un "dato oculto" que recuperamos con currentData().
        self.gt = QtWidgets.QWidget(); lgt = QtWidgets.QVBoxLayout(self.gt); lgt.setContentsMargins(0, 0, 0, 0)
        self.type_combo = QtWidgets.QComboBox()
        self.type_combo.setSizePolicy(QtWidgets.QSizePolicy.Ignored, QtWidgets.QSizePolicy.Fixed)
        self.type_combo.setSizeAdjustPolicy(QtWidgets.QComboBox.AdjustToMinimumContentsLengthWithIcon)
        for label, layer in TIPOS:
            self.type_combo.addItem(swatch_icon(layer_qcolor(layer)), label, layer)
        self.type_combo.setCurrentIndex(0); self.type_combo.currentIndexChanged.connect(lambda _: self._redraw())
        self.chk_ab = QtWidgets.QCheckBox("Abandonado (línea ──/── W ──)")
        self.chk_ext_same = QtWidgets.QCheckBox("Al extender un extremo: continuar la misma utilidad")
        self.chk_ext_same.setChecked(True)
        lgt.addWidget(self.type_combo); lgt.addWidget(self.chk_ab); lgt.addWidget(self.chk_ext_same)

        # Combo de ORIENTACIÓN — COMPARTIDO por Multileader y Leader.
        # Cuando el usuario abre "Multileader" o "Leader" lo REPARENTAMOS al slot
        # de esa sección (una única instancia; su estado —H/V/D— se mantiene).
        self.orient_combo = QtWidgets.QComboBox()
        for oid, lbl in LEADER_ORIENT: self.orient_combo.addItem(lbl, oid)
        self.orient_combo.currentIndexChanged.connect(lambda _: self._update_ui())

        # Grupo "Contenido del Multileader" (checkbox + texto libre + lista Excel)
        self.ga = QtWidgets.QGroupBox("Contenido del texto"); lga = QtWidgets.QVBoxLayout(self.ga)
        self.chk_custom = QtWidgets.QCheckBox("Usar texto personalizado"); self.chk_custom.toggled.connect(self._toggle_custom)
        lga.addWidget(self.chk_custom)
        self.txt_edit = QtWidgets.QLineEdit(); self.txt_edit.setPlaceholderText("texto personalizado…"); self.txt_edit.setEnabled(False)
        lga.addWidget(self.txt_edit)
        self.lbl_textos = QtWidgets.QLabel("Textos (columna TEXTO del Excel):"); lga.addWidget(self.lbl_textos)
        # QListWidget es una lista simple de líneas de texto (una por fila).
        self.text_list = QtWidgets.QListWidget(); self.text_list.setMaximumHeight(140); lga.addWidget(self.text_list)
        self.lbl_lead_hint = QtWidgets.QLabel("<i>Colocas varios seguidos; Esc para salir.</i>"); lga.addWidget(self.lbl_lead_hint)

        # Grupo "Estilo de texto" (fuente, altura, negrita + rotación).
        # COMPARTIDO por Multileader y Texto libre. La rotación solo aplica a
        # textos libres; la mostramos/ocultamos según la sección abierta.
        self.gtxt = QtWidgets.QGroupBox("Estilo de texto"); lgx = QtWidgets.QVBoxLayout(self.gtxt)
        # QFontComboBox = combo que lista todas las fuentes instaladas en el sistema.
        self.font_combo = QtWidgets.QFontComboBox(); self.font_combo.setCurrentFont(QtGui.QFont(C.TEXT_FONT))
        self.font_combo.setSizePolicy(QtWidgets.QSizePolicy.Ignored, QtWidgets.QSizePolicy.Fixed)
        self.font_combo.currentFontChanged.connect(lambda _: self._style_changed())
        r = QtWidgets.QHBoxLayout(); r.addWidget(QtWidgets.QLabel("Altura (pies):"))
        b_minus = QtWidgets.QPushButton("−"); b_minus.setFixedWidth(30); b_minus.clicked.connect(lambda: self._bump_size(-0.5))
        # QDoubleSpinBox = campo numérico decimal con incremento por flechas (aquí ocultas).
        self.size_spin = QtWidgets.QDoubleSpinBox(); self.size_spin.setRange(0.5, 200); self.size_spin.setValue(3.0)
        self.size_spin.setSingleStep(0.5); self.size_spin.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)
        self.size_spin.valueChanged.connect(lambda _: self._style_changed())
        b_plus = QtWidgets.QPushButton("+"); b_plus.setFixedWidth(30); b_plus.clicked.connect(lambda: self._bump_size(0.5))
        r.addWidget(b_minus); r.addWidget(self.size_spin); r.addWidget(b_plus)
        self.chk_bold = QtWidgets.QCheckBox("Negrita"); self.chk_bold.toggled.connect(lambda _: self._style_changed())
        lgx.addWidget(self.font_combo); lgx.addLayout(r); lgx.addWidget(self.chk_bold)
        # Rotación (0-360°) — solo se usa para textos libres.
        self.rot_row = QtWidgets.QWidget(); rr2 = QtWidgets.QHBoxLayout(self.rot_row); rr2.setContentsMargins(0, 0, 0, 0)
        rr2.addWidget(QtWidgets.QLabel("Rotación (°):"))
        rb_l = QtWidgets.QPushButton("⟲"); rb_l.setFixedWidth(30); rb_l.clicked.connect(lambda: self._bump_rot(-1))
        self.rot_spin = QtWidgets.QSpinBox(); self.rot_spin.setRange(0, 360); self.rot_spin.setSingleStep(1); self.rot_spin.setWrapping(True)
        self.rot_spin.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons); self.rot_spin.valueChanged.connect(lambda _: self._style_changed())
        rb_r = QtWidgets.QPushButton("⟳"); rb_r.setFixedWidth(30); rb_r.clicked.connect(lambda: self._bump_rot(1))
        rr2.addWidget(rb_l); rr2.addWidget(self.rot_spin); rr2.addWidget(rb_r)
        lgx.addWidget(self.rot_row)
        lgx.addWidget(QtWidgets.QLabel("<i>Enter aplica · Ctrl+Shift+Enter salta de línea</i>"))

        # Grupo "En curso" (aparece cuando estás dibujando una utilidad o zona)
        self.gcur = QtWidgets.QGroupBox("En curso"); lc = QtWidgets.QHBoxLayout(self.gcur)
        self.btn_fin = QtWidgets.QPushButton("Finalizar (Enter)"); self.btn_fin.clicked.connect(self._on_enter)
        b_up = QtWidgets.QPushButton("Deshacer punto"); b_up.clicked.connect(self.undo)
        lc.addWidget(self.btn_fin); lc.addWidget(b_up)

        # ═══════════════════════════════════════════════════════════════════════
        # AHORA: creamos las secciones del acordeón y colocamos los widgets.
        # Los "slots" (self._slot_*) son QVBoxLayouts vacíos que quedan reservados
        # dentro de cada página para recibir los widgets compartidos por reparent.
        # ═══════════════════════════════════════════════════════════════════════

        # ── Sección: Vista y páginas ──
        p, l = _page("📄  Vista y páginas", "view")
        l.addWidget(QtWidgets.QLabel("Página:")); l.addWidget(self.gp)
        l.addWidget(QtWidgets.QLabel("Transparencia del PDF:")); l.addWidget(self.gtr)
        l.addStretch(1)

        # ── Sección: Dibujar utilidad ──
        p, l = _page("✏  Dibujar utilidad", "pipe")
        l.addWidget(self.btn_pipe)
        l.addWidget(QtWidgets.QLabel("Tipo de utilidad:"))
        l.addWidget(self.gt)
        self._slot_gcur_pipe = QtWidgets.QVBoxLayout(); l.addLayout(self._slot_gcur_pipe)   # slot: aquí va gcur al dibujar
        l.addStretch(1)

        # (La sección "Multileader" del acordeón está deshabilitada; su tab del
        # inventario también. La infraestructura de Multileader se conserva por si
        # se reactiva, pero NO se muestra al usuario.)

        # ── Sección: Leader (flecha simple) ──
        p, l = _page("↘  Leader (flecha simple)", "leader")
        l.addWidget(self.btn_leader_simple)
        l.addWidget(QtWidgets.QLabel("Orientación:"))
        self._slot_orient_ld = QtWidgets.QVBoxLayout(); l.addLayout(self._slot_orient_ld)   # slot: orient_combo
        l.addWidget(QtWidgets.QLabel("<i>El Leader es solo flecha, sin texto.</i>"))
        l.addStretch(1)

        # ── Sección: Texto libre ──
        p, l = _page("T  Texto libre", "text")
        l.addWidget(self.btn_text)
        self._slot_style_tx = QtWidgets.QVBoxLayout(); l.addLayout(self._slot_style_tx)     # slot: gtxt (estilo)
        l.addStretch(1)

        # ── Sección: Borrar zona ──
        p, l = _page("▭  Borrar zona", "erase")
        l.addWidget(self.btn_erase)
        _lbl = QtWidgets.QLabel("<i>Clic para agregar vértices, Enter cierra. "
                                     "Al exportar borra el plano dentro del polígono.</i>")
        _lbl.setWordWrap(True); l.addWidget(_lbl)
        self._slot_gcur_erase = QtWidgets.QVBoxLayout(); l.addLayout(self._slot_gcur_erase)  # slot: gcur al borrar
        l.addStretch(1)

        # (Georreferenciación y Cotas/red 3D se hacen una vez por proyecto — se
        #  mueven al menú "Herramientas" y menú "Georreferencia".)

        # Al abrir una sección del acordeón: reparenta los widgets compartidos y
        # regresa el modo a "idle" (así no quedan mezclados los estados).
        self.toolbox.currentChanged.connect(self._on_toolbox_change)

        scroll = QtWidgets.QScrollArea()
        scroll.setWidget(self.toolbox)
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        lv.addWidget(scroll, 1)
        ldock.setWidget(left); ldock.setMinimumWidth(260)
        self.addDockWidget(QtCore.Qt.LeftDockWidgetArea, ldock)

        # (La inicialización de los widgets compartidos ocurre al final de _build_ui,
        # cuando el dock derecho —self.tabs, etc.— ya existe.)

        # ── DOCK DERECHO: inventario y selección ──
        rdock = QtWidgets.QDockWidget("Inventario", self); rdock.setFeatures(QtWidgets.QDockWidget.NoDockWidgetFeatures)
        right = QtWidgets.QWidget(); rv = QtWidgets.QVBoxLayout(right)
        self.tabs = QtWidgets.QTabWidget()
        self.pipe_list = QtWidgets.QListWidget(); self.pipe_list.currentRowChanged.connect(self._sel_pipe)
        self.lead_list = QtWidgets.QListWidget(); self.lead_list.currentRowChanged.connect(self._sel_leader)
        self.sleader_list = QtWidgets.QListWidget(); self.sleader_list.currentRowChanged.connect(self._sel_sleader)
        self.txt_marks_list = QtWidgets.QListWidget(); self.txt_marks_list.currentRowChanged.connect(self._sel_text)
        self.region_list = QtWidgets.QListWidget(); self.region_list.currentRowChanged.connect(self._sel_region)
        self.region_list.itemChanged.connect(self._region_toggled)
        self.bz_list = QtWidgets.QListWidget(); self.bz_list.currentRowChanged.connect(self._sel_bz)
        self.tabs.addTab(self.pipe_list, "Utilidades"); #self.tabs.addTab(self.lead_list, "Multileaders")
        self.tabs.addTab(self.sleader_list, "Leaders")
        self.tabs.addTab(self.txt_marks_list, "Textos"); self.tabs.addTab(self.region_list, "Zonas")
        self.tabs.addTab(self.bz_list, "Buzones")
        # Menú contextual (clic derecho) en cada lista visible del inventario
        # (la lista de Multileaders no se registra porque su pestaña está oculta)
        for listw, tab_idx in ((self.pipe_list, TAB_PIPE),
                               (self.sleader_list, TAB_LEADER), (self.txt_marks_list, TAB_TEXT),
                               (self.region_list, TAB_REGION)):
            listw.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
            listw.customContextMenuRequested.connect(
                lambda pos, lw=listw, ti=tab_idx: self._list_context_menu(lw, ti, pos))
        self.tabs.currentChanged.connect(self._tab_changed); rv.addWidget(self.tabs, 1)
        # Propiedades de la utilidad seleccionada (nombre, diámetro, unidad) → XDATA en el DXF
        self.gprop = QtWidgets.QGroupBox("Propiedades de la utilidad"); fpr = QtWidgets.QFormLayout(self.gprop)
        self.prop_name = QtWidgets.QLineEdit(); self.prop_name.editingFinished.connect(self._prop_changed)
        # El diámetro va SIEMPRE en PULGADAS y SOLO de la lista estándar del
        # catálogo (12,15,18,…): un desplegable NO editable, sin valores libres,
        # para que coincida 1:1 con un tamaño real del catálogo de Civil 3D.
        # Es independiente de la unidad de trabajo (que rige coordenadas/cotas).
        # El diámetro ya no es un campo del UI: se deriva automáticamente del
        # "Tamaño (catálogo)" elegido. p["diam"] se calcula al guardar propiedades.
        fpr.addRow("Nombre:", self.prop_name)
        # Campos de la utilidad usados por el JSON de red 3.0 y por el DXF:
        #   - material: texto libre (p.ej. "HDPE"); viaja al JSON como `material`.
        #   - part (pieza): nombre del tipo de pieza; viaja al JSON como `part`.
        #   - network_type: "auto" (=lo decide la capa) / "pipe" (con buzones)
        #     / "pressure" (línea a presión). Viaja al JSON como `network_type`.
        #   - invert inicio/fin (m): las COTAS de fondo de la tubería en cada
        #     extremo; imprescindibles para reconstruir la red 3D.
        # Material: desplegable con los valores exactos de Civil 3D (no texto libre).
        self.prop_material = QtWidgets.QComboBox()
        for m in PIPE_MATERIALS:
            self.prop_material.addItem(m)
        self.prop_material.currentIndexChanged.connect(lambda _: self._prop_changed())
        self.prop_part = QtWidgets.QLineEdit(); self.prop_part.setPlaceholderText("p.ej. 900 mm Corrugated HDPE Pipe")
        self.prop_part.editingFinished.connect(self._prop_changed)
        self.prop_nettype = QtWidgets.QComboBox()
        self.prop_nettype.addItem("Automático (según la capa)", "")
        self.prop_nettype.addItem("Con buzones (pipe)", "pipe")
        self.prop_nettype.addItem("A presión (pressure)", "pressure")
        self.prop_nettype.currentIndexChanged.connect(lambda _: self._prop_changed())
        self.prop_inv0 = QtWidgets.QDoubleSpinBox(); self.prop_inv1 = QtWidgets.QDoubleSpinBox()
        for sp in (self.prop_inv0, self.prop_inv1):
            sp.setRange(-100000, 100000); sp.setDecimals(3); sp.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)
            sp.valueChanged.connect(lambda _: self._prop_changed())
        fpr.addRow("Material:", self.prop_material)
        fpr.addRow("Tipo de red:", self.prop_nettype)
        # Labels dinámicas: se recomponen al cambiar la unidad de trabajo.
        # Nombre igual a Civil 3D: "Elevación de rasante" (no "Invert").
        self.lbl_prop_inv0 = QtWidgets.QLabel("Elev. de rasante inicial (ft):"); fpr.addRow(self.lbl_prop_inv0, self.prop_inv0)
        self.lbl_prop_inv1 = QtWidgets.QLabel("Elev. de rasante final (ft):");   fpr.addRow(self.lbl_prop_inv1, self.prop_inv1)
        fpr.addRow("Part (pieza):", self.prop_part)
        # Familia + tamaño del catálogo Civil 3D para esta pipe (solo gravedad).
        # Para presión y conduit no aplica: presión usa el sub-catálogo por material
        # y conduit se deja como polyline simple.
        self.prop_family = QtWidgets.QComboBox()
        self.prop_family.currentIndexChanged.connect(self._pipe_family_changed)
        self.prop_size = QtWidgets.QComboBox()
        self.prop_size.currentIndexChanged.connect(lambda _: self._prop_changed())
        self.lbl_prop_family = QtWidgets.QLabel("Familia (catálogo):")
        self.lbl_prop_size = QtWidgets.QLabel("Tamaño (catálogo):")
        fpr.addRow(self.lbl_prop_family, self.prop_family)
        fpr.addRow(self.lbl_prop_size, self.prop_size)

        # Lista de vértices intermedios con checkbox "sin buzón aquí". Al marcarlo,
        # el addin no insertará una structure en ese vértice (útil para quiebres
        # donde el conducto sigue de largo sin manhole). Los extremos (primer y
        # último vértice) SIEMPRE llevan structure y no se muestran aquí.
        self.lbl_prop_noman = QtWidgets.QLabel("Vértices intermedios sin buzón:")
        self.prop_noman = QtWidgets.QListWidget()
        self.prop_noman.setMaximumHeight(110)
        self.prop_noman.itemChanged.connect(self._pipe_noman_changed)
        fpr.addRow(self.lbl_prop_noman, self.prop_noman)

        # Override de familia/tamaño POR SEGMENTO (opcional). Por defecto todos los
        # tramos usan la familia y el tamaño globales de arriba; este botón abre un
        # diálogo para variar por tramo (útil para reducciones, cambios de material,
        # o marcar un segmento como distinto sin dividir la utilidad).
        self.btn_seg_override = QtWidgets.QPushButton("Editar por segmento…")
        self.btn_seg_override.clicked.connect(self._open_seg_override_dialog)
        self.lbl_seg_override = QtWidgets.QLabel("Variación por tramo:")
        fpr.addRow(self.lbl_seg_override, self.btn_seg_override)

        rv.addWidget(self.gprop)
        # ── Propiedades del buzón seleccionado (tab Buzones) ───────────────────
        self.gprop_bz = QtWidgets.QGroupBox("Propiedades del buzón"); fbz = QtWidgets.QFormLayout(self.gprop_bz)
        self.bz_cod = QtWidgets.QLineEdit(); self.bz_cod.editingFinished.connect(self._bz_prop_changed)
        self.bz_rim = QtWidgets.QDoubleSpinBox(); self.bz_sump = QtWidgets.QDoubleSpinBox()
        for sp in (self.bz_rim, self.bz_sump):
            sp.setRange(-100000, 100000); sp.setDecimals(3); sp.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)
            sp.valueChanged.connect(lambda _v: self._bz_prop_changed())
        self.bz_family = QtWidgets.QComboBox(); self.bz_family.currentIndexChanged.connect(self._bz_family_changed)
        self.bz_size = QtWidgets.QComboBox(); self.bz_size.currentIndexChanged.connect(self._bz_prop_changed)
        self.bz_cover = QtWidgets.QComboBox(); self.bz_cover.addItems(["Sí", "No"])
        self.bz_cover.currentIndexChanged.connect(self._bz_prop_changed)
        self.bz_net_lbl = QtWidgets.QLabel("—")
        self.bz_origin_lbl = QtWidgets.QLabel("—")
        fbz.addRow("Código:", self.bz_cod)
        self.bz_rim_lbl = QtWidgets.QLabel(f"Rim ({self.work_unit}):"); fbz.addRow(self.bz_rim_lbl, self.bz_rim)
        self.bz_sump_lbl = QtWidgets.QLabel(f"Sump ({self.work_unit}):"); fbz.addRow(self.bz_sump_lbl, self.bz_sump)
        fbz.addRow("Familia:", self.bz_family)
        fbz.addRow("Tamaño:", self.bz_size)
        fbz.addRow("Tapa:", self.bz_cover)
        fbz.addRow("Red:", self.bz_net_lbl)
        fbz.addRow("Origen:", self.bz_origin_lbl)
        self.btn_bz_addsize = QtWidgets.QPushButton("➕ Agregar tamaño personalizado…")
        self.btn_bz_addsize.clicked.connect(self._bz_add_custom_size_current)
        fbz.addRow("", self.btn_bz_addsize)
        # Checkbox de etiquetas — entre la lista de buzones (tab) y el panel de propiedades.
        self.chk_bz_labels = QtWidgets.QCheckBox(
            "Mostrar etiquetas (código) al lado del buzón en el lienzo y en el DXF exportado")
        self.chk_bz_labels.setChecked(bool(self.show_bz_labels))
        def _toggle_bz_labels(v):
            self.show_bz_labels = bool(v); self._redraw()
        self.chk_bz_labels.toggled.connect(_toggle_bz_labels)
        rv.addWidget(self.chk_bz_labels)
        rv.addWidget(self.gprop_bz)
        # Mensaje guía cuando estás en la tab Buzones pero no seleccionaste nada.
        self.lbl_bz_hint = QtWidgets.QLabel(
            "Haz clic en un buzón de la lista (o en su círculo en el lienzo) para ver y editar sus propiedades.")
        self.lbl_bz_hint.setWordWrap(True)
        self.lbl_bz_hint.setStyleSheet("color:#f0d060; padding:8px; background:#333a4a; border-radius:4px;")
        rv.addWidget(self.lbl_bz_hint)
        self.gprop_bz.setVisible(False); self.lbl_bz_hint.setVisible(False)
        self._bz_prop_guard = False                 # evita reentradas al setear valores desde el modelo
        rr = QtWidgets.QGridLayout()
        self.btn_ct = QtWidgets.QPushButton("Cambiar tipo"); self.btn_ct.clicked.connect(self.change_pipe_type)
        self.btn_mv = QtWidgets.QPushButton("Editar/mover"); self.btn_mv.clicked.connect(self.enter_move)
        self.btn_edit = QtWidgets.QPushButton("Editar texto"); self.btn_edit.clicked.connect(self.edit_selected_text)
        self.btn_del = QtWidgets.QPushButton("Eliminar"); self.btn_del.clicked.connect(self.delete_selected)
        rr.addWidget(self.btn_ct, 0, 0); rr.addWidget(self.btn_mv, 0, 1)
        rr.addWidget(self.btn_edit, 1, 0); rr.addWidget(self.btn_del, 1, 1)
        rv.addLayout(rr)
        rdock.setWidget(right); rdock.setMinimumWidth(250)
        self.addDockWidget(QtCore.Qt.RightDockWidgetArea, rdock)

        # ── Barra de estado: modo · info · contadores en vivo · escala · georref ──
        self.status = self.statusBar(); self.status.setSizeGripEnabled(False)
        self.lbl_mode = QtWidgets.QLabel("Modo: inactivo"); self.lbl_mode.setStyleSheet("color:#adc6ff;")
        self.status.addWidget(self.lbl_mode)
        self.status.addWidget(QtWidgets.QLabel("│"))
        self.lbl_info = QtWidgets.QLabel(""); self.lbl_info.setStyleSheet("color:#8c909f;"); self.status.addWidget(self.lbl_info, 1)
        self.lbl_snap = QtWidgets.QLabel("Imán: OFF"); self.lbl_coords = QtWidgets.QLabel("X —  Y —")
        # Contadores en vivo: N utilidades · N leaders · N textos · dirty
        self.lbl_counts = QtWidgets.QLabel("—")
        self.lbl_dirty = QtWidgets.QLabel("")   # muestra "●" cuando hay cambios sin guardar
        self.lbl_scale = QtWidgets.QLabel("Escala —"); self.lbl_geo = QtWidgets.QLabel("Georref: no")
        for w in (self.lbl_snap, self.lbl_coords, self.lbl_counts, self.lbl_dirty, self.lbl_scale, self.lbl_geo):
            w.setStyleSheet("color:#c2c6d6;"); self.status.addPermanentWidget(w)
        self.canvas.moved.connect(self._update_coords)
        self._update_geo_status()
        self._info("Abre o arrastra un PDF/proyecto.")

        # Todo listo: coloca los widgets compartidos del acordeón en la sección
        # inicial (esto necesita que self.tabs, self.gprop y self.lbl_mode existan).
        self._on_toolbox_change(self.toolbox.currentIndex())
        self._refresh_unit_labels()                # etiquetas de campo con la unidad activa

    def _menu_act(self, menu, text, fn, sc=None):
        a = QtGui.QAction(text, self); a.triggered.connect(fn)
        if sc: a.setShortcut(sc)
        menu.addAction(a); return a

    def _apply_style(self):
        self.setStyleSheet("""
            QMainWindow,QDockWidget,QGroupBox,QScrollArea,QMenuBar,QMenu{background:#2b2b2b;color:#eee;}
            QMenuBar::item:selected,QMenu::item:selected{background:#3c5a99;}
            QGroupBox{border:1px solid #444;margin-top:8px;padding:6px;border-radius:4px;font-weight:bold;}
            QGroupBox::title{subcontrol-origin:margin;left:8px;color:#bbb;}
            QListWidget,QLineEdit,QComboBox,QSpinBox,QDoubleSpinBox,QFontComboBox,QTabWidget::pane{background:#333;color:#eee;border:1px solid #555;}
            QTabBar::tab{background:#333;color:#ccc;padding:5px;} QTabBar::tab:selected{background:#3c5a99;color:white;}
            QPushButton{background:#3c5a99;color:white;border:none;padding:6px;border-radius:4px;}
            QPushButton:hover{background:#4a6fbf;} QLabel,QCheckBox{color:#ddd;font-weight:normal;}
            QSplitter::handle{background:#555;height:6px;}""")
        self.chk_snap.setChecked(False)

    def _shortcuts(self):
        QtGui.QShortcut(QtGui.QKeySequence("Ctrl+T"), self, self.enter_move)
        for k in (QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter):
            QtGui.QShortcut(QtGui.QKeySequence(k), self, self._on_enter)
        QtGui.QShortcut(QtGui.QKeySequence(QtCore.Qt.Key_Escape), self, self._on_escape)
        QtGui.QShortcut(QtGui.QKeySequence("Ctrl+C"), self, self._copy_sel)
        QtGui.QShortcut(QtGui.QKeySequence("Ctrl+V"), self, self._paste_sel)
        QtGui.QShortcut(QtGui.QKeySequence(QtCore.Qt.Key_Delete), self, self._delete_shortcut)

    def _delete_shortcut(self):
        # Suprimir borra el elemento seleccionado, salvo mientras se escribe texto
        if self._editor is not None: return
        fw = QtWidgets.QApplication.focusWidget()
        if isinstance(fw, (QtWidgets.QLineEdit, QtWidgets.QTextEdit, QtWidgets.QAbstractSpinBox)): return
        self.delete_selected()

    # ─────────────────────────── páginas ───────────────────────────
    def _update_page_label(self):
        if self.doc:
            self.page_edit.setText(str(self.page_idx + 1)); self.lbl_page.setText(f" / {self.doc.page_count} ")
            self.page_edit.setEnabled(True)
            self.btn_prev.setEnabled(self.page_idx > 0)
            self.btn_next.setEnabled(self.page_idx < self.doc.page_count - 1)
        else:
            self.page_edit.setText(""); self.page_edit.setEnabled(False); self.lbl_page.setText(" / — ")
            self.btn_prev.setEnabled(False); self.btn_next.setEnabled(False)

    def _goto_page_edit(self):
        if not self.doc: return
        try: n = int(self.page_edit.text()) - 1
        except ValueError: self._update_page_label(); return
        n = max(0, min(n, self.doc.page_count - 1))
        if n != self.page_idx:
            if not self._confirm_discard(): self._update_page_label(); return
            self.page_idx = n; self._load_page(n)
        else:
            self._update_page_label()

    def _prev_page(self):
        if self.doc and self.page_idx > 0:
            if not self._confirm_discard(): return
            self.page_idx -= 1; self._load_page(self.page_idx)

    def _next_page(self):
        if self.doc and self.page_idx < self.doc.page_count - 1:
            if not self._confirm_discard(): return
            self.page_idx += 1; self._load_page(self.page_idx)

    # ─────────────────────────── Estado / modos ───────────────────────────
    def _on_enter(self):
        if self.mode == "pipe": self.finish_pipe()
        elif self.mode == "erase": self.finish_erase()
        elif self.mode in ("leader1", "leader2", "leader3"):
            self.set_mode("idle"); self._info("Comando Leader finalizado (Enter)")
        elif self.mode == "move":
            self.set_mode("idle"); self._info("Edición terminada (Enter)")

    def _on_escape(self):
        if self.mode == "pipe" and self.cur_pts:
            self._push(); self.cur_pts = []; self._extending = False; self._ext_pipe = None; self._ext_at = None
            self._update_ui(); self._redraw(); self._info("Puntos cancelados")
        elif self.mode == "erase" and self._erase_pts:
            self._erase_pts = []; self._redraw(); self._info("Zona cancelada")
        elif self.mode == "insert_bz":
            self.set_mode("idle"); self._info("Inserción de buzón cancelada")
        elif self.sel_pipe >= 0 or self.sel_leader >= 0 or self.sel_region >= 0 or self.sel_text >= 0:
            self._deselect_all(); self._info("Selección quitada")
        else:
            self.set_mode("idle"); self._info("Salió del modo")

    def _deselect_all(self):
        self.sel_pipe = self.sel_leader = self.sel_region = self.sel_text = self.sel_bz = -1
        for lst in (self.pipe_list, self.lead_list, self.sleader_list, self.txt_marks_list, self.region_list,
                    getattr(self, "bz_list", None)):
            if lst is None: continue
            lst.blockSignals(True); lst.setCurrentRow(-1); lst.clearSelection(); lst.blockSignals(False)
        if self.mode == "move": self.set_mode("idle")
        self._update_ui(); self._redraw()

    def _info(self, m): self.lbl_info.setText(m)

    def _zoom_in(self): self.canvas.scale(1.25, 1.25)
    def _zoom_out(self): self.canvas.scale(0.8, 0.8)

    def _toggle_snap(self, v):
        self.snap = v; self.lbl_snap.setText(f"Imán: {'ON' if v else 'OFF'}")
        self.lbl_snap.setStyleSheet("color:%s;" % ("#5fd35f" if v else "#c2c6d6"))

    def _update_coords(self, x, y):
        if self.canvas.pixmap_item is None: return
        cx, cy = self._to_cad(x, y); self.lbl_coords.setText(f"X {cx:,.1f}  Y {cy:,.1f}")

    def _update_geo_status(self):
        if self.georef.active():
            unit = georef_mod.epsg_unit(self.georef.epsg)
            rms = f" · RMS {self.georef.rms:.2f} {unit}" if self.georef.rms is not None else ""
            self.lbl_geo.setText(f"Georref: EPSG:{self.georef.epsg}{rms}")
            self.lbl_geo.setStyleSheet("color:#5fd35f;")
        else:
            self.lbl_geo.setText("Georref: no (escala titleblock)")
            self.lbl_geo.setStyleSheet("color:#e0c060;")

    def _update_ui(self):
        m = self.mode
        def st(btn, on): btn.setStyleSheet(BTN_ON if on else BTN_OFF)
        in_leader = m in ("leader1", "leader2", "leader3")
        st(self.btn_pipe, m == "pipe")
        st(self.btn_leader_simple, in_leader)
        st(self.btn_text, m == "text"); st(self.btn_erase, m == "erase")
        self.btn_pipe.setText("■  Salir de dibujar utilidad" if m == "pipe" else "✏  Dibujar utilidad")
        self.btn_leader_simple.setText("●  Coloque Leader…" if in_leader else "↘  Colocar Leader")
        self.btn_erase.setText("■  Terminar zona (Enter)" if m == "erase" else "▭  Borrar zona (polígono)")
        ti = self._current_tab()
        self.gtxt.setTitle("Estilo de texto")
        self.gprop.setVisible(ti == TAB_PIPE and self.sel_pipe >= 0)
        # Panel de propiedades del buzón: visible en tab Buzones (aunque sin selección
        # se muestra el groupbox con campos deshabilitados para que el user vea que existe).
        if hasattr(self, "gprop_bz"):
            self.gprop_bz.setVisible(ti == TAB_BZ)
        if hasattr(self, "chk_bz_labels"):
            self.chk_bz_labels.setVisible(ti == TAB_BZ)
        # "En curso": solo mientras hay puntos en curso. gcur ya fue movido a la
        # sección correcta por set_mode; aquí solo habilitamos Finalizar y mostramos.
        active_draw = (m == "pipe" and len(self.cur_pts) >= 1) or (m == "erase" and len(self._erase_pts) >= 1)
        self.gcur.setVisible(active_draw)
        self.btn_fin.setEnabled((m == "pipe" and len(self.cur_pts) >= 2) or (m == "erase" and len(self._erase_pts) >= 3))
        ti = self._current_tab()
        self.btn_ct.setVisible(ti == TAB_PIPE)
        self.btn_mv.setVisible(ti in (TAB_PIPE, TAB_LEADER, TAB_TEXT, TAB_REGION))
        self.btn_mv.setText("Mover" if ti == TAB_TEXT else "Editar/mover")
        self.btn_edit.setVisible(ti in (TAB_ML, TAB_TEXT))
        diag = self.orient_combo.currentData() == "d"
        lead1 = "Modo: Leader — clic en la cabeza de flecha (dónde señala)"
        lead2 = ("Modo: Leader — clic en el inicio del landing (bisagra)" if diag
                 else "Modo: Leader — clic en el final del cuerpo")
        lead3 = "Modo: Leader — clic en el final del cuerpo"
        self.lbl_mode.setText({"idle": "Modo: inactivo  ·  clic en el dibujo para seleccionar",
                               "pipe": ("Modo: EXTENDIENDO desde el vértice — clic agrega puntos, Enter finaliza"
                                        if self._extending else "Modo: dibujar utilidad  ·  Enter finaliza"),
                               "leader1": lead1,
                               "leader2": lead2,
                               "leader3": lead3,
                               "text": "Modo: texto libre — clic donde escribir · Enter aplica",
                               "erase": "Modo: borrar zona — clic para el polígono, Enter cierra",
                               "move": "Modo: editar — arrastra vértice · clic en tramo inserta · clic-en-vértice extiende (F) · clic derecho elimina"}.get(m, ""))

    def set_mode(self, m):
        """Cambia el "modo" del programa (qué está haciendo ahora el usuario):
        pipe (dibujando utilidad), leader1/2/3 (colocando Multileader/Leader),
        text (escribiendo texto libre), erase (borrando zona), move (editando),
        idle (sin nada activo).

        Además REPARENTA el grupo "En curso" (self.gcur) a la sección adecuada
        del acordeón según el modo, para que los botones Finalizar/Deshacer
        aparezcan DENTRO de la sección donde estás trabajando."""
        if m not in ("leader1", "leader2", "leader3"): self._pending = None
        if m != "erase": self._erase_pts = []
        if m != "pipe": self._extending = False
        self.mode = m
        # Colocar el panel "En curso" en su sección correspondiente (si existe)
        if hasattr(self, "_slot_gcur_pipe"):
            if m == "pipe":
                self._place_widget(self.gcur, self._slot_gcur_pipe)
                self.gcur.show()
            elif m == "erase":
                self._place_widget(self.gcur, self._slot_gcur_erase)
                self.gcur.show()
            else:
                self.gcur.hide()
        self._update_ui(); self._redraw()

    # El inventario usa IDs lógicos (TAB_PIPE, TAB_ML, …) que NO coinciden con la
    # posición visible de la pestaña cuando alguna está oculta (p. ej. Multileaders).
    # Estos helpers traducen widget↔constante, así el código es robusto ante
    # pestañas ocultas o reordenadas.
    def _tab_map(self):
        m = {self.pipe_list: TAB_PIPE, self.lead_list: TAB_ML, self.sleader_list: TAB_LEADER,
             self.txt_marks_list: TAB_TEXT, self.region_list: TAB_REGION}
        if hasattr(self, "bz_list"): m[self.bz_list] = TAB_BZ
        return m

    def _current_tab(self):
        return self._tab_map().get(self.tabs.currentWidget(), TAB_PIPE)

    def _show_tab(self, logical):
        w = {v: k for k, v in self._tab_map().items()}.get(logical)
        idx = self.tabs.indexOf(w) if w is not None else -1
        if idx >= 0:
            self.tabs.setCurrentIndex(idx)

    def _tab_changed(self, _):
        ti = self._current_tab()                    # sel_leader se comparte entre ML y Leaders: re-sincronizar
        if ti == TAB_ML: self.sel_leader = self._leader_at_row(self.lead_list, self.lead_list.currentRow())
        elif ti == TAB_LEADER: self.sel_leader = self._leader_at_row(self.sleader_list, self.sleader_list.currentRow())
        if self.mode == "move": self.set_mode("idle")   # no seguir editando al cambiar de pestaña
        if ti == TAB_BZ: self._sync_bz_panel()
        self._update_ui(); self._redraw()
    # Los "toggle_*" alternan entre "modo activo" e "idle" (sin nada activo).
    # Además abren su sección del acordeón para que las opciones sean visibles.
    def toggle_pipe(self):
        if self.mode == "pipe": self.set_mode("idle")
        else: self._open_section("pipe"); self.set_mode("pipe")
    def toggle_text_mode(self):
        if self.mode == "text": self.set_mode("idle")
        else: self._open_section("text"); self.set_mode("text")
    def toggle_erase(self):
        if self.mode == "erase": self.set_mode("idle")
        else: self._open_section("erase"); self.set_mode("erase")
    def active_layer(self):
        d = self.type_combo.currentData(); return d if d else "AGUA"

    def _toggle_custom(self, on):
        self.txt_edit.setEnabled(on); self.text_list.setEnabled(not on)
        if on:
            self.text_list.blockSignals(True); self.text_list.setCurrentRow(-1); self.text_list.clearSelection(); self.text_list.blockSignals(False)
        else:
            self.txt_edit.clear()

    # ─────────────────────────── undo/redo ───────────────────────────
    def _snap_state(self):
        return copy.deepcopy(dict(cur_pts=self.cur_pts, pipes=self.pipes, leaders=self.leaders,
                                  text_marks=self.text_marks, erase_regions=self.erase_regions,
                                  structures=self.structures))

    def _push(self):
        self._undo.append(self._snap_state()); self._redo.clear(); self._dirty = True
        if len(self._undo) > 400: self._undo.pop(0)

    def _restore(self, s):
        self.cur_pts, self.pipes = s["cur_pts"], s["pipes"]
        self.leaders, self.text_marks = s["leaders"], s["text_marks"]
        self.erase_regions = s.get("erase_regions", [])
        self.structures = s.get("structures", [])
        self._refresh_lists(); self._update_ui(); self._redraw()

    def undo(self):
        if self._undo: self._redo.append(self._snap_state()); self._restore(self._undo.pop()); self._info("Deshacer")

    def redo(self):
        if self._redo: self._undo.append(self._snap_state()); self._restore(self._redo.pop()); self._info("Rehacer")

    # ─────────────────────────── abrir ───────────────────────────
    def _busy(self, m="Procesando…"):
        QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.WaitCursor); self._info(m); QtWidgets.QApplication.processEvents()
    def _unbusy(self): QtWidgets.QApplication.restoreOverrideCursor()

    def open_path(self, path):
        low = path.lower()
        if low.endswith(".pdf"): self._open_pdf_path(path)
        elif low.endswith(".digproj"): self._open_project_path(path)
        elif low.endswith((".xlsx", ".xlsm")): self._read_excel(path)

    def open_pdf(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Abrir PDF", DOWNLOADS, "PDF (*.pdf)")
        if path: self._open_pdf_path(path)

    def _open_pdf_path(self, path):
        if not self._confirm_discard(): return
        self._busy("Abriendo PDF…")
        try:
            self.pdf_path = path; self.project_path = None; self.doc = fitz.open(path)
            self.page_idx = 0; self._load_page(0)
        finally: self._unbusy()

    def _load_page(self, idx):
        self._close_editor()
        page = self.doc[idx]; self.page_idx = idx; self.scale = VP.detect_scale(page)
        self.rot = page.rotation; mbx = page.mediabox; self.W, self.H = mbx.width, mbx.height
        self.derot = page.derotation_matrix
        pix = page.get_pixmap(matrix=fitz.Matrix(self.zoom, self.zoom), alpha=False)
        self.pageH_px = pix.height
        # tamaño de marca acotado: evita textos/Multileaders gigantes por escala mal detectada
        self.leader_hpx = max(14.0, min(LEADER_TEXT_FT / self.scale * self.zoom, self.pageH_px * 0.05))
        buf = bytes(pix.samples)
        qimg = QtGui.QImage(buf, pix.width, pix.height, pix.stride, QtGui.QImage.Format_RGB888).copy()
        arr = np.frombuffer(buf, np.uint8).reshape(pix.height, pix.stride)[:, :pix.width * 3].reshape(pix.height, pix.width, 3)
        self.gray = (0.299 * arr[:, :, 0] + 0.587 * arr[:, :, 1] + 0.114 * arr[:, :, 2]).astype(np.uint8)
        self._overlay = []
        self.canvas.set_image(qimg)
        self._reset_model(); self._update_page_label()
        self.lbl_scale.setText(f"Escala 1\"={self.scale*72:.0f}'")
        self._info(f"Página {idx + 1} cargada.")

    def _reset_model(self):
        self.cur_pts = []; self.pipes = []; self.leaders = []; self.text_marks = []
        self.erase_regions = []; self._erase_pts = []; self.structures = []
        self.sel_pipe = self.sel_leader = self.sel_region = self.sel_text = -1
        self._overlay = []; self._close_editor(); self._dirty = False; self._extending = False
        self.georef = georef_mod.Georef()          # cada página/PDF nuevo empieza sin georreferencia
        self._undo.clear(); self._redo.clear()
        self.set_mode("idle"); self._refresh_lists(); self._redraw(); self._update_geo_status()

    # ─────────────────────────── proyecto ───────────────────────────
    def _write_project(self, path):
        self._busy("Guardando proyecto…")
        try:
            model = dict(pipes=self.pipes, leaders=self.leaders, text_marks=self.text_marks,
                         erase_regions=self.erase_regions, structures=self.structures,
                         georef=self.georef.to_dict(),
                         work_unit=self.work_unit,                # unidad de trabajo del proyecto
                         tf=dict(scale=self.scale, zoom=self.zoom, rot=self.rot, W=self.W, H=self.H,
                                 derot=[self.derot.a, self.derot.b, self.derot.c, self.derot.d, self.derot.e, self.derot.f]),
                         pdf_name=os.path.basename(self.pdf_path or ""), version=VERSION)
            ba = QtCore.QByteArray(); buf = QtCore.QBuffer(ba); buf.open(QtCore.QIODevice.WriteOnly)
            self.canvas.pixmap_item.pixmap().save(buf, "PNG")
            with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
                z.writestr("model.json", json.dumps(model)); z.writestr("page.png", bytes(ba))
                pdf_bytes = self._get_pdf_bytes()
                if pdf_bytes:
                    z.writestr("source.pdf", pdf_bytes)
            self.project_path = path; self._dirty = False; self._info(f"Proyecto guardado: {os.path.basename(path)}")
        finally: self._unbusy()

    def _get_pdf_bytes(self):
        if self.doc:
            try: return self.doc.tobytes(deflate=True)
            except Exception: pass
        if self.pdf_path and os.path.isfile(self.pdf_path):
            try:
                with open(self.pdf_path, "rb") as f: return f.read()
            except Exception: pass
        return None

    def _cleanup_tmp_pdf(self):
        tmp = getattr(self, '_tmp_pdf', None)
        if tmp and os.path.isfile(tmp):
            try: os.remove(tmp)
            except Exception: pass
        self._tmp_pdf = None

    def save_project(self):
        if self.canvas.pixmap_item is None:
            QtWidgets.QMessageBox.information(self, "Nada que guardar", "Abre un PDF o proyecto primero."); return
        if self.project_path: self._write_project(self.project_path)
        else: self.save_project_as()

    def save_project_as(self):
        if self.canvas.pixmap_item is None:
            QtWidgets.QMessageBox.information(self, "Nada que guardar", "Abre un PDF o proyecto primero."); return
        base = self.project_path or os.path.join(DOWNLOADS, "proyecto.digproj")
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Guardar proyecto como", base, "Proyecto (*.digproj)")
        if path: self._write_project(path)

    def open_project(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Abrir proyecto", DOWNLOADS, "Proyecto (*.digproj)")
        if path: self._open_project_path(path)

    def _open_project_path(self, path):
        if not self._confirm_discard(): return
        self._busy("Abriendo proyecto…")
        try:
            self._cleanup_tmp_pdf()
            with zipfile.ZipFile(path) as z:
                model = json.loads(z.read("model.json")); png = z.read("page.png")
                if "source.pdf" in z.namelist():
                    tmp_pdf = path + ".src.pdf"
                    with open(tmp_pdf, "wb") as fp: fp.write(z.read("source.pdf"))
                    self._tmp_pdf = tmp_pdf
                else:
                    tmp_pdf = None
            qimg = QtGui.QImage.fromData(png, "PNG")
            self._overlay = []; self._close_editor()
            self.canvas.set_image(qimg); self.gray = qimage_to_gray(qimg)
            tf = model["tf"]; self.scale = tf["scale"]; self.zoom = tf["zoom"]; self.rot = tf["rot"]
            self.W, self.H = tf["W"], tf["H"]; self.derot = fitz.Matrix(*tf["derot"])
            self.pageH_px = qimg.height()
            self.leader_hpx = max(14.0, min(LEADER_TEXT_FT / self.scale * self.zoom, self.pageH_px * 0.05))
            if tmp_pdf:
                self.pdf_path = tmp_pdf; self.doc = fitz.open(tmp_pdf)
            else:
                self.pdf_path = None; self.doc = None
            self.project_path = path; self.page_idx = 0
            self.pipes = model.get("pipes", []); self.leaders = model.get("leaders", [])
            self.text_marks = model.get("text_marks", [])
            self.erase_regions = [r if isinstance(r, dict) else {"pts": r, "enabled": True}
                                  for r in model.get("erase_regions", [])]
            self.structures = model.get("structures", [])   # retrocompat: proyectos viejos sin buzones
            self.georef = georef_mod.Georef.from_dict(model.get("georef"))   # retrocompat: sin georref → escala
            # Unidad de trabajo SIEMPRE pies (los diámetros van en pulgadas, por campo).
            # Ya no hay selector; ignoramos el work_unit guardado en proyectos viejos.
            self.work_unit = "ft"
            for p in self.pipes: p["unit"] = "ft"
            self.cur_pts = []; self._erase_pts = []; self.sel_pipe = self.sel_leader = self.sel_region = self.sel_text = -1
            self._undo.clear(); self._redo.clear(); self._dirty = False
            self.set_mode("idle"); self._refresh_lists(); self._update_page_label(); self._redraw()
            self.lbl_scale.setText(f"Escala 1\"={self.scale*72:.0f}'"); self._update_geo_status()
            self._refresh_unit_labels()
            self._info(f"Proyecto abierto ({len(self.pipes)} utilidades). Ctrl+S guarda en este mismo archivo.")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", str(e))
        finally: self._unbusy()

    def _confirm_discard(self):
        if not self._dirty or self.canvas.pixmap_item is None: return True
        r = QtWidgets.QMessageBox.question(
            self, "Cambios sin guardar", "Hay cambios sin guardar. ¿Deseas guardarlos?",
            QtWidgets.QMessageBox.Save | QtWidgets.QMessageBox.Discard | QtWidgets.QMessageBox.Cancel)
        if r == QtWidgets.QMessageBox.Cancel: return False
        if r == QtWidgets.QMessageBox.Save:
            self.save_project(); return not self._dirty
        return True

    def close_project(self):
        if self.canvas.pixmap_item is None: return
        if not self._confirm_discard(): return
        self._cleanup_tmp_pdf()
        self.canvas.scene().clear(); self.canvas.pixmap_item = None
        self.pdf_path = None; self.doc = None; self.project_path = None; self.gray = None
        self.pipes = []; self.leaders = []; self.text_marks = []; self.erase_regions = []; self.structures = []
        self.cur_pts = []; self._erase_pts = []; self._overlay = []; self._close_editor()
        self.sel_pipe = self.sel_leader = self.sel_region = self.sel_text = -1
        self._undo.clear(); self._redo.clear(); self._dirty = False
        self.georef = georef_mod.Georef()
        self.set_mode("idle"); self._refresh_lists(); self._update_page_label(); self._info("Proyecto cerrado.")

    def closeEvent(self, e):
        if self._confirm_discard():
            self._cleanup_tmp_pdf(); e.accept()
        else: e.ignore()

    # ─────────────────────────── clics ───────────────────────────
    def on_click(self, x, y, button):
        if self.canvas.pixmap_item is None: return
        if button == QtCore.Qt.RightButton:
            if self.mode == "pipe": self.finish_pipe()
            elif self.mode == "erase": self.finish_erase()
            elif self.mode == "move": self._delete_vertex(x, y)
            return
        if self.mode == "pipe":
            self._push(); self.cur_pts.append(self._snap(x, y)); self._update_ui(); self._redraw()
        elif self.mode == "erase":
            self._erase_pts.append((x, y)); self._update_ui(); self._redraw()
        elif self.mode == "text":
            if self._editor is not None: return   # el clic confirma el texto abierto (no abre otro)
            self._new_free_text(x, y)
        elif self.mode == "leader1":
            self._pending["arrow"] = self._snap(x, y); self.mode = "leader2"; self._update_ui()
        elif self.mode == "leader2":
            hx, hy = self._pending["arrow"]; o = self.orient_combo.currentData()
            if o == "d":                                 # diagonal: 2º clic = inicio del landing (bisagra)
                self._pending["landing"] = self._snap(x, y); self.mode = "leader3"; self._update_ui(); return
            tail = (x, hy) if o == "h" else (hx, y)      # h/v: 2º clic = final del cuerpo, recto al eje
            self._add_simple_leader((hx, hy), tail, o); return
        elif self.mode == "leader3":                         # Leader diagonal: 3er clic = final del cuerpo
            self._add_simple_leader(self._pending["arrow"], (x, y), "d",
                                    landing=self._pending.get("landing")); return
        elif self.mode == "insert_bz":
            self._do_insert_manhole(x, y)
        elif self.mode == "idle":
            self._pick(x, y)

    def _add_simple_leader(self, head, tail, orient, landing=None):
        """Coloca un Leader simple (solo flecha, sin texto) y queda listo para el siguiente.
        orient 'h'/'v' → 2 clics (cabeza→final del cuerpo, recto al eje).
        orient 'd'     → 3 clics (cabeza → inicio del landing/bisagra → final del cuerpo)."""
        self._push()
        self.leaders.append({"text": "", "orient": orient, "simple": True,
                             "arrow": head, "tp": tail if tail else head, "landing": landing,
                             "font": self.font_combo.currentFont().family(),
                             "size_ft": self.size_spin.value(), "bold": self.chk_bold.isChecked()})
        self._pending = {"arrow": None, "simple": True}; self.mode = "leader1"
        self._refresh_lists(); self._update_ui(); self._redraw()
        self._info("Leader colocado. Clic en la cabeza de flecha del siguiente (Esc para salir).")

    def on_dblclick(self, x, y):
        if self.mode not in ("idle", "move"): return
        thr = 14.0 / max(1e-6, self.canvas.transform().m11())
        for i, ld in enumerate(self.leaders):
            if not (ld.get("arrow") and ld.get("tp")): continue
            geo = self._leader_geo(ld); lx, ly = geo["label_pos"]; H = geo["H"]
            lines = ld["text"].split("\n"); tw = max((len(s) for s in lines), default=1) * H * 0.55; th = len(lines) * H
            hit_text = (lx - 8 <= x <= lx + tw + 8 and ly - 8 <= y <= ly + th + 8)
            if hit_text or math.hypot(ld["tp"][0] - x, ld["tp"][1] - y) < thr:
                self._edit_leader_text(i); return
        for i, tm in enumerate(self.text_marks):
            if self._text_hit(tm, x, y):
                self._edit_text_mark(i); return

    def _text_hit(self, tm, x, y):
        h = self._px_for_ft(tm["size_ft"]) if "size_ft" in tm else tm.get("h", 16)
        lines = tm["text"].split("\n")
        w = max((len(s) for s in lines), default=1) * h * 0.55; th = len(lines) * h
        px, py = tm["pos"]; return px - 6 <= x <= px + w + 6 and py - 6 <= y <= py + th + 6

    def _pick(self, x, y):
        thr = 10.0 / max(1e-6, self.canvas.transform().m11())
        # buzones/intersecciones primero (círculo pequeño): dist ≤ 8 px del centro
        bz_thr = 8.0 / max(1e-6, self.canvas.transform().m11())
        best_bz, bd_bz = -1, bz_thr
        for i, s in enumerate(self.structures):
            if s.get("world"): continue                 # los importados no se dibujan en el lienzo
            sx, sy = s.get("x"), s.get("y")
            if sx is None or sy is None: continue
            d = math.hypot(x - sx, y - sy)
            if d < bd_bz: bd_bz, best_bz = d, i
        if best_bz >= 0:
            self._no_center = True; self._show_tab(TAB_BZ); self.bz_list.setCurrentRow(best_bz)
            self._no_center = False; return
        for i, tm in enumerate(self.text_marks):        # textos primero (blancos pequeños)
            if self._text_hit(tm, x, y):
                self._no_center = True; self._show_tab(TAB_TEXT); self.txt_marks_list.setCurrentRow(i)
                self._no_center = False; return
        for i, ld in enumerate(self.leaders):           # leaders: por la línea o el texto
            if not (ld.get("arrow") and ld.get("tp")): continue
            geo = self._leader_geo(ld); hit = False
            for s in geo["segs"]:
                for a, b in zip(s, s[1:]):
                    if G.pt_seg_dist(x, y, a[0], a[1], b[0], b[1]) < thr: hit = True; break
                if hit: break
            lx, ly = geo["label_pos"]; H = geo["H"]
            tw = max((len(t) for t in ld["text"].split("\n")), default=1) * H * 0.6; tt = ld["text"].count("\n") + 1
            if not hit and lx - 8 <= x <= lx + tw + 8 and ly - 8 <= y <= ly + tt * H + 8: hit = True
            if hit:
                self._select_leader(i); return
        best, bd = -1, thr
        for i, p in enumerate(self.pipes):
            if not p.get("pts"): continue               # tramos importados (world) no están en el lienzo
            for a, b in zip(p["pts"], p["pts"][1:]):
                d = G.pt_seg_dist(x, y, a[0], a[1], b[0], b[1])
                if d < bd: bd, best = d, i
        if best >= 0:
            self._no_center = True; self._show_tab(TAB_PIPE); self.pipe_list.setCurrentRow(best)
            self._no_center = False

    def _snap(self, x, y):
        if not self.snap: return (x, y)
        return G.snap_point(self.gray, x, y, self.snap_r)

    # ─────────────────────────── editar / mover ───────────────────────────
    def _current_kind(self):
        ti = self._current_tab()
        if ti == TAB_PIPE and 0 <= self.sel_pipe < len(self.pipes): return "pipe"
        if ti == TAB_LEADER and 0 <= self.sel_leader < len(self.leaders) and self.leaders[self.sel_leader].get("simple"): return "leader"
        if ti == TAB_TEXT and 0 <= self.sel_text < len(self.text_marks): return "text"
        if ti == TAB_REGION and 0 <= self.sel_region < len(self.erase_regions): return "region"
        return None

    def enter_move(self):
        kind = self._current_kind()
        # Ctrl+T con un Leader simple seleccionado: asegura la pestaña Leaders y entra a editar sus vértices
        if not kind and 0 <= self.sel_leader < len(self.leaders) and self.leaders[self.sel_leader].get("simple"):
            self._select_leader(self.sel_leader, center=True); kind = "leader"
        if kind == "leader":
            self.set_mode("move"); self._info("Editar Leader: arrastra un vértice (posición/longitud) o el trazo para mover. Enter/Esc termina.")
        elif kind:
            self.set_mode("move"); self._info("Arrastra para mover · clic en vértice extiende (F) · clic derecho elimina")
        else:
            self._info("Selecciona primero una utilidad, Leader, texto o zona")

    def _thr(self): return 12.0 / max(1e-6, self.canvas.transform().m11())

    def _segments(self, pts, closed):
        segs = list(zip(range(len(pts) - 1), pts, pts[1:]))
        if closed and len(pts) >= 3: segs.append((len(pts) - 1, pts[-1], pts[0]))
        return segs

    def begin_move(self, x, y):
        kind = self._current_kind()
        if not kind: return
        self._push(); self._moved = False; self._move_kind = kind
        self._press_xy = (x, y); self._last_xy = (x, y); thr = self._thr()
        if kind == "text":
            self._move0 = (x, y); self._drag_vertex = None; self._edit_pts = None; return
        if kind == "leader":
            ld = self.leaders[self.sel_leader]; self._edit_leader = ld
            pts = [tuple(ld["arrow"])]
            if ld.get("landing"): pts.append(tuple(ld["landing"]))
            pts.append(tuple(ld["tp"]))
            self._edit_pts = pts; self._edit_closed = False
            vi, vd = -1, thr
            for i, (px, py) in enumerate(pts):
                d = math.hypot(px - x, py - y)
                if d < vd: vd, vi = d, i
            self._drag_vertex = vi if vi >= 0 else None       # vértice cercano → arrastra; si no, mueve todo
            self._move0 = None if vi >= 0 else (x, y); return
        pts = self.pipes[self.sel_pipe]["pts"] if kind == "pipe" else self.erase_regions[self.sel_region]["pts"]
        if not pts: return                              # tramo importado (world): no editable en el lienzo
        self._edit_pts = pts; self._edit_closed = (kind == "region")
        vi, vd = -1, thr
        for i, (px, py) in enumerate(pts):
            d = math.hypot(px - x, py - y)
            if d < vd: vd, vi = d, i
        if vi >= 0:
            self._drag_vertex = vi; self._move0 = None; return
        si, sd = -1, thr
        for idx, a, b in self._segments(pts, self._edit_closed):
            d = G.pt_seg_dist(x, y, a[0], a[1], b[0], b[1])
            if d < sd: sd, si = d, idx
        if si >= 0:
            pts.insert(si + 1, (x, y)); self._drag_vertex = si + 1; self._move0 = None; self._moved = True
            self._refresh_lists(); return
        self._drag_vertex = None; self._move0 = (x, y)

    def do_move(self, x, y):
        self._moved = True; self._last_xy = (x, y)
        if self._move_kind == "text" and 0 <= self.sel_text < len(self.text_marks):
            if self._move0:
                dx, dy = x - self._move0[0], y - self._move0[1]; self._move0 = (x, y)
                px, py = self.text_marks[self.sel_text]["pos"]
                self.text_marks[self.sel_text]["pos"] = (px + dx, py + dy); self._redraw()
            return
        if self._move_kind == "leader":
            pts = self._edit_pts
            if pts is None: return
            if self._drag_vertex is not None:
                pts[self._drag_vertex] = (x, y)               # mueve un vértice (posición/longitud)
            elif self._move0 is not None:
                dx, dy = x - self._move0[0], y - self._move0[1]; self._move0 = (x, y)
                for i in range(len(pts)): pts[i] = (pts[i][0] + dx, pts[i][1] + dy)
            else:
                return
            self._sync_leader(); self._redraw(); return
        pts = self._edit_pts
        if pts is None: return
        if self._drag_vertex is not None:
            pts[self._drag_vertex] = (x, y); self._redraw(); return
        if self._move0 is not None:
            dx, dy = x - self._move0[0], y - self._move0[1]; self._move0 = (x, y)
            for i in range(len(pts)): pts[i] = (pts[i][0] + dx, pts[i][1] + dy)
            self._redraw()

    def end_move(self):
        # clic (casi sin arrastrar) sobre un vértice de una tubería → extender
        dist = math.hypot(self._last_xy[0] - self._press_xy[0], self._last_xy[1] - self._press_xy[1]) \
            if (self._last_xy and self._press_xy) else 0
        if (self._move_kind == "pipe" and self._drag_vertex is not None and dist < 6
                and 0 <= self.sel_pipe < len(self.pipes)):
            self._move0 = None; self._move_kind = None
            self._start_extension(self.sel_pipe, self._drag_vertex); self._drag_vertex = None; return
        self._move0 = None; self._drag_vertex = None; self._move_kind = None; self._edit_leader = None

    def _sync_leader(self):
        """Vuelca los vértices editados (self._edit_pts) al leader (arrow / landing / tp)."""
        ld = self._edit_leader; pts = self._edit_pts
        if not ld or not pts: return
        ld["arrow"] = pts[0]
        if ld.get("landing"): ld["landing"] = pts[1]; ld["tp"] = pts[2]
        else: ld["tp"] = pts[-1]

    def _start_extension(self, pi, vi):
        # Sin ventana emergente: la casilla 'continuar la misma utilidad' decide.
        pts = self.pipes[pi]["pts"]; vpos = tuple(pts[vi]); is_end = (vi == 0 or vi == len(pts) - 1)
        same = self.chk_ext_same.isChecked() and is_end
        self._extending = True; self._ext_layer = self.pipes[pi]["layer"]; self.cur_pts = [vpos]
        if same:
            self._ext_pipe = pi; self._ext_at = "start" if vi == 0 else "end"
            self._info("Continuando la MISMA utilidad: clic para agregar puntos, Enter finaliza.")
        else:
            self._ext_pipe = None; self._ext_at = None
            if self.chk_ext_same.isChecked() and not is_end:
                self._info("Solo desde un extremo se continúa; se creará una utilidad NUEVA. Clic para agregar, Enter finaliza.")
            else:
                self._info("Utilidad NUEVA (rama en F): clic para agregar puntos, Enter finaliza.")
        self.set_mode("pipe")

    def _delete_vertex(self, x, y):
        kind = self._current_kind()
        if kind not in ("pipe", "region"): return
        pts = self.pipes[self.sel_pipe]["pts"] if kind == "pipe" else self.erase_regions[self.sel_region]["pts"]
        floor = 3 if kind == "region" else 2
        if len(pts) <= floor: self._info(f"Necesita al menos {floor} puntos"); return
        thr = self._thr(); vi, vd = -1, thr
        for i, (px, py) in enumerate(pts):
            d = math.hypot(px - x, py - y)
            if d < vd: vd, vi = d, i
        if vi >= 0:
            self._push(); pts.pop(vi); self._refresh_lists(); self._redraw(); self._info("Vértice eliminado")

    # ─────────────────────────── utilidades ───────────────────────────
    def finish_pipe(self):
        if self._extending and self._ext_pipe is not None and 0 <= self._ext_pipe < len(self.pipes):
            extra = self.cur_pts[1:]                       # el 1er punto es el vértice existente
            if extra:
                self._push(); pts = self.pipes[self._ext_pipe]["pts"]
                if self._ext_at == "end": pts.extend(extra)
                else: self.pipes[self._ext_pipe]["pts"] = list(reversed(extra)) + pts
        elif len(self.cur_pts) >= 2:
            layer = self._ext_layer if self._extending else self.active_layer()
            ab = False if self._extending else self.chk_ab.isChecked()
            self._push(); self.pipes.append({"layer": layer, "pts": self.cur_pts[:], "ab": ab,
                                             "diam": float(PIPE_DIAMETERS_IN[0]), "diam_unit": "in",
                                             "material": DEFAULT_PIPE_MATERIAL})
        self.cur_pts = []; self._extending = False; self._ext_layer = None; self._ext_pipe = None; self._ext_at = None
        self._refresh_lists(); self._update_ui(); self._redraw()

    def _sel_pipe(self, r):
        """Se llama cuando el usuario selecciona una utilidad en el inventario
        (lista derecha). Actualiza el panel de Propiedades con SUS datos.

        `self._prop_guard` es un pequeño truco: bloquea temporalmente los
        callbacks de los campos mientras los rellenamos con valores. Sin él,
        `setValue`/`setText` dispararía `_prop_changed()` y guardaría los
        datos ANTES de terminar de cargarlos (círculo vicioso)."""
        self.sel_pipe = r
        if 0 <= r < len(self.pipes):
            p = self.pipes[r]; pts = p.get("pts")
            if not self._no_center and pts:                # los tramos importados (world) no tienen pts
                mid = pts[len(pts) // 2]; self.canvas.centerOn(mid[0], mid[1])
            self._prop_guard = True
            self.prop_name.setText(p.get("name", ""))
            # El diámetro se deriva del "Tamaño" del catálogo (elegido más abajo).
            self.prop_part.setText(p.get("part", ""))
            self.prop_inv0.setValue(p.get("inv_start") or 0.0); self.prop_inv1.setValue(p.get("inv_end") or 0.0)
            mi = self.prop_material.findText(p.get("material") or DEFAULT_PIPE_MATERIAL)
            self.prop_material.setCurrentIndex(mi if mi >= 0 else 0)
            # findData busca el índice del combo cuya "data" (dato oculto) coincide
            # con "" | "pipe" | "pressure"; si no encuentra devuelve -1 → índice 0.
            idx = self.prop_nettype.findData(p.get("net_type", "") or "")
            self.prop_nettype.setCurrentIndex(idx if idx >= 0 else 0)
            self._reload_pipe_families(p)
            self._reload_pipe_noman(p)
            self._prop_guard = False
        self._update_ui(); self._redraw()

    def _reload_pipe_noman(self, p):
        """Repuebla la lista de vértices intermedios con checkbox 'sin buzón'.
        Extremos (primer y último) no se muestran — siempre llevan structure.
        También controla la visibilidad del botón de override por segmento."""
        self.prop_noman.blockSignals(True); self.prop_noman.clear()
        pts = p.get("pts") or []
        n = len(pts)
        sin = set(p.get("no_manhole_verts") or [])
        # Mostrar solo si hay al menos un vértice intermedio (n >= 3).
        show = n >= 3
        self.lbl_prop_noman.setVisible(show); self.prop_noman.setVisible(show)
        if show:
            for i in range(1, n - 1):
                it = QtWidgets.QListWidgetItem(f"Vértice {i + 1}  ({pts[i][0]:.0f}, {pts[i][1]:.0f})")
                it.setFlags(it.flags() | QtCore.Qt.ItemIsUserCheckable)
                it.setCheckState(QtCore.Qt.Checked if i in sin else QtCore.Qt.Unchecked)
                it.setData(QtCore.Qt.UserRole, i)
                self.prop_noman.addItem(it)
        self.prop_noman.blockSignals(False)
        # Override por segmento: solo tiene sentido si hay al menos 2 tramos.
        show_seg = n >= 3
        self.lbl_seg_override.setVisible(show_seg); self.btn_seg_override.setVisible(show_seg)
        n_over = len(p.get("seg_overrides") or {})
        self.btn_seg_override.setText(
            f"Editar por segmento…  ({n_over} override{'s' if n_over != 1 else ''})"
            if n_over else "Editar por segmento…")

    def _open_seg_override_dialog(self):
        """Diálogo con una fila por tramo (vertex i → i+1), un combo de familia y
        otro de tamaño en cada una. La primera opción de cada combo es 'usar la de
        la pipe' (sin override). Guarda en p['seg_overrides']."""
        if not (0 <= self.sel_pipe < len(self.pipes)): return
        import civil_catalog as _cc
        p = self.pipes[self.sel_pipe]
        pts = p.get("pts") or []
        n_seg = len(pts) - 1
        if n_seg < 1: return
        kind = self._pipe_net_kind(p)
        fams = ([] if not self.civil_year else
                (_cc.pressure_pipes(self.civil_year) if kind == "pressure"
                 else _cc.imperial_pipes(self.civil_year)))

        dlg = QtWidgets.QDialog(self); dlg.setWindowTitle("Variar familia/tamaño por segmento")
        dlg.resize(720, 480)
        v = QtWidgets.QVBoxLayout(dlg)
        v.addWidget(QtWidgets.QLabel(
            "<i>Deja '(usar la de la pipe)' para heredar la familia/tamaño global. "
            "Solo se envía al DXF lo que sí varíe.</i>"))
        tbl = QtWidgets.QTableWidget(n_seg, 3)
        tbl.setHorizontalHeaderLabels(["Tramo", "Familia", "Tamaño"])
        tbl.verticalHeader().setVisible(False)
        tbl.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        tbl.horizontalHeader().setStretchLastSection(True)
        tbl.setColumnWidth(0, 100); tbl.setColumnWidth(1, 320)
        v.addWidget(tbl, 1)

        current = dict(p.get("seg_overrides") or {})
        # Las keys pueden venir como strings si se cargaron desde JSON — normalizo.
        current = {int(k): dict(v) for k, v in current.items()}

        row_widgets = []
        for i in range(n_seg):
            lab = QtWidgets.QTableWidgetItem(f"{i + 1} → {i + 2}")
            lab.setFlags(QtCore.Qt.ItemIsEnabled)
            tbl.setItem(i, 0, lab)

            cb_fam = QtWidgets.QComboBox()
            cb_fam.addItem("(usar la de la pipe)", "")
            for f in fams:
                cb_fam.addItem(f"{f['pretty']}  [{f['subfolder']}]", f["id"])
            cb_size = QtWidgets.QComboBox()
            cb_size.addItem("(usar la de la pipe)", "")

            tbl.setCellWidget(i, 1, cb_fam)
            tbl.setCellWidget(i, 2, cb_size)
            row_widgets.append((cb_fam, cb_size))

            def _refill_sizes(cbf=cb_fam, cbs=cb_size, current_size=""):
                fid = cbf.currentData() or ""
                cbs.blockSignals(True); cbs.clear()
                cbs.addItem("(usar la de la pipe)", "")
                if fid:
                    sz = (_cc.pressure_pipe_sizes(self.civil_year, fid) if kind == "pressure"
                          else _cc.pipe_sizes(self.civil_year, fid))
                    for s in sz: cbs.addItem(s, s)
                if current_size:
                    for j in range(cbs.count()):
                        if cbs.itemData(j) == current_size: cbs.setCurrentIndex(j); break
                cbs.blockSignals(False)

            cb_fam.currentIndexChanged.connect(lambda _, f=_refill_sizes: f())

            # Preseleccionar override existente para esta fila.
            ov = current.get(i)
            if ov and ov.get("pipe_family"):
                for j in range(cb_fam.count()):
                    if cb_fam.itemData(j) == ov["pipe_family"]:
                        cb_fam.setCurrentIndex(j); break
                _refill_sizes(current_size=ov.get("pipe_size", ""))

        bb = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        bb.accepted.connect(dlg.accept); bb.rejected.connect(dlg.reject); v.addWidget(bb)
        if dlg.exec() != QtWidgets.QDialog.Accepted: return

        # Recolectar overrides no vacíos.
        self._push()
        new_over = {}
        for i, (cb_fam, cb_size) in enumerate(row_widgets):
            fid = cb_fam.currentData() or ""
            sz = cb_size.currentData() or ""
            if fid or sz:
                entry = {}
                if fid: entry["pipe_family"] = fid
                if sz:  entry["pipe_size"] = sz
                new_over[i] = entry
        p["seg_overrides"] = new_over
        self._dirty = True
        self._reload_pipe_noman(p)                   # actualiza el contador en el botón
        self._redraw()

    def _pipe_noman_changed(self, item):
        if self._prop_guard: return
        if not (0 <= self.sel_pipe < len(self.pipes)): return
        p = self.pipes[self.sel_pipe]; self._push()
        idx = item.data(QtCore.Qt.UserRole)
        sin = set(p.get("no_manhole_verts") or [])
        if item.checkState() == QtCore.Qt.Checked: sin.add(idx)
        else: sin.discard(idx)
        p["no_manhole_verts"] = sorted(sin)
        self._dirty = True
        self._redraw()

    def _pipe_net_kind(self, p):
        """Devuelve 'gravity' | 'pressure' | 'conduit' según la capa del pipe."""
        from model import network_kind
        return network_kind(p.get("layer") or "")

    def _reload_pipe_families(self, p):
        """Repuebla los combos prop_family y prop_size según la capa del pipe y el
        catálogo Civil 3D seleccionado. Aplica para todos los tipos:
          - gravity y conduit → catálogo imperial de pipes (PVC/HDPE/DI/concreto/CMP…)
          - pressure → catálogo AWWA sub-material (Flanged/PushOn/PVC/HDPE/…)"""
        import civil_catalog as _cc
        self.prop_family.blockSignals(True); self.prop_family.clear()
        self.prop_size.blockSignals(True); self.prop_size.clear()
        kind = self._pipe_net_kind(p)
        show = kind in ("gravity", "pressure", "conduit") and bool(self.civil_year)
        self.lbl_prop_family.setVisible(show); self.prop_family.setVisible(show)
        self.lbl_prop_size.setVisible(show); self.prop_size.setVisible(show)
        if not show:
            self.prop_family.blockSignals(False); self.prop_size.blockSignals(False); return
        fams = (_cc.pressure_pipes(self.civil_year) if kind == "pressure"
                else _cc.imperial_pipes(self.civil_year))
        self.prop_family.addItem("(por defecto)", "")
        for f in fams:
            idx = self.prop_family.count()
            self.prop_family.addItem(f"{f['pretty']}  [{f['subfolder']}]", f["id"])
            img = f.get("img_path")
            tip = f"<b>{f['pretty']}</b><br><i>{f['subfolder']}</i>"
            if img: tip += f"<br><img src='file:///{img.replace(chr(92), '/')}' width='220'>"
            self.prop_family.setItemData(idx, tip, QtCore.Qt.ToolTipRole)
        cur = p.get("pipe_family", "") or ""
        for i in range(self.prop_family.count()):
            if self.prop_family.itemData(i) == cur:
                self.prop_family.setCurrentIndex(i); break
        self._load_pipe_sizes(kind, cur, p.get("pipe_size", "") or "")
        self.prop_family.blockSignals(False); self.prop_size.blockSignals(False)

    def _load_pipe_sizes(self, kind, fid, current):
        import civil_catalog as _cc
        self.prop_size.blockSignals(True); self.prop_size.clear()
        if not fid or not self.civil_year:
            self.prop_size.addItem("(sin familia)", ""); self.prop_size.setEnabled(False)
            self.prop_size.blockSignals(False); return
        sizes = (_cc.pressure_pipe_sizes(self.civil_year, fid) if kind == "pressure"
                 else _cc.pipe_sizes(self.civil_year, fid))
        if not sizes:
            self.prop_size.addItem("(sin tamaños)", ""); self.prop_size.setEnabled(False)
        else:
            self.prop_size.setEnabled(True); self.prop_size.addItem("(por defecto)", "")
            for sz in sizes: self.prop_size.addItem(sz, sz)
            if current:
                for i in range(self.prop_size.count()):
                    if self.prop_size.itemData(i) == current:
                        self.prop_size.setCurrentIndex(i); break
        self.prop_size.blockSignals(False)

    def _pipe_family_changed(self, _idx):
        if self._prop_guard: return
        if not (0 <= self.sel_pipe < len(self.pipes)): return
        p = self.pipes[self.sel_pipe]
        fid = self.prop_family.currentData() or ""
        p["pipe_family"] = fid; p["pipe_size"] = ""
        self._load_pipe_sizes(self._pipe_net_kind(p), fid, "")
        self._dirty = True; self._redraw()

    def _leader_at_row(self, lst, r):
        """Índice real en self.leaders del item de la fila r (o -1)."""
        it = lst.item(r) if r is not None and r >= 0 else None
        return it.data(QtCore.Qt.UserRole) if it is not None else -1

    def _sel_leader(self, r):                              # pestaña Multileaders
        """Al seleccionar un Multileader en el inventario derecho, abrimos la
        sección Multileader del acordeón izquierdo — así el usuario ve el
        estilo/orientación y puede editarlos."""
        i = self._leader_at_row(self.lead_list, r); self.sel_leader = i
        if 0 <= i < len(self.leaders):
            ld = self.leaders[i]
            if not self._no_center and ld.get("tp"): self.canvas.centerOn(ld["tp"][0], ld["tp"][1])
            self._open_section("ml")
            self._style_guard = True
            self.font_combo.setCurrentFont(QtGui.QFont(ld.get("font", C.TEXT_FONT)))
            self.size_spin.setValue(ld.get("size_ft", LEADER_TEXT_FT)); self.chk_bold.setChecked(bool(ld.get("bold")))
            self._style_guard = False
        self._update_ui(); self._redraw()

    def _sel_sleader(self, r):                             # pestaña Leaders (solo flechas)
        i = self._leader_at_row(self.sleader_list, r); self.sel_leader = i
        if 0 <= i < len(self.leaders):
            ld = self.leaders[i]
            if not self._no_center and ld.get("tp"): self.canvas.centerOn(ld["tp"][0], ld["tp"][1])
        self._update_ui(); self._redraw()

    def _select_leader(self, i, center=False):
        """Selecciona el leader nº i en la pestaña que le corresponde (ML o Leaders)."""
        if not (0 <= i < len(self.leaders)): return
        simple = self.leaders[i].get("simple")
        lst = self.sleader_list if simple else self.lead_list
        row = next((r for r in range(lst.count()) if lst.item(r).data(QtCore.Qt.UserRole) == i), -1)
        self._no_center = not center
        self._show_tab(TAB_LEADER if simple else TAB_ML); lst.setCurrentRow(row)
        self._no_center = False

    def _sel_text(self, r):
        """Al seleccionar un texto libre, abrimos la sección Texto libre del
        acordeón para que el estilo (fuente/altura/negrita/rotación) sea editable."""
        self.sel_text = r
        if 0 <= r < len(self.text_marks):
            tm = self.text_marks[r]
            if not self._no_center: self.canvas.centerOn(tm["pos"][0], tm["pos"][1])
            self._open_section("text")
            self._style_guard = True
            self.font_combo.setCurrentFont(QtGui.QFont(tm.get("font", C.TEXT_FONT)))
            self.size_spin.setValue(tm.get("size_ft", 3.0)); self.chk_bold.setChecked(bool(tm.get("bold")))
            self.rot_spin.setValue(int(tm.get("rot", 0)) % 360); self._style_guard = False
        self._update_ui(); self._redraw()

    def _sel_region(self, r):
        self.sel_region = r
        if not self._no_center and 0 <= r < len(self.erase_regions):
            pts = self.erase_regions[r]["pts"]
            cx = sum(p[0] for p in pts) / len(pts); cy = sum(p[1] for p in pts) / len(pts); self.canvas.centerOn(cx, cy)
        self._update_ui(); self._redraw()

    def _region_toggled(self, item):
        r = self.region_list.row(item)
        if 0 <= r < len(self.erase_regions):
            self.erase_regions[r]["enabled"] = (item.checkState() == QtCore.Qt.Checked); self._dirty = True; self._redraw()

    # ─────────────────────── Acordeón: cambios de sección ────────────────────────
    def _place_widget(self, w, target_layout):
        """Mueve un widget al layout indicado (Qt reasigna su padre automáticamente).
        Es como "mover una caja" de un estante a otro. Si el widget ya estaba en un
        layout, primero lo quitamos de ese layout (removeWidget). Si no hacemos esto,
        Qt puede dejar celdas huérfanas o mostrar el widget en dos sitios."""
        if w.parent() is not None:
            pl = w.parent().layout()
            if pl is not None:
                pl.removeWidget(w)
        target_layout.addWidget(w)

    def _on_toolbox_change(self, idx):
        """Al abrir una sección distinta del acordeón:
          1) Salimos del modo activo (evita mezclar 'colocar Multileader' con
             el usuario abriendo la sección Texto por error).
          2) Reparentamos los widgets COMPARTIDOS a la sección correspondiente:
             - orient_combo → Leader
             - gtxt (estilo) → Texto libre
             - rot_row visible solo en Texto libre
          3) Refrescamos la UI (etiqueta del modo, botones activos, etc.)."""
        # Descubrimos qué sección ("key") corresponde al índice.
        key = next((k for k, i in self._sec_idx.items() if i == idx), None)
        if self.mode not in ("idle",):
            self.set_mode("idle")
        if key == "leader":
            self._place_widget(self.orient_combo, self._slot_orient_ld)
        elif key == "text":
            self._place_widget(self.gtxt, self._slot_style_tx)
            self.rot_row.setVisible(True)
        self._update_ui()

    def _open_section(self, key):
        """Abre programáticamente una sección del acordeón por su nombre lógico.
        La usamos p.ej. cuando el usuario selecciona un Multileader en el
        inventario: abrimos automáticamente la sección Multileader para que
        vea su estilo/orientación y pueda editarlos."""
        i = self._sec_idx.get(key)
        if i is not None and self.toolbox.currentIndex() != i:
            self.toolbox.setCurrentIndex(i)

    def _bump_size(self, delta):
        """Sube o baja la altura del texto en pasos de 0.5 pies."""
        self.size_spin.setValue(round(max(0.5, self.size_spin.value() + delta), 2))

    def _bump_rot(self, delta):
        self.rot_spin.setValue((self.rot_spin.value() + delta) % 360)

    def _bump_opacity(self, delta_pct):
        val = self.canvas.pdf_opacity + delta_pct / 100.0
        self.canvas.set_pdf_opacity(val)
        self.lbl_opacity.setText(f"{round(self.canvas.pdf_opacity * 100)}%")

    def _px_for_ft(self, ft):
        raw = ft / self.scale * self.zoom if self.scale else ft * self.zoom
        cap = self.pageH_px * 0.06 if self.pageH_px else 200
        return max(8.0, min(raw, cap))

    def _style_changed(self):
        if self._style_guard: return
        ti = self._current_tab()
        if ti == TAB_TEXT and 0 <= self.sel_text < len(self.text_marks):
            tm = self.text_marks[self.sel_text]; self._push()
            tm["font"] = self.font_combo.currentFont().family(); tm["size_ft"] = self.size_spin.value()
            tm["bold"] = self.chk_bold.isChecked(); tm["rot"] = self.rot_spin.value() % 360
            self._redraw()
        elif ti == TAB_ML and 0 <= self.sel_leader < len(self.leaders):
            ld = self.leaders[self.sel_leader]; self._push()
            ld["font"] = self.font_combo.currentFont().family(); ld["size_ft"] = self.size_spin.value()
            ld["bold"] = self.chk_bold.isChecked()
            self._redraw()

    def _prop_changed(self):
        """Callback: cualquier cambio en el panel Propiedades escribe al modelo.
        `self._push()` guarda un snapshot para Ctrl+Z (deshacer).
        La UNIDAD de diámetro/invert es la del proyecto (self.work_unit), no per-pipe."""
        if self._prop_guard: return
        if self._current_tab() == TAB_PIPE and 0 <= self.sel_pipe < len(self.pipes):
            p = self.pipes[self.sel_pipe]; self._push()
            p["name"] = self.prop_name.text().strip()
            p["diam_unit"] = "in"                                        # el diámetro nunca va en pies
            p["unit"] = self.work_unit                                  # unidad de trabajo (coords/cotas)
            p["part"] = self.prop_part.text().strip()
            p["inv_start"] = self.prop_inv0.value(); p["inv_end"] = self.prop_inv1.value()
            p["material"] = self.prop_material.currentText()
            p["net_type"] = self.prop_nettype.currentData() or ""
            # Familia + tamaño del catálogo Civil 3D. El diámetro se deriva del tamaño.
            if self.prop_family.isVisible():
                p["pipe_family"] = self.prop_family.currentData() or ""
                p["pipe_size"] = self.prop_size.currentData() or "" if self.prop_size.isEnabled() else ""
            # p["diam"] se calcula del pipe_size (p.ej. "24 in" → 24.0). Sin tamaño → 0.
            p["diam"] = _extract_diam_from_size(p.get("pipe_size", ""))
            self._refresh_lists()

    def _refresh_unit_labels(self):
        """Etiquetas de campo fijas: cotas en PIES, diámetro en PULGADAS.
        (Ya no hay selector de unidad; todo va por campo.)"""
        # (Diámetro se muestra vía combo de tamaño del catálogo, no necesita etiqueta aquí)
        if hasattr(self, "lbl_prop_inv0"): self.lbl_prop_inv0.setText("Elev. de rasante inicial (ft):")
        if hasattr(self, "lbl_prop_inv1"): self.lbl_prop_inv1.setText("Elev. de rasante final (ft):")

    def _refresh_counts(self):
        """Contadores en vivo en la barra de estado: utilidades, leaders, textos, zonas."""
        if not hasattr(self, "lbl_counts"): return
        n_p = len(self.pipes); n_l = len(self.leaders); n_t = len(self.text_marks); n_z = len(self.erase_regions)
        self.lbl_counts.setText(f"{n_p} util · {n_l} lead · {n_t} txt · {n_z} zona")
        # marca de "sin guardar"
        if hasattr(self, "lbl_dirty"):
            if self._dirty:
                self.lbl_dirty.setText("●"); self.lbl_dirty.setStyleSheet("color:#e0c060;")
            else:
                self.lbl_dirty.setText("")

    def change_pipe_type(self):
        if 0 <= self.sel_pipe < len(self.pipes):
            self._push(); self.pipes[self.sel_pipe]["layer"] = self.active_layer()
            self.pipes[self.sel_pipe]["ab"] = self.chk_ab.isChecked(); self._refresh_lists(); self._redraw()

    def edit_selected_text(self):
        ti = self._current_tab()
        if ti == TAB_ML and 0 <= self.sel_leader < len(self.leaders): self._edit_leader_text(self.sel_leader)
        elif ti == TAB_TEXT and 0 <= self.sel_text < len(self.text_marks): self._edit_text_mark(self.sel_text)

    def _list_context_menu(self, listw, tab_idx, pos):
        item = listw.itemAt(pos)
        if item is None: return
        if self._current_tab() != tab_idx: self._show_tab(tab_idx)
        listw.setCurrentRow(listw.row(item))          # selecciona la fila bajo el cursor
        menu = QtWidgets.QMenu(self)
        if tab_idx == TAB_PIPE:
            self._menu_act(menu, "Cambiar tipo", self.change_pipe_type)
            self._menu_act(menu, "Editar/mover", self.enter_move)
        elif tab_idx == TAB_ML:
            self._menu_act(menu, "Editar texto", self.edit_selected_text)
        elif tab_idx == TAB_LEADER:
            self._menu_act(menu, "Editar/mover", self.enter_move)
        elif tab_idx == TAB_TEXT:
            self._menu_act(menu, "Mover", self.enter_move)
            self._menu_act(menu, "Editar texto", self.edit_selected_text)
        elif tab_idx == TAB_REGION:
            self._menu_act(menu, "Editar/mover", self.enter_move)
        menu.addSeparator()
        self._menu_act(menu, "Eliminar", self.delete_selected)
        menu.exec(listw.viewport().mapToGlobal(pos))

    def delete_selected(self):
        ti = self._current_tab()
        if ti == TAB_PIPE and 0 <= self.sel_pipe < len(self.pipes):
            self._push(); self.pipes.pop(self.sel_pipe); self.sel_pipe = -1
        elif ti in (TAB_ML, TAB_LEADER) and 0 <= self.sel_leader < len(self.leaders):
            self._push(); self.leaders.pop(self.sel_leader); self.sel_leader = -1
        elif ti == TAB_TEXT and 0 <= self.sel_text < len(self.text_marks):
            self._push(); self.text_marks.pop(self.sel_text); self.sel_text = -1
        elif ti == TAB_REGION and 0 <= self.sel_region < len(self.erase_regions):
            self._push(); self.erase_regions.pop(self.sel_region); self.sel_region = -1
        self._refresh_lists(); self._redraw()

    def _copy_sel(self):
        ti = self._current_tab()
        if ti in (TAB_ML, TAB_LEADER) and 0 <= self.sel_leader < len(self.leaders): self._clip = ("leader", copy.deepcopy(self.leaders[self.sel_leader]))
        elif ti == TAB_PIPE and 0 <= self.sel_pipe < len(self.pipes): self._clip = ("pipe", copy.deepcopy(self.pipes[self.sel_pipe]))
        elif ti == TAB_TEXT and 0 <= self.sel_text < len(self.text_marks): self._clip = ("text", copy.deepcopy(self.text_marks[self.sel_text]))
        else: self._info("Selecciona algo para copiar."); return
        self._info("Copiado. Ctrl+V para pegar una copia.")

    def _paste_sel(self):
        if not self._clip: return
        kind, obj = self._clip; o = copy.deepcopy(obj); d = 30; self._push()
        if kind == "leader":
            if o.get("arrow"): o["arrow"] = (o["arrow"][0] + d, o["arrow"][1] + d)
            if o.get("landing"): o["landing"] = (o["landing"][0] + d, o["landing"][1] + d)
            if o.get("tp"): o["tp"] = (o["tp"][0] + d, o["tp"][1] + d)
            self.leaders.append(o); self._refresh_lists(); self._select_leader(len(self.leaders) - 1)
        elif kind == "pipe":
            o["pts"] = [(x + d, y + d) for (x, y) in o["pts"]]; self.pipes.append(o)
            self._show_tab(TAB_PIPE); self._refresh_lists(); self.pipe_list.setCurrentRow(len(self.pipes) - 1)
        elif kind == "text":
            o["pos"] = (o["pos"][0] + d, o["pos"][1] + d); self.text_marks.append(o)
            self._show_tab(TAB_TEXT); self._refresh_lists(); self.txt_marks_list.setCurrentRow(len(self.text_marks) - 1)
        self._redraw(); self._info("Pegado (copia desplazada).")

    def _refresh_lists(self):
        self._refresh_counts()
        self.pipe_list.blockSignals(True); self.pipe_list.clear()
        for i, p in enumerate(self.pipes, 1):
            tag = " (AB)" if p.get("ab") else ""
            nm = f" · {p['name']}" if p.get("name") else ""
            n = len(p.get("pts") or [])
            info = f"red:{p.get('net', '')}" if p.get("world") else str(n)
            it = QtWidgets.QListWidgetItem(swatch_icon(layer_qcolor(p["layer"])), f"{i}. {p['layer']}{tag}{nm} ({info})")
            it.setForeground(QtGui.QColor("white")); self.pipe_list.addItem(it)
        self.pipe_list.blockSignals(False)
        self.lead_list.blockSignals(True); self.lead_list.clear()
        self.sleader_list.blockSignals(True); self.sleader_list.clear()
        nml = ns = 0
        for i, ld in enumerate(self.leaders):
            if ld.get("simple"):
                ns += 1; o = {"h": "horizontal", "v": "vertical", "d": "diagonal"}.get(ld.get("orient", "d"), "")
                it = QtWidgets.QListWidgetItem(f"{ns}. Leader {o}".rstrip())
                it.setData(QtCore.Qt.UserRole, i); self.sleader_list.addItem(it)
            else:
                nml += 1; it = QtWidgets.QListWidgetItem(f"{nml}. {ld['text'][:24].replace(chr(10), ' / ')}")
                it.setData(QtCore.Qt.UserRole, i); self.lead_list.addItem(it)
        self.lead_list.blockSignals(False); self.sleader_list.blockSignals(False)
        self.txt_marks_list.blockSignals(True); self.txt_marks_list.clear()
        for i, tm in enumerate(self.text_marks, 1): self.txt_marks_list.addItem(f"{i}. {tm['text'][:28].replace(chr(10), ' / ')}")
        self.txt_marks_list.blockSignals(False)
        self.region_list.blockSignals(True); self.region_list.clear()
        for i, rg in enumerate(self.erase_regions, 1):
            it = QtWidgets.QListWidgetItem(f"Zona {i} ({len(rg['pts'])} vértices)")
            it.setFlags(it.flags() | QtCore.Qt.ItemIsUserCheckable)
            it.setCheckState(QtCore.Qt.Checked if rg.get("enabled", True) else QtCore.Qt.Unchecked)
            it.setForeground(QtGui.QColor("white")); self.region_list.addItem(it)
        self.region_list.blockSignals(False)
        # Refrescar tab Buzones + panel de propiedades del buzón seleccionado.
        self._rebuild_structures()
        self.bz_list.blockSignals(True); self.bz_list.clear()
        for i, s in enumerate(self.structures, 1):
            fam = s.get("part") or "(sin familia)"
            sz = f"  {s['part_size']}" if s.get("part_size") else ""
            emoji = "🟠" if s.get("net") == "conduit" else "🔵"
            it = QtWidgets.QListWidgetItem(f"{emoji} {s.get('cod', '?')}  ·  {fam}{sz}")
            self.bz_list.addItem(it)
        self.bz_list.blockSignals(False)
        self._sync_bz_panel()

    # ─────────────────────────── borrar zona ───────────────────────────
    def finish_erase(self):
        if len(self._erase_pts) < 3: return
        self._push(); poly = self._erase_pts[:]; self.erase_regions.append({"pts": poly, "enabled": True})
        self.pipes = [p for p in self.pipes if not (p.get("pts") and all(point_in_poly(px, py, poly) for (px, py) in p["pts"]))]
        self.leaders = [l for l in self.leaders if not (l.get("tp") and point_in_poly(l["tp"][0], l["tp"][1], poly))]
        self.text_marks = [t for t in self.text_marks if not point_in_poly(t["pos"][0], t["pos"][1], poly)]
        self._erase_pts = []; self.set_mode("idle"); self._refresh_lists()
        self._info("Zona agregada: al exportar también se borra la geometría base del plano dentro de ella.")

    # ─────────────────────────── Multileader ───────────────────────────
    def _read_leader_text(self):
        if self.chk_custom.isChecked(): return self.txt_edit.text().strip()
        it = self.text_list.currentItem()
        return it.text() if it and self.text_list.currentRow() >= 0 else ""

    def start_leader(self, simple=True):
        """Entra al modo de colocación de Leader (solo flecha, sin texto). La
        orientación se lee de `orient_combo` AL MOMENTO del clic final para que
        el usuario pueda cambiarla dentro del modo. `simple` se mantiene solo
        por retrocompat de callers antiguos; siempre es True ahora."""
        self._pending = {"arrow": None, "simple": True}
        self._open_section("leader")
        self.set_mode("leader1")
        if self.orient_combo.currentData() == "d":
            self._info("Leader diagonal: cabeza → inicio del landing (bisagra) → final del cuerpo. Enter/Esc para salir.")
        else:
            self._info("Leader: cabeza de flecha → final del cuerpo. Enter/Esc para salir.")

    def _edit_leader_text(self, idx):
        ld = self.leaders[idx]; tp = ld["tp"]
        if ld.get("simple"):                                 # el Leader simple no tiene texto que editar
            self._info("El Leader simple no lleva texto."); return
        def commit(val):
            self._close_editor()
            if val.strip(): self._push(); ld["text"] = val.rstrip("\n"); self._refresh_lists()
            self._redraw()
        self._open_editor(tp[0], tp[1] - self.leader_hpx, ld["text"], commit)

    def _leader_geo(self, ld):
        """Geometría del Multileader (px). La 'cola' (parte de la línea junto al texto)
        se adapta al largo del texto. La punta se orienta con segs[0][1]."""
        ax, ay = ld["arrow"]; tx, ty = ld["tp"]
        ftsize = ld.get("size_ft", LEADER_TEXT_FT)
        H = self._px_for_ft(ftsize)
        lines = ld["text"].split("\n"); maxlen = max((len(s) for s in lines), default=1); nlines = len(lines)
        tw = max(maxlen * H * 0.6, H * 2); th = nlines * H; gap = H * 0.5; near = H * 0.22
        if ld.get("simple"):                               # LEADER simple: solo flecha, sin texto
            lp = ld.get("landing")
            if lp:                                         # diagonal con landing: cabeza → bisagra → final
                segs = [[(ax, ay), (lp[0], lp[1]), (tx, ty)]]
            else:                                          # recto h/v: cabeza → final del cuerpo
                segs = [[(ax, ay), (tx, ty)]]
            end = segs[0][-1]
            return dict(segs=segs, label_pos=end, rot=0, side="right", verts_px=segs[0], insert_px=end,
                        dogleg=0.0, H=H, cad_h=ftsize, tcenter_px=end, cad_rot=0)
        orient = ld.get("orient", "h")
        if orient == "v":                                  # recto vertical; texto vertical junto a la cola
            signY = -1 if ty < ay else 1
            L = max(abs(ty - ay), tw + gap); ey = ay + signY * L; my = (ay + ey) / 2
            side = "top" if signY < 0 else "bottom"
            lbl = (ax + H * 0.08, my + tw / 2)              # rot -90, centrado a lo largo, pegado a la línea
            segs = [[(ax, ay), (ax, ey)]]
            return dict(segs=segs, label_pos=lbl, rot=-90, side=side, verts_px=segs[0], insert_px=(ax, ey),
                        dogleg=0.0, H=H, cad_h=ftsize, tcenter_px=(ax + th / 2 + H * 0.08, my), cad_rot=90)
        if orient == "h":                                  # recto horizontal; texto encima de la cola
            signX = 1 if tx >= ax else -1
            L = max(abs(tx - ax), tw + gap); ex = ax + signX * L
            lblx = ex - tw if signX > 0 else ex
            side = "right" if signX > 0 else "left"
            segs = [[(ax, ay), (ex, ay)]]
            return dict(segs=segs, label_pos=(lblx, ay - th - near), rot=0, side=side, verts_px=segs[0], insert_px=(ex, ay),
                        dogleg=0.0, H=H, cad_h=ftsize, tcenter_px=(lblx + tw / 2, ay - near - th / 2), cad_rot=0)
        # diagonal: flecha → 2º clic → landing horizontal → texto encima del landing
        right = tx >= ax; sgn = 1 if right else -1
        lx = tx + sgn * (tw + gap); text_x = min(tx, lx) + gap
        side = "right" if right else "left"
        lbl = (text_x, ty - H - near)                      # 1ª línea encima; extras al otro lado
        segs = [[(ax, ay), (tx, ty), (lx, ty)]]
        return dict(segs=segs, label_pos=lbl, rot=0, side=side, verts_px=segs[0], insert_px=(lx, ty),
                    dogleg=0.0, H=H, cad_h=ftsize, tcenter_px=(text_x + tw / 2, ty - H - near + th / 2), cad_rot=0)

    # ─────────────────────────── texto libre ───────────────────────────
    def _new_free_text(self, x, y):
        def commit(val):
            self._close_editor()
            if val.strip():
                self._push()
                self.text_marks.append({"pos": (x, y), "text": val.rstrip("\n"),
                                        "size_ft": self.size_spin.value(), "font": self.font_combo.currentFont().family(),
                                        "bold": self.chk_bold.isChecked(), "rot": self.rot_spin.value() % 360, "free": True})
                self._refresh_lists()
            self._redraw()
        self._open_editor(x, y, "", commit)

    def _edit_text_mark(self, idx):
        tm = self.text_marks[idx]
        def commit(val):
            self._close_editor()
            if val.strip(): self._push(); tm["text"] = val.rstrip("\n"); self._refresh_lists()
            self._redraw()
        self._open_editor(tm["pos"][0], tm["pos"][1], tm["text"], commit)

    def _open_editor(self, x, y, initial, on_commit, w=220):
        # Caja flotante hija del viewport (no escala con el zoom y recibe el teclado
        # de forma fiable: Enter aplica, Ctrl+Shift+Enter salta de línea, clic fuera aplica).
        self._close_editor()
        ed = InlineEdit(initial); ed.setParent(self.canvas.viewport())
        ed.setFixedWidth(w); ed.setFixedHeight(64)
        vp = self.canvas.mapFromScene(QtCore.QPointF(x, y))
        ed.move(vp); ed.show(); ed.raise_()
        ed.committed.connect(on_commit); self._editor = ed
        ed.setFocus(QtCore.Qt.OtherFocusReason); ed.selectAll()

    def _close_editor(self):
        if self._editor:
            ed = self._editor; self._editor = None
            try: ed.hide(); ed.deleteLater()
            except Exception: pass

    # ─────────────────────────── dibujo ───────────────────────────
    def _redraw(self):
        sc = self.canvas.scene()
        for it in self._overlay:
            try: sc.removeItem(it)
            except (RuntimeError, ValueError): pass
        self._overlay = []
        if self.canvas.pixmap_item is None: return
        # zonas de borrado — DETRÁS de todo (solo tapan el PDF)
        if self._erase_pts:
            self._poly(self._erase_pts, QtGui.QColor(255, 220, 0), 1.8, dots=True, z=Z_MARK)
        for i, rg in enumerate(self.erase_regions):
            qp = QtGui.QPolygonF([QtCore.QPointF(px, py) for (px, py) in rg["pts"]])
            enabled = rg.get("enabled", True); sel = (i == self.sel_region)
            if not enabled:
                pen = QtGui.QPen(QtGui.QColor(150, 150, 150), 1.2, QtCore.Qt.DashLine); brush = QtGui.QBrush(QtGui.QColor(200, 200, 200, 30))
            elif sel:
                pen = QtGui.QPen(QtGui.QColor(255, 40, 40), 2.0); brush = QtGui.QBrush(QtGui.QColor(255, 255, 255, 128))
            else:
                pen = QtGui.QPen(QtGui.QColor(255, 255, 255), 1.6); brush = QtGui.QBrush(QtGui.QColor(255, 255, 255, 255))
            pen.setCosmetic(True); it = sc.addPolygon(qp, pen, brush); it.setZValue(Z_ERASE); self._overlay.append(it)
            if sel and self.mode == "move": self._handles(rg["pts"])
        # utilidades
        for i, p in enumerate(self.pipes):
            if not p.get("pts"): continue               # tramos importados (world): no se dibujan
            sel = (i == self.sel_pipe)
            self._poly(p["pts"], layer_qcolor(p["layer"]), 4.0 if sel else 2.0, z=Z_MARK)
            # Marcador magenta en los vértices "sin buzón" de la pipe seleccionada,
            # para identificarlos de un vistazo (además del checkbox del panel).
            if sel:
                sin = set(p.get("no_manhole_verts") or [])
                if sin:
                    pen_nm = QtGui.QPen(QtGui.QColor(255, 40, 200), 2.0); pen_nm.setCosmetic(True)
                    brush_nm = QtGui.QBrush(QtGui.QColor(255, 40, 200))
                    for vi in sin:
                        if 0 <= vi < len(p["pts"]):
                            vx, vy = p["pts"][vi]
                            it = sc.addEllipse(vx - 7, vy - 7, 14, 14, pen_nm, brush_nm)
                            it.setZValue(Z_MARK + 2); self._overlay.append(it)
            if sel and self.mode == "move": self._handles(p["pts"])
        self._poly(self.cur_pts, layer_qcolor(self._ext_layer or self.active_layer()), 2.0, dots=True, z=Z_MARK)
        # leaders
        anno = aci_qcolor(8)
        for i, ld in enumerate(self.leaders):
            if not (ld.get("arrow") and ld.get("tp")): continue
            col = QtGui.QColor(120, 220, 120) if i == self.sel_leader else anno
            geo = self._leader_geo(ld)
            for s in geo["segs"]: self._poly(s, col, 1.6, z=Z_MARK)
            self._arrow(ld["arrow"], geo["segs"][0][1], col)   # punta orientada a lo largo de la línea
            if i == self.sel_leader and self.mode == "move" and ld.get("simple"):
                self._handles(geo["segs"][0])                  # vértices editables (cabeza / bisagra / final)
            if ld["text"]:                                     # Leader simple no lleva texto
                t = sc.addText(ld["text"]); t.setDefaultTextColor(col); t.document().setDocumentMargin(0)
                f = t.font(); f.setPixelSize(int(geo["H"])); t.setFont(f)
                if geo["rot"]: t.setRotation(geo["rot"])
                t.setPos(geo["label_pos"][0], geo["label_pos"][1]); t.setZValue(Z_MARK); self._overlay.append(t)
        # buzones — círculo relleno con el color del pipe al que pertenecen.
        # Se dibuja por encima de las utilidades (mismo z que MARK). Si show_bz_labels
        # está activo, el código se dibuja al lado con una fuente pequeña blanca.
        self._draw_structures()
        # textos
        for i, tm in enumerate(self.text_marks):
            t = sc.addText(tm["text"]); t.setDefaultTextColor(QtGui.QColor(120, 220, 120)); t.document().setDocumentMargin(0)
            hpx = self._px_for_ft(tm["size_ft"]) if "size_ft" in tm else tm.get("h", 16)
            f = t.font(); f.setPixelSize(max(6, int(hpx)))
            if tm.get("font"): f.setFamily(tm["font"])
            f.setBold(bool(tm.get("bold"))); t.setFont(f)
            t.setPos(tm["pos"][0], tm["pos"][1])
            if tm.get("rot"): t.setRotation(-tm["rot"])       # rot en grados CCW; Qt gira en sentido horario
            t.setZValue(Z_MARK); self._overlay.append(t)
            if i == self.sel_text and self._current_tab() == TAB_TEXT:
                br = t.boundingRect(); pen = QtGui.QPen(QtGui.QColor(255, 180, 40)); pen.setCosmetic(True)
                rit = sc.addRect(tm["pos"][0], tm["pos"][1], br.width(), br.height(), pen); rit.setZValue(Z_MARK); self._overlay.append(rit)

    def _draw_structures(self):
        """Dibuja cada buzón como un círculo relleno con el color del pipe al que
        pertenece (mismo vértice). Si show_bz_labels está activo, escribe el código
        del buzón al lado del círculo."""
        if not self.structures: return
        sc = self.canvas.scene(); tol2 = 14.0 ** 2
        # Precomputa color por buzón: mira los pipes NO importados y toma el layer
        # del primero cuyo vértice coincida (dist² ≤ tol²).
        def _color_for(s):
            sx, sy = s.get("x"), s.get("y")
            if sx is None or sy is None: return QtGui.QColor(200, 200, 200)
            if s.get("world") or s.get("net") == "pressure":
                # Para presión (sin vértice de pipe dibujado) o buzones importados,
                # gris claro (no hay línea de referencia visible).
                pass
            for p in self.pipes:
                if p.get("world") or not p.get("pts"): continue
                for (vx, vy) in p["pts"]:
                    if (vx - sx) ** 2 + (vy - sy) ** 2 <= tol2:
                        return layer_qcolor(p["layer"])
            return QtGui.QColor(180, 180, 180)     # buzón sin pipe cercano (raro)
        pen = QtGui.QPen(QtGui.QColor(255, 255, 255), 1.2); pen.setCosmetic(True)
        pen_sel = QtGui.QPen(QtGui.QColor(255, 220, 40), 2.5); pen_sel.setCosmetic(True)
        R = 6.0                                     # radio en px (independiente del zoom por _cosmetic pen)
        for i, s in enumerate(self.structures):
            sx, sy = s.get("x"), s.get("y")
            if sx is None or sy is None: continue
            if s.get("world"): continue             # los importados (Excel) están en coord mundo, no lienzo
            col = _color_for(s); brush = QtGui.QBrush(col)
            use_pen = pen_sel if i == getattr(self, "sel_bz", -1) else pen
            r_use = R + 1.5 if i == getattr(self, "sel_bz", -1) else R
            it = sc.addEllipse(sx - r_use, sy - r_use, 2 * r_use, 2 * r_use, use_pen, brush)
            it.setZValue(Z_MARK + 1); self._overlay.append(it)
            if self.show_bz_labels and s.get("cod"):
                t = sc.addText(s["cod"]); t.setDefaultTextColor(QtGui.QColor(180, 180, 180))
                t.document().setDocumentMargin(0)
                f = t.font(); f.setPixelSize(11); f.setBold(True); t.setFont(f)
                t.setPos(sx + R + 2, sy - R - 2); t.setZValue(Z_MARK + 1); self._overlay.append(t)

    def _handles(self, pts):
        sc = self.canvas.scene(); pen = QtGui.QPen(QtGui.QColor(255, 255, 255)); pen.setCosmetic(True)
        for (vx, vy) in pts:
            it = sc.addRect(vx - 5, vy - 5, 10, 10, pen, QtGui.QBrush(QtGui.QColor(255, 180, 40)))
            it.setZValue(Z_HANDLE); self._overlay.append(it)

    def _poly(self, pts, color, width, dots=False, z=Z_MARK):
        sc = self.canvas.scene(); pen = QtGui.QPen(color, width); pen.setCosmetic(True)
        for a, b in zip(pts, pts[1:]):
            it = sc.addLine(a[0], a[1], b[0], b[1], pen); it.setZValue(z); self._overlay.append(it)
        if dots:
            for (x, y) in pts:
                it = sc.addEllipse(x - 3, y - 3, 6, 6, pen, QtGui.QBrush(color)); it.setZValue(z); self._overlay.append(it)

    def _arrow(self, a, b, color):
        ang = math.atan2(a[1] - b[1], a[0] - b[0]); L = self.leader_hpx * 0.8
        p1 = (a[0] - L * math.cos(ang - 0.4), a[1] - L * math.sin(ang - 0.4))
        p2 = (a[0] - L * math.cos(ang + 0.4), a[1] - L * math.sin(ang + 0.4))
        poly = QtGui.QPolygonF([QtCore.QPointF(*a), QtCore.QPointF(*p1), QtCore.QPointF(*p2)])
        it = self.canvas.scene().addPolygon(poly, QtGui.QPen(color), QtGui.QBrush(color)); it.setZValue(Z_MARK); self._overlay.append(it)

    # ─────────────────────────── coords ───────────────────────────
    def _to_cad(self, x, y):
        # Compuerta única: si hay georreferencia activa, píxel→UTM real; si no, escala del titleblock.
        if self.georef.active():
            return self.georef.to_world(x, y)
        return G.to_cad(x, y, self.scale, self.rot, self.W, self.H, self.derot, self.zoom)

    def _georef_base_doc(self, doc):
        """Con georreferencia activa, transforma TODO el plano base al mismo sistema
        REAL que las anotaciones. El plano base viene en coordenadas de titleblock
        (G.to_cad sin georref); ajustamos una afín base→real muestreando 3 puntos
        (base = G.to_cad(px), real = georef.to_world(px)) y la aplicamos a todas
        las entidades del modelspace. Así base + anotaciones quedan alineados."""
        from ezdxf.math import Matrix44
        pm = self.canvas.pixmap_item.pixmap()
        w, h = pm.width(), pm.height()
        samples = [(0.1 * w, 0.1 * h), (0.9 * w, 0.15 * h), (0.15 * w, 0.9 * h)]
        src = [G.to_cad(x, y, self.scale, self.rot, self.W, self.H, self.derot, self.zoom) for (x, y) in samples]
        dst = [self.georef.to_world(x, y) for (x, y) in samples]
        M, _, _ = georef_mod.fit(src, dst, "affine")
        a, b, c = M[0]; d, e, f = M[1]
        mat = Matrix44((a, d, 0, 0), (b, e, 0, 0), (0, 0, 1, 0), (c, f, 0, 1))
        msp = doc.modelspace()
        for ent in list(msp):
            try:
                ent.transform(mat)
            except Exception:
                pass                                     # entidad que no soporta transform: se deja

    def _set_geodata(self, doc):
        """Incrusta el sistema de coordenadas (NAD83 California Zona V, pies US =
        EPSG:2229) como GeoData, para que el CAD reconozca el plano geolocalizado.
        Best-effort: si algo falla, no rompe la exportación."""
        if not self.georef.active() or int(self.georef.epsg) != 2229:
            return
        try:
            from pyproj import Transformer
            b = self._plan_bbox_real(); cx = (b[0] + b[2]) / 2.0; cy = (b[1] + b[3]) / 2.0
            lon, lat = Transformer.from_crs(2229, 4326, always_xy=True).transform(cx, cy)
            gd = doc.modelspace().new_geodata()
            gd.coordinate_system_definition = "CA83VF"       # código Autodesk: NAD83 CA Zona V, pie US
            gd.dxf.design_point = (cx, cy, 0)                # punto en coords del dibujo (pies 2229)
            gd.dxf.reference_point = (lon, lat, 0)           # su lon/lat (grados)
            gd.dxf.north_direction = (0, 1)
            gd.dxf.coordinate_type = gd.PROJECTED_GRID
        except Exception:
            pass

    def _set_plan_view(self, doc):
        """Hace que el DXF se abra en vista de PLANTA (top) y encuadrado al dibujo,
        para que no aparezca como una hoja inclinada ni diminuta al abrirlo."""
        try:
            from ezdxf import bbox
            ext = bbox.extents(doc.modelspace())
            if not ext.has_data:
                return
            cx = (ext.extmin.x + ext.extmax.x) / 2.0
            cy = (ext.extmin.y + ext.extmax.y) / 2.0
            h = (ext.extmax.y - ext.extmin.y) or (ext.extmax.x - ext.extmin.x) or 100.0
            doc.set_modelspace_vport(h * 1.15, center=(cx, cy))   # vista top, centrada
        except Exception:
            pass

    def _maybe_export_dwg(self, doc, out):
        """Si el usuario lo activó, genera también un .dwg junto al .dxf usando el
        ODA File Converter (ezdxf.addons.odafc). Si ODA no está, avisa y deja el DXF."""
        if not (getattr(self, "act_dwg", None) and self.act_dwg.isChecked()):
            return
        dwg = os.path.splitext(out)[0] + ".dwg"
        try:
            from ezdxf.addons import odafc
            if not odafc.is_installed():
                QtWidgets.QMessageBox.information(self, "DWG",
                    "Para generar DWG necesitas instalar el ODA File Converter (gratuito).\n"
                    "Se guardó solo el DXF; ábrelo en tu CAD y «Guardar como DWG» si lo necesitas ahora.")
                return
            odafc.export_dwg(doc, dwg, replace=True)
            self._info(f"DWG generado: {os.path.basename(dwg)}")
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "DWG", f"No se pudo generar el DWG (se conserva el DXF).\n\n{e}")

    def _plan_bbox_real(self):
        """Recuadro del plano en coordenadas reales (del georef), desde las esquinas
        de la página. Devuelve (xmin, ymin, xmax, ymax)."""
        pm = self.canvas.pixmap_item.pixmap()
        w, h = pm.width(), pm.height()
        cs = [self.georef.to_world(x, y) for (x, y) in ((0, 0), (w, 0), (w, h), (0, h))]
        xs = [c[0] for c in cs]; ys = [c[1] for c in cs]
        return (min(xs), min(ys), max(xs), max(ys))

    def _maybe_add_la_reference(self, doc):
        """Si el usuario lo activó y el plano está georreferenciado a EPSG:2229,
        descarga de NavigateLA las calles (y opcional parcelas) del área y las
        añade como capas de referencia. Falla en silencio con aviso."""
        want_streets = getattr(self, "act_la_ref", None) and self.act_la_ref.isChecked()
        want_parcels = getattr(self, "act_la_parcels", None) and self.act_la_parcels.isChecked()
        if not (want_streets or want_parcels):
            return
        if not self.georef.active() or int(self.georef.epsg) != 2229:
            QtWidgets.QMessageBox.information(self, "Capas de LA",
                "Las capas reales de LA solo se pueden agregar si el plano está "
                "georreferenciado a EPSG:2229 (State Plane de LA)."); return
        try:
            from geo.la_reference import add_reference_layers
            nc, npa = add_reference_layers(doc, self._plan_bbox_real(),
                                           streets=bool(want_streets), parcels=bool(want_parcels))
            self._info(f"Capas de LA agregadas: {nc} tramos de calle, {npa} parcelas.")
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Capas de LA",
                f"No se pudieron descargar las capas de LA (¿internet?).\n\n{e}")

    # ─────────────────────────── exportar ───────────────────────────
    def run_pipeline(self, mode="todo"):
        """mode: 'todo' = PDF digitalizado + anotaciones · 'pdf' = solo el PDF ·
        'anot' = solo las anotaciones dibujadas en el programa."""
        if self.canvas.pixmap_item is None:
            QtWidgets.QMessageBox.information(self, "Nada", "Abre un PDF o proyecto."); return
        need_pdf = mode in ("todo", "pdf")
        if need_pdf and (not self.pdf_path or not os.path.isfile(self.pdf_path)):
            QtWidgets.QMessageBox.information(self, "Sin PDF",
                "No se encontró el PDF original. Se exportarán solo las anotaciones (utilidades, leaders, textos).")
            mode = "anot"; need_pdf = False
        base = os.path.splitext(os.path.basename(self.pdf_path))[0] if self.pdf_path else "proyecto"
        suffix = {"todo": "_completo", "pdf": "_plano", "anot": "_anotaciones"}[mode]
        out, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Guardar DXF", os.path.join(DOWNLOADS, base + suffix + ".dxf"), "DXF (*.dxf)")
        if not out: return
        self._out = out; self._mode = mode
        if not need_pdf:                                   # solo anotaciones: sin pipeline
            try:
                doc = ezdxf.new("R2010", setup=True); C.apply_imperial_header(doc)
                self._merge_into(doc, marks=True)
                self._maybe_add_la_reference(doc)        # capas reales de LA (si se activó)
                self._set_geodata(doc)                   # geolocaliza el DXF en EPSG:2229
                self._set_plan_view(doc)                 # abrir en vista 2D/planta, encuadrado
                doc.saveas(out)
                self._maybe_export_dwg(doc, out)         # además .dwg si se activó (ODA)
                QtWidgets.QMessageBox.information(self, "Listo", f"Exportado (solo anotaciones):\n{out}")
                self._info("DXF de anotaciones exportado.")
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, "Error", str(e))
            return
        self._tmp = out + ".base.tmp.dxf"
        self._prog = QtWidgets.QProgressDialog("Digitalizando el plano…", None, 0, 0, self)
        self._prog.setWindowTitle("Procesando"); self._prog.setWindowModality(QtCore.Qt.WindowModal)
        self._prog.setCancelButton(None); self._prog.show()
        self._worker = PipelineWorker(self.pdf_path, self._tmp); self._worker.done.connect(self._pipeline_done); self._worker.start()

    def _pipeline_done(self, tmp, err):
        if getattr(self, "_prog", None): self._prog.close()
        if err: QtWidgets.QMessageBox.critical(self, "Error al digitalizar", err); return
        try:
            marks = (self._mode == "todo")               # 'pdf' = solo el plano, sin anotaciones
            doc = ezdxf.readfile(tmp); C.apply_imperial_header(doc)   # reafirma imperial ($MEASUREMENT=0) tras leer el plano base
            if self.georef.active():                     # lleva el plano base al mismo sistema REAL que las anotaciones
                self._georef_base_doc(doc)
            self._merge_into(doc, marks=marks)
            self._maybe_add_la_reference(doc)            # capas reales de LA (si se activó)
            self._set_geodata(doc)                       # geolocaliza el DXF en EPSG:2229
            self._set_plan_view(doc)                     # abrir en vista 2D/planta, encuadrado
            doc.saveas(self._out)
            self._maybe_export_dwg(doc, self._out)       # además .dwg si se activó (ODA)
            if os.path.exists(tmp): os.remove(tmp)
            if marks:
                nreg = sum(1 for r in self.erase_regions if r.get("enabled", True))
                msg = (f"Exportado (PDF + anotaciones):\n{self._out}\n\n{len(self.pipes)} utilidades, "
                       f"{len(self.leaders)} leaders, {len(self.text_marks)} textos, {nreg} zonas borradas.")
            else:
                msg = f"Exportado (solo el PDF digitalizado):\n{self._out}"
            QtWidgets.QMessageBox.information(self, "Listo", msg); self._info("DXF exportado.")
        except Exception as e:
            import traceback; QtWidgets.QMessageBox.critical(self, "Error al guardar", f"{e}\n{traceback.format_exc()}")

    def _merge_into(self, doc, marks=True):
        dxf_export.merge_into(self, doc, marks=marks)

    # ─────────────────────────── Red 3D: buzones / cotas ───────────────────────────
    def insert_manhole(self):
        """Activa el modo 'clic sobre una línea existente' para insertar un buzón en
        ese punto. El buzón se materializa como un vértice extra en la polilínea (y
        _rebuild_structures lo recoge como buzón nuevo)."""
        if not self.pipes:
            self._info("No hay líneas dibujadas para insertar un buzón.")
            return
        self.set_mode("insert_bz")
        self._info("Clic sobre una línea para insertar un buzón (Esc para salir).")

    def _do_insert_manhole(self, x, y):
        from model import network_kind
        thr = 14.0 / max(1e-6, self.canvas.transform().m11())
        best = (None, -1, thr)                      # (pipe_index, seg_index, dist)
        for pi, p in enumerate(self.pipes):
            pts = p.get("pts")
            if not pts or len(pts) < 2: continue    # tramos importados de Excel (world) no editables
            if network_kind(p.get("layer") or "") == "pressure": continue   # presión no lleva buzones
            for idx, a, b in self._segments(pts, False):
                d = G.pt_seg_dist(x, y, a[0], a[1], b[0], b[1])
                if d < best[2]: best = (pi, idx, d)
        if best[0] is None:
            self._info("Los buzones/cajas solo se insertan en redes de gravedad o conduit (no en presión).")
            self.set_mode("idle"); return
        self._push()
        pi, si, _ = best
        self.pipes[pi]["pts"].insert(si + 1, (x, y))
        self._rebuild_structures()
        self._refresh_lists(); self._redraw()
        self._info("Buzón insertado. Edítalo en la tab Buzones.")
        self.set_mode("idle")

    def _rebuild_structures(self):
        """Detecta buzones por los VÉRTICES (extremos + intermedios) de las tuberías
        dibujadas:
          - Gravedad (SS/SD) → prefijo BZ- (buzones cilíndricos con tapa).
          - Conduit (eléctrico/telecom) → prefijo CAJA- (cajas de registro/vaults).
          - Presión (agua/gas) → sin nodos automáticos.
        Preserva ediciones (cod/rim/sump/part/part_size/covered) por coincidencia
        de coordenada. Los buzones importados de Excel (world) se conservan aparte."""
        from model import network_kind
        tol = 14.0
        def near(a, b): return math.hypot(a[0] - b[0], a[1] - b[1]) <= tol
        # Descarta buzones espurios de versiones previas con net inválida (p.ej. "pressure").
        old = [s for s in self.structures
               if not s.get("world") and (s.get("net") or "gravity") in ("gravity", "conduit")]
        world = [s for s in self.structures if s.get("world")]
        detected = []
        for p in self.pipes:
            if p.get("world"): continue
            kind = network_kind(p.get("layer") or "")
            if kind not in ("gravity", "conduit"): continue    # presión no lleva nodos automáticos
            pts = p.get("pts")
            if not pts or len(pts) < 2: continue
            for pt in pts:                              # todos los vértices (extremos + intermedios)
                if not any(near(pt, (s["x"], s["y"])) for s in detected):
                    detected.append({"cod": "", "x": pt[0], "y": pt[1], "rim": None,
                                     "sump": None, "part": "", "part_size": "",
                                     "net": kind, "covered": True, "world": False})
        for s in detected:                                 # reasigna ediciones previas por coordenada
            for o in old:
                if near((s["x"], s["y"]), (o.get("x", -1e9), o.get("y", -1e9))):
                    s.update(cod=o.get("cod", ""), rim=o.get("rim"), sump=o.get("sump"),
                             part=o.get("part", ""), part_size=o.get("part_size", ""),
                             covered=bool(o.get("covered", True))); break
        # Códigos únicos: BZ-N para gravedad, CAJA-N para conduit.
        used = {s.get("cod", "") for s in world + detected if s.get("cod")}
        cnt_bz = cnt_caja = 1
        for s in detected:
            if s.get("cod"): continue
            prefix = "CAJA-" if s.get("net") == "conduit" else "BZ-"
            if prefix == "BZ-":
                while f"BZ-{cnt_bz}" in used: cnt_bz += 1
                s["cod"] = f"BZ-{cnt_bz}"; used.add(s["cod"]); cnt_bz += 1
            else:
                while f"CAJA-{cnt_caja}" in used: cnt_caja += 1
                s["cod"] = f"CAJA-{cnt_caja}"; used.add(s["cod"]); cnt_caja += 1
        self.structures = world + detected; self._dirty = True

    # ── Tab Buzones: selección, panel de propiedades y edición ──────────────
    def _sel_bz(self, row):
        """La selección en la lista de buzones cambió: sincroniza panel + canvas."""
        if row < 0 or row >= len(self.structures):
            self.sel_bz = -1
        else:
            self.sel_bz = row
            s = self.structures[row]
            if not s.get("world") and not self._no_center:
                x, y = s.get("x"), s.get("y")
                if x is not None and y is not None:
                    self.canvas.centerOn(float(x), float(y))
        self._sync_bz_panel(); self._update_ui(); self._redraw()

    def _sync_bz_panel(self):
        """Carga los valores del buzón self.sel_bz en el panel de propiedades."""
        if not hasattr(self, "gprop_bz"): return
        import civil_catalog as _cc
        self._bz_prop_guard = True
        try:
            in_tab = self._current_tab() == TAB_BZ
            self.gprop_bz.setVisible(in_tab)
            has_sel = 0 <= self.sel_bz < len(self.structures)
            # Habilitar/deshabilitar todos los controles del groupbox según haya selección
            for w in (self.bz_cod, self.bz_rim, self.bz_sump, self.bz_family, self.bz_size,
                      self.bz_cover, self.btn_bz_addsize):
                w.setEnabled(has_sel)
            if not has_sel:
                self.gprop_bz.setTitle("Propiedades del buzón — selecciona uno de la lista")
                return
            s = self.structures[self.sel_bz]
            net = s.get("net") or "gravity"
            self.gprop_bz.setTitle("Propiedades de la caja" if net == "conduit"
                                    else "Propiedades del buzón")
            self.bz_cod.setText(s.get("cod", ""))
            self.bz_rim.setValue(float(s.get("rim") or 0.0))
            self.bz_sump.setValue(float(s.get("sump") or 0.0))
            self.bz_cover.setCurrentIndex(0 if s.get("covered", True) else 1)
            self.bz_net_lbl.setText("conduit (eléctrico/telecom)" if net == "conduit" else "gravedad")
            self.bz_origin_lbl.setText("Excel" if s.get("world") else "dibujo")
            # Familias del catálogo imperial de estructuras (gravedad).
            self.bz_family.blockSignals(True); self.bz_family.clear()
            fams = _cc.imperial_structures(self.civil_year) if self.civil_year else []
            self.bz_family.addItem("(por defecto)", "")
            for f in fams:
                idx = self.bz_family.count()
                self.bz_family.addItem(f"{f['pretty']}  [{f['subfolder']}]", f["id"])
                img = f.get("img_path")
                tip = f"<b>{f['pretty']}</b><br><i>{f['subfolder']}</i>"
                if img:
                    tip += f"<br><img src='file:///{img.replace(chr(92), '/')}' width='220'>"
                self.bz_family.setItemData(idx, tip, QtCore.Qt.ToolTipRole)
            cur_fid = s.get("part", "") or ""
            for i in range(self.bz_family.count()):
                if self.bz_family.itemData(i) == cur_fid:
                    self.bz_family.setCurrentIndex(i); break
            self.bz_family.blockSignals(False)
            self._load_bz_sizes(cur_fid, s.get("part_size", "") or "")
            self.btn_bz_addsize.setEnabled(bool(cur_fid))
        finally:
            self._bz_prop_guard = False

    def _load_bz_sizes(self, fid, current):
        """Repuebla self.bz_size según la familia (siempre catálogo de gravedad)."""
        import civil_catalog as _cc
        self.bz_size.blockSignals(True); self.bz_size.clear()
        if not fid or not self.civil_year:
            self.bz_size.addItem("(sin familia)", ""); self.bz_size.setEnabled(False)
            self.bz_size.blockSignals(False); return
        sizes = _cc.structure_sizes(self.civil_year, fid)
        if not sizes:
            self.bz_size.addItem("(sin tamaños detectados)", ""); self.bz_size.setEnabled(False)
        else:
            self.bz_size.setEnabled(True); self.bz_size.addItem("(por defecto)", "")
            for sz in sizes: self.bz_size.addItem(sz, sz)
            if current:
                for i in range(self.bz_size.count()):
                    if self.bz_size.itemData(i) == current:
                        self.bz_size.setCurrentIndex(i); break
        self.bz_size.blockSignals(False)

    def _bz_family_changed(self, _idx):
        if self._bz_prop_guard: return
        if not (0 <= self.sel_bz < len(self.structures)): return
        s = self.structures[self.sel_bz]
        fid = self.bz_family.currentData() or ""
        s["part"] = fid; s["part_size"] = ""      # al cambiar familia se resetea el tamaño
        self._load_bz_sizes(fid, "")
        self.btn_bz_addsize.setEnabled(bool(fid))
        self._dirty = True
        self._refresh_bz_list_item(self.sel_bz)
        self._redraw()

    def _bz_prop_changed(self):
        if self._bz_prop_guard: return
        if not (0 <= self.sel_bz < len(self.structures)): return
        s = self.structures[self.sel_bz]
        cod_new = self.bz_cod.text().strip()
        if cod_new and cod_new != s.get("cod", ""):
            # Validar unicidad
            if any(o.get("cod") == cod_new for i, o in enumerate(self.structures) if i != self.sel_bz):
                QtWidgets.QMessageBox.warning(self, "Código repetido",
                    f"Ya existe un buzón con código '{cod_new}'. Elige otro.")
                self._bz_prop_guard = True; self.bz_cod.setText(s.get("cod", "")); self._bz_prop_guard = False
                return
            s["cod"] = cod_new
        s["rim"] = float(self.bz_rim.value()) if self.bz_rim.value() != 0.0 else s.get("rim")
        s["sump"] = float(self.bz_sump.value()) if self.bz_sump.value() != 0.0 else s.get("sump")
        # Si el usuario dejó los spins en 0.0 pero el valor original era 0 o None, respetar 0.
        s["rim"] = float(self.bz_rim.value()); s["sump"] = float(self.bz_sump.value())
        if self.bz_size.isEnabled():
            s["part_size"] = self.bz_size.currentData() or ""
        s["covered"] = (self.bz_cover.currentIndex() == 0)
        self._dirty = True
        self._refresh_bz_list_item(self.sel_bz)
        self._redraw()

    def _refresh_bz_list_item(self, row):
        if not (0 <= row < len(self.structures)): return
        s = self.structures[row]
        fam = s.get("part") or "(sin familia)"
        sz = f"  {s['part_size']}" if s.get("part_size") else ""
        emoji = "🟠" if s.get("net") == "conduit" else "🔵"
        item = self.bz_list.item(row)
        if item: item.setText(f"{emoji} {s.get('cod', '?')}  ·  {fam}{sz}")

    def _bz_add_custom_size_current(self):
        if not (0 <= self.sel_bz < len(self.structures)): return
        s = self.structures[self.sel_bz]
        fid = s.get("part") or ""
        if not fid:
            QtWidgets.QMessageBox.information(self, "Sin familia",
                "Elige primero una familia para agregarle un tamaño."); return
        self._add_structure_size_dialog(self, fid, self.bz_family, self.bz_size)
        # Después de agregar, guardo el nuevo tamaño en el buzón actual.
        self._bz_prop_changed()

    def _add_structure_size_dialog(self, parent, fid, cbf, cbs, kind="structure"):
        """Abre un diálogo con un campo por cada parámetro de la familia (leído del
        XML del catálogo). Al aceptar, escribe los nuevos <Item> en el .xml y refresca
        los combos de familia (todos los que apunten a fid) y el combo de tamaño.
        `kind` es 'structure' o 'pipe' (Bancoductos/Bancos Tubos usan el mismo XML)."""
        import civil_catalog as _cc
        params = _cc.family_params(self.civil_year, fid, kind)
        if not params:
            QtWidgets.QMessageBox.warning(parent, "Sin parámetros",
                f"No pude leer parámetros del catálogo para {fid}."); return
        dlg = QtWidgets.QDialog(parent); dlg.setWindowTitle("Agregar tamaño personalizado"); dlg.resize(560, 420)
        v = QtWidgets.QVBoxLayout(dlg)
        v.addWidget(QtWidgets.QLabel(
            f"<b>{fid}</b><br>Ingresa el valor nuevo para cada parámetro. Deja en blanco "
            f"los que quieras dejar en su valor por defecto (primer valor existente).<br>"
            f"El tamaño quedará guardado en el catálogo Civil 3D y podrá reutilizarse en otros proyectos."))
        form = QtWidgets.QFormLayout(); v.addLayout(form)
        edits = {}
        for p in params:
            unit = p["unit"] or ""
            label = f"{p['desc']} ({p['name']}, {unit})" if unit else f"{p['desc']} ({p['name']})"
            le = QtWidgets.QLineEdit()
            hint = ", ".join(sorted({f"{float(x):g}" for x in p["items"]},
                                     key=lambda s: float(s))) if p["items"] else ""
            if hint: le.setPlaceholderText(f"existentes: {hint}")
            form.addRow(label, le); edits[p["name"]] = le
        bb = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        bb.accepted.connect(dlg.accept); bb.rejected.connect(dlg.reject); v.addWidget(bb)
        if dlg.exec() != QtWidgets.QDialog.Accepted: return
        values = {name: le.text().strip() for name, le in edits.items() if le.text().strip()}
        if not values:
            QtWidgets.QMessageBox.information(parent, "Nada que agregar",
                "No ingresaste ningún valor."); return
        res = _cc.add_family_size(self.civil_year, fid, values, kind)
        if not res.get("ok"):
            QtWidgets.QMessageBox.critical(parent, "Error", res.get("error", "?")); return
        added = res.get("added") or {}
        skipped = res.get("skipped") or {}
        msg = []
        if added:
            msg.append("Agregado al catálogo:\n  · " +
                       "\n  · ".join(f"{k} = {v}" for k, v in added.items()))
        if skipped:
            msg.append("Omitido:\n  · " +
                       "\n  · ".join(f"{k}: {v}" for k, v in skipped.items()))
        QtWidgets.QMessageBox.information(parent, "Catálogo actualizado", "\n\n".join(msg))
        # Refrescar el combo de tamaño de la fila actual (si nos pasaron uno).
        if cbs is not None:
            sizes = (_cc.pipe_sizes(self.civil_year, fid) if kind == "pipe"
                     else _cc.structure_sizes(self.civil_year, fid))
            cbs.blockSignals(True); cbs.clear()
            if sizes:
                cbs.setEnabled(True); cbs.addItem("(por defecto)", "")
                for sz in sizes: cbs.addItem(sz, sz)
            else:
                cbs.setEnabled(False); cbs.addItem("(sin tamaños detectados)", "")
            cbs.blockSignals(False)

    # ─────────────────────────── Editor de Catálogo Civil 3D ───────────────────────────
    def open_catalog_editor(self):
        """Diálogo para navegar el catálogo (US Imperial Pipes + Structures) de la
        versión Civil 3D actual, ver los tamaños de cada familia y agregar tamaños
        nuevos. Edita directamente los XML del catálogo (con backup .xml.bak)."""
        import civil_catalog as _cc
        if self.civil_year is None:
            QtWidgets.QMessageBox.warning(self, "Sin Civil 3D",
                "No detecté ninguna instalación de Civil 3D."); return

        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle(f"Editor de catálogo — Civil 3D {self.civil_year}")
        dlg.resize(880, 560)
        lay = QtWidgets.QHBoxLayout(dlg)

        # Árbol izquierdo: agrupado por subcarpeta.
        tree = QtWidgets.QTreeWidget()
        tree.setHeaderLabels(["Familia"])
        tree.setMinimumWidth(360)
        lay.addWidget(tree, 1)

        pipes = _cc.imperial_pipes(self.civil_year)
        structs = _cc.imperial_structures(self.civil_year)
        # Agrupar por subfolder, marcando kind en el item para saber a qué XML apuntar.
        groups = {}
        for f in pipes:
            groups.setdefault(("pipe", f["subfolder"]), []).append(f)
        for f in structs:
            groups.setdefault(("structure", f["subfolder"]), []).append(f)
        for (kind, sub), fams in sorted(groups.items(), key=lambda kv: (kv[0][0], kv[0][1])):
            head = QtWidgets.QTreeWidgetItem([f"{sub}  ({'tuberías' if kind=='pipe' else 'estructuras'})"])
            head.setFlags(QtCore.Qt.ItemIsEnabled)
            tree.addTopLevelItem(head)
            for f in fams:
                it = QtWidgets.QTreeWidgetItem([f["pretty"]])
                it.setData(0, QtCore.Qt.UserRole, (kind, f["id"]))
                head.addChild(it)
            head.setExpanded(True)

        # Panel derecho: detalles + tamaños actuales + botón agregar.
        right = QtWidgets.QWidget(); rv = QtWidgets.QVBoxLayout(right)
        lbl_fam = QtWidgets.QLabel("<i>Selecciona una familia a la izquierda.</i>")
        lbl_fam.setWordWrap(True); rv.addWidget(lbl_fam)
        rv.addWidget(QtWidgets.QLabel("Tamaños actuales:"))
        sizes_list = QtWidgets.QListWidget(); rv.addWidget(sizes_list, 1)
        btn_row = QtWidgets.QHBoxLayout()
        btn_add = QtWidgets.QPushButton("Agregar tamaño…"); btn_add.setEnabled(False)
        btn_row.addWidget(btn_add); btn_row.addStretch(1)
        rv.addLayout(btn_row)
        rv.addWidget(QtWidgets.QLabel(
            "<i>Los tamaños agregados quedan en el XML del catálogo Civil 3D "
            "(<code>ProgramData\\Autodesk\\C3D &lt;año&gt;\\...</code>). Para "
            "que aparezcan en un dibujo, ábrelo en Civil 3D y añade la familia "
            "a la Parts List (o usa el comando <code>AGREGAR_BANCOS_Y_BUZONES</code>).</i>"))
        lay.addWidget(right, 1)

        current = {"kind": None, "fid": None}

        def _refresh_sizes():
            sizes_list.clear()
            k, fid = current["kind"], current["fid"]
            if not fid: return
            sz = (_cc.pipe_sizes(self.civil_year, fid) if k == "pipe"
                  else _cc.structure_sizes(self.civil_year, fid))
            if not sz:
                sizes_list.addItem("(sin tamaños detectados)")
            else:
                for s in sz: sizes_list.addItem(s)

        def _on_sel():
            items = tree.selectedItems()
            if not items:
                current["kind"] = current["fid"] = None
                lbl_fam.setText("<i>Selecciona una familia a la izquierda.</i>")
                btn_add.setEnabled(False); sizes_list.clear(); return
            it = items[0]; data = it.data(0, QtCore.Qt.UserRole)
            if not data:
                current["kind"] = current["fid"] = None
                lbl_fam.setText("<i>Selecciona una familia (no un grupo).</i>")
                btn_add.setEnabled(False); sizes_list.clear(); return
            k, fid = data
            current["kind"] = k; current["fid"] = fid
            lbl_fam.setText(f"<b>{it.text(0)}</b><br>"
                            f"Tipo: {'tubería' if k == 'pipe' else 'estructura'} · <code>{fid}</code>")
            params = _cc.family_params(self.civil_year, fid, k)
            btn_add.setEnabled(bool(params))
            if not params and k == "pipe":
                lbl_fam.setText(lbl_fam.text() + "<br><i>Esta familia usa parámetros continuos "
                                                 "(no lista fija) — no editable desde aquí.</i>")
            _refresh_sizes()

        def _on_add():
            k, fid = current["kind"], current["fid"]
            if not fid: return
            self._add_structure_size_dialog(dlg, fid, None, None, kind=k)
            _refresh_sizes()

        tree.itemSelectionChanged.connect(_on_sel)
        btn_add.clicked.connect(_on_add)

        bb = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Close)
        bb.rejected.connect(dlg.reject); bb.accepted.connect(dlg.accept)
        rv.addWidget(bb)
        dlg.exec()

    def import_network_excel(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Importar Excel de red", DOWNLOADS, "Excel (*.xlsx *.xlsm)")
        if not path: return
        self._busy("Leyendo Excel de red…")
        try:
            nets, warns = excel_import.read_network_workbook(path)
        except Exception as e:
            self._unbusy(); QtWidgets.QMessageBox.critical(self, "Error al leer Excel", str(e)); return
        self._unbusy(); self._push()
        # reemplaza lo importado previamente (world); conserva lo dibujado
        self.structures = [s for s in self.structures if not s.get("world")]
        self.pipes = [p for p in self.pipes if not p.get("world")]
        n_s = n_p = 0
        # El Excel del cliente viene en METROS (estándar topográfico); pasamos
        # coordenadas, cotas y diámetros a la unidad de trabajo del proyecto
        # (ft o in) para que TODO en el modelo esté en la misma unidad que el JSON.
        u = self.work_unit
        def to_u(v): return G.convert_length(v, "m", u)
        for name, nd in nets.items():
            for s in nd["structures"]:
                self.structures.append({
                    "cod": s["cod"], "x": to_u(s["x"]), "y": to_u(s["y"]),
                    "rim": to_u(s["rim"]), "sump": to_u(s["sump"]),
                    "part": s["part"], "net": name, "world": True}); n_s += 1
            for pp in nd["pipes"]:
                self.pipes.append({
                    "layer": "ALCANTARILLADO", "pts": [], "world": True, "net": name,
                    "wstart": (to_u(pp["xi"]), to_u(pp["yi"])),
                    "wend":   (to_u(pp["xf"]), to_u(pp["yf"])),
                    "from": pp["from"], "to": pp["to"], "id": pp["id"], "name": pp["id"],
                    "inv_start": to_u(pp["zi"]), "inv_end": to_u(pp["zf"]),
                    "diam": to_u(pp["diam"]) or 0.0, "unit": u,
                    "part": pp["part"]}); n_p += 1
        self._refresh_lists(); self._redraw()
        msg = f"Importado: {len(nets)} red(es), {n_s} buzones, {n_p} tuberías."
        if warns: msg += "\n\nAvisos:\n- " + "\n- ".join(warns[:12])
        QtWidgets.QMessageBox.information(self, "Importación de red", msg); self._info(msg.split(chr(10))[0])

    # ─────────────────────────── Georreferenciación ───────────────────────────
    def clear_georef(self):
        if not self.georef.active():
            self._info("El plano no está georreferenciado."); return
        self._dirty = True; self.georef = georef_mod.Georef()
        self._update_geo_status(); self._info("Georreferencia quitada; se usa la escala del titleblock.")

    def open_georef(self):
        if self.canvas.pixmap_item is None:
            QtWidgets.QMessageBox.information(self, "Sin plano", "Abre un PDF o proyecto primero."); return
        try:
            from geo.georef_dialog import GeorefDialog
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Falta un componente",
                "La georreferenciación necesita PySide6-WebEngine, pyproj y scikit-image.\n\n"
                "Instálalos con tu Python 3.12:\n"
                r"  C:\Users\Deyvy\AppData\Local\Programs\Python\Python312\python.exe -m pip install "
                "PySide6-WebEngine pyproj scikit-image" f"\n\nDetalle: {e}")
            return
        img = self.canvas.pixmap_item.pixmap().toImage()
        dlg = GeorefDialog(self, img, self.georef)
        if dlg.exec() == QtWidgets.QDialog.Accepted and dlg.result_georef is not None:
            self.georef = dlg.result_georef; self._dirty = True
            self._update_geo_status(); self._redraw()
            unit = georef_mod.epsg_unit(self.georef.epsg)
            rms = self.georef.rms if self.georef.rms is not None else 0.0
            self._info(f"Georreferenciado (EPSG:{self.georef.epsg}, RMS {rms:.2f} {unit}). "
                       "Las coordenadas exportadas ahora son reales.")
            QtWidgets.QMessageBox.information(
                self, "Georreferenciación",
                f"✓ Plano georreferenciado\n\nEPSG: {self.georef.epsg}\nRMS: {rms:.2f} {unit}")

    # ─────────────────────────── Excel ───────────────────────────
    def open_excel(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Abrir Excel", DOWNLOADS, "Excel (*.xlsx *.xlsm)")
        if path: self._read_excel(path)

    def _read_excel(self, path):
        self._busy("Leyendo Excel…")
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True); texts = []
            for ws in wb.worksheets:
                try:
                    hc = hr = None
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value is not None and str(cell.value).strip().upper() in ("TEXTO", "TEXTOS"):
                                hc, hr = cell.column, cell.row; break
                        if hc: break
                    if hc is None: continue
                    for row in ws.iter_rows(min_row=hr + 1, min_col=hc, max_col=hc):
                        val = row[0].value
                        if val is not None and str(val).strip(): texts.append(str(val).strip())
                except Exception: continue
            self.text_list.clear(); self.text_list.addItems(texts)
            self._info(f"Excel: {len(texts)} textos de la columna TEXTO." if texts else "No hallé columna TEXTO/TEXTOS.")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error al leer Excel", str(e))
        finally: self._unbusy()

    # ─────────────────────────── Ayuda ───────────────────────────
    def _show_html(self, title, html, w=780, h=660):
        dlg = QtWidgets.QDialog(self); dlg.setWindowTitle(title); dlg.resize(w, h)
        lay = QtWidgets.QVBoxLayout(dlg); tb = QtWidgets.QTextBrowser(); tb.setOpenExternalLinks(True)
        tb.setStyleSheet("background:#1e1e1e;color:#e8e8e8;font-size:14px;"); tb.setHtml(html)
        btn = QtWidgets.QPushButton("Cerrar"); btn.clicked.connect(dlg.accept)
        lay.addWidget(tb); lay.addWidget(btn); dlg.exec()

    def show_about(self):
        dlg = QtWidgets.QDialog(self); dlg.setWindowTitle("Acerca de"); dlg.resize(760, 680)
        lay = QtWidgets.QVBoxLayout(dlg)
        head = QtWidgets.QLabel(
            f"<h2>Digitalizador de planos — utilidades</h2>"
            f"<p><b>Versión {VERSION}</b> · para ingeniería civil (agua, alcantarillado, gas, "
            f"eléctrico, telefonía, drenaje).</p>"
            f"<p>Convierte un PDF de plano a DXF y te deja marcar utilidades, Multileaders y notas "
            f"sobre la imagen, exportando todo en las mismas coordenadas.</p>"
            f"<p style='color:#888;'>Historial de versiones (la más reciente arriba, la más antigua "
            f"abajo). Haz clic en cada versión para desplegarla.</p>")
        head.setWordWrap(True); head.setStyleSheet("color:#e8e8e8;"); lay.addWidget(head)
        box = QtWidgets.QToolBox()
        box.setStyleSheet("QToolBox::tab{background:#333;color:#ddd;border:1px solid #555;}"
                          "QToolBox::tab:selected{background:#3c5a99;color:white;font-weight:bold;}")
        icon = {"added": ("#5fd35f", "✚ nueva"), "removed": ("#e06060", "✖ quitada"),
                "fixed": ("#6cc5e0", "✎ corregida"), "changed": ("#e0c060", "↻ cambiada"),
                "base": ("#cfcfcf", "•")}
        default = ("#cfcfcf", "•")
        for ver, items in CHANGELOG:
            tb = QtWidgets.QTextBrowser(); tb.setStyleSheet("background:#1e1e1e;color:#e8e8e8;border:none;")
            lis = "".join(f'<li style="color:{icon.get(s, default)[0]};margin-bottom:4px;">'
                          f'<b>[{icon.get(s, default)[1]}]</b> {t}</li>' for s, t in items)
            tb.setHtml(f"<ul>{lis}</ul>"); box.addItem(tb, f"v{ver}")
        lay.addWidget(box, 1)
        btn = QtWidgets.QPushButton("Cerrar"); btn.clicked.connect(dlg.accept); lay.addWidget(btn)
        dlg.exec()

    def show_manual(self):
        """Ventana del manual de usuario. Es HTML sencillo dentro de un
        QTextBrowser (visor de texto enriquecido); nada de red ni servidor."""
        html = """
        <h2>Manual rapido</h2>
        <p><i>App para marcar planos PDF (agua, alcantarillado, drenaje, gas, electricidad,
        telefonia) y exportar a AutoCAD/Civil 3D via DXF con la red 3D como XDATA.</i></p>

        <h3>1. Abrir plano</h3>
        <p><b>Archivo -> Abrir PDF...</b> (o arrastralo). Cambia de pagina y ajusta
        transparencia en la seccion <b>Vista y paginas</b>. Rueda = zoom, boton central = pan.</p>

        <h3>2. Dibujar utilidad</h3>
        <p>Acordeon <b>Dibujar utilidad</b> -> elige el tipo -> <b>Dibujar utilidad</b> ->
        clic para cada vertice, <b>Enter</b> finaliza. Con "Imán" activo se pega a lineas del PDF.</p>

        <h3>3. Propiedades de la tuberia</h3>
        <p>Selecciona la utilidad en el inventario. En <b>Propiedades</b> pon:
        <b>Diametro</b> (lista fija en pulgadas del catalogo imperial),
        <b>Elev. de rasante inicial/final</b> (pies),
        <b>Material</b> (lista fija que matchea Civil 3D). Sin esto la red 3D queda vacia.</p>

        <h3>4. Leader (flecha)</h3>
        <p>Acordeon <b>Leader</b> -> orientacion (H/V/D) -> clic en la cabeza -> clic en el final
        del cuerpo (2 clics H/V, 3 clics D).</p>

        <h3>5. Texto libre</h3>
        <p>Acordeon <b>Texto libre</b>. Clic donde escribir, <b>Enter</b> aplica,
        <b>Ctrl+Shift+Enter</b> = salto de linea. Editable con doble clic.</p>

        <h3>6. Borrar zona / MEMBRETE</h3>
        <p>Acordeon <b>Borrar zona</b>: polilinea, Enter cierra. Al exportar borra el plano
        dentro de la zona. Ademas, el cajetin/marco del PDF se detecta automatico y va a la
        capa <b>MEMBRETE</b> (congelable en CAD).</p>

        <h3>7. Buzones y red 3D</h3>
        <p><b>Herramientas -> Gestionar buzones...</b> Los buzones se crean automatico en los
        vertices de tuberias de gravedad (alcantarillado, drenaje). El plugin de Civil 3D
        (<code>IMPORTAR_RED</code>) los crea con estilo cilindrico y catalogo imperial.
        Tambien puedes <b>Importar Excel de red</b> con hojas BUZONES y TUBERIAS.</p>

        <h3>8. Georreferenciacion (opcional)</h3>
        <p><b>Herramientas -> Georreferenciar...</b> Solo si necesitas coordenadas UTM aproximadas
        para anteproyecto. El dato topografico real viene del levantamiento.</p>

        <h3>9. Guardar y exportar</h3>
        <ul>
          <li><b>Ctrl+S</b> guarda el proyecto <code>.digproj</code> (recuperable con todo lo marcado).</li>
          <li><b>Exportar DXF</b> genera el DXF completo (PDF digitalizado + tus marcados + red 3D
              en XDATA). Se abre luego en Civil 3D con el comando <code>IMPORTAR_RED</code>.</li>
        </ul>
        """
        self._show_html("Manual de usuario", html, 880, 780)

    def show_shortcuts(self):
        rows = [("Ctrl+Z / Ctrl+Shift+Z", "Deshacer / Rehacer"),
                ("Enter", "Aplicar: finaliza utilidad/zona, o agrega texto/edición"),
                ("Ctrl+Shift+Enter", "Salto de línea dentro de un texto"),
                ("Escape", "Quitar la selección; si no hay, salir del modo"),
                ("Ctrl+T", "Editar/mover lo seleccionado"),
                ("Ctrl+S / Ctrl+Shift+S", "Guardar proyecto / Guardar como…"),
                ("Ctrl+W", "Cerrar proyecto (pregunta si hay cambios)"),
                ("Doble clic", "Sobre un texto: editarlo"),
                ("Clic derecho", "Finaliza línea/zona; en editar, elimina el vértice"),
                ("◀ ▶", "Página anterior / siguiente"),
                ("Rueda", "Zoom · Botón central + arrastrar: desplazar")]
        body = "".join(f'<tr><td style="padding:4px 14px;color:#8bd;"><b>{k}</b></td>'
                       f'<td style="padding:4px;">{d}</td></tr>' for k, d in rows)
        self._show_html("Atajos de teclado", f"<h2>Atajos de teclado</h2><table>{body}</table>", 640, 500)

    # ─────────────────────────── drag & drop ───────────────────────────
    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls(): e.acceptProposedAction()

    def dropEvent(self, e):
        for u in e.mimeData().urls(): self.open_path(u.toLocalFile()); break


def main():
    # Necesario para QWebEngineView (mapa de georreferenciación); inofensivo si no se usa.
    QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_ShareOpenGLContexts)
    app = QtWidgets.QApplication(sys.argv)
    win = Main(); win.show()
    if len(sys.argv) > 1:
        win.open_path(sys.argv[1])
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
