"""
marcar_utilidades.py — Digitalizar planos y marcar utilidades.

App de escritorio (PySide6) para ingeniería civil: abre el PDF de un plano,
marca utilidades, coloca Multileaders con nomenclatura, escribe texto libre,
borra zonas y exporta un único DXF (base digitalizada + tu marcado) en las
mismas coordenadas. Ver menú Ayuda → Manual de usuario.
"""
import sys, os, copy, math, json, zipfile
import fitz
import numpy as np
import ezdxf
from ezdxf.enums import TextEntityAlignment
from PySide6 import QtCore, QtGui, QtWidgets

import config as C
import vector_pipeline as VP

VERSION = "0.6.7"

TIPOS = [
    ("Agua (W)", "AGUA"), ("Alcantarillado (SS)", "ALCANTARILLADO"),
    ("Drenaje (SD)", "DRENAJE"), ("Gas (G)", "GAS"),
    ("Eléctrico (E)", "ELECTRICO"), ("Eléctrico aéreo E(OH)", "ELECTRICO_AEREO"),
    ("Telefonía (T)", "TELECOM"), ("Telefonía aérea T(OH)", "TELECOM_AEREO"),
]
ACI_RGB = {1: (255, 60, 60), 2: (235, 215, 40), 3: (60, 210, 60), 4: (60, 210, 210),
           5: (90, 140, 255), 6: (230, 90, 230), 7: (235, 235, 235), 8: (150, 150, 150),
           30: (255, 150, 40)}
LEADER_TEXT_FT = 3.0
DOWNLOADS = os.path.join(os.path.expanduser("~"), "Downloads")
LEADER_ORIENT = [("h", "Horizontal"), ("v", "Vertical"), ("d", "Diagonal")]
BTN_ON = "background:#2e9e4f;color:white;font-weight:bold;padding:8px;border-radius:4px;"
BTN_OFF = "background:#3c5a99;color:white;padding:8px;border-radius:4px;"
Z_PDF, Z_ERASE, Z_MARK, Z_HANDLE = 0, 1, 5, 8
# Índices de las pestañas del inventario (derecha)
TAB_PIPE, TAB_ML, TAB_LEADER, TAB_TEXT, TAB_REGION = 0, 1, 2, 3, 4

CHANGELOG = [
    ("0.6.7", [
        ("changed", "Al exportar a DXF, cada Multileader se escribe como entidad MULTILEADER nativa (flecha + directriz + texto); si el visor no lo soporta, se dibuja explícito como respaldo."),
        ("fixed", "Ayuda → 'Acerca de…' ya no aparecía en blanco (fallaba al mostrar entradas de tipo 'changed' del historial)."),
    ]),
    ("0.6.6", [
        ("added", "Con un Leader seleccionado, Ctrl+T entra directo a editar sus vértices (posición/longitud)."),
        ("changed", "Al exportar a DXF, los Leaders simples se escriben como entidad LEADER nativa (con punta de flecha) siguiendo sus vértices; si el visor no lo soporta, se dibujan explícitos como respaldo."),
    ]),
    ("0.6.5", [
        ("added", "Pestaña de inventario 'Leaders' independiente de 'Multileaders'."),
        ("added", "Editar Leaders (horizontal/vertical/diagonal): selecciona el Leader, pulsa 'Editar/mover' y arrastra sus vértices (posición y longitud) o el trazo completo. Enter/Esc termina la edición."),
    ]),
    ("0.6.4", [
        ("added", "Leader diagonal con landing (3 clics): cabeza de flecha → inicio del landing (bisagra) → final del cuerpo."),
        ("added", "En modo Leader, Enter finaliza el comando (además de Esc)."),
    ]),
    ("0.6.3", [
        ("added", "'Colocar Leader' tiene orientación (Horizontal/Vertical/Diagonal): 1er clic = cabeza de flecha, 2º clic = final del cuerpo. En Horizontal/Vertical el cuerpo se mantiene recto al eje."),
        ("changed", "En modo Leader el panel se titula 'Leader' y muestra solo la Orientación; se oculta el estilo de texto (el Leader no lleva texto)."),
    ]),
    ("0.6.2", [
        ("changed", "'Colocar Leader' ahora coloca solo la flecha, sin texto: 1er clic = cabeza de flecha, 2º clic = final del cuerpo."),
        ("fixed", "El Leader simple ya no se mezclaba con el Multileader al encadenar colocaciones (se perdía el modo)."),
    ]),
    ("0.6.1", [
        ("added", "El botón Exportar DXF es un desplegable: PDF + anotaciones, solo el PDF, o solo las anotaciones."),
        ("added", "Nueva herramienta 'Colocar Leader': una directriz recta simple (punta → texto), aparte del Multileader."),
    ]),
    ("0.6.0", [
        ("fixed", "El plano base ya NO se restilea: el linetype con letra (─ W ─) se aplica solo a la utilidad que dibujas (por entidad), no a la capa."),
        ("fixed", "El resaltador/anotaciones del PDF (rellenos semitransparentes) ya no se digitalizan."),
        ("fixed", "El Multileader se digitaliza como geometría exacta (línea + punta + texto) agrupada, igual que en la vista previa: el texto y la cola ya salen bien."),
    ]),
    ("0.5.2", [
        ("fixed", "El texto del Multileader vertical ahora queda pegado a la línea (antes salía muy separado)."),
        ("fixed", "Al digitalizar, la entidad MULTILEADER sigue exactamente la línea dibujada (no reencamina las diagonales/verticales)."),
    ]),
    ("0.5.1", [
        ("added", "El tipo de utilidad ahora es un menú desplegable (con su color), más compacto."),
    ]),
    ("0.5.0", [
        ("added", "El Multileader tiene estilo de texto propio (fuente, altura, negrita) editable, y se exporta a CAD con ese estilo."),
        ("added", "Extender vértice sin ventana emergente: casilla 'continuar la misma utilidad' (si es un extremo) o rama nueva."),
        ("added", "Cada utilidad tiene propiedades: nombre, diámetro y unidad (pies/pulgadas), que se exportan como dato (XDATA) al DXF."),
        ("fixed", "Multileader horizontal/vertical: la cola se adapta al largo del texto y el texto queda bien ubicado respecto a la línea y la flecha."),
        ("fixed", "El Multileader se ve al instante al colocarlo (ya no hace falta pulsar Escape)."),
        ("fixed", "En modo Multileader puedes cambiar orientación y texto y se aplican al siguiente que coloques."),
        ("added", "Al editar un texto puedes rotarlo con imán a 0/45/90/135/180/225/270/315°."),
        ("fixed", "El cambio de página con Enter funciona."),
        ("added", "Seleccionar Multileader con clic en el dibujo."),
        ("added", "Ctrl+C / Ctrl+V para duplicar el elemento seleccionado (Multileader, utilidad o texto)."),
    ]),
    ("0.4.1", [
        ("fixed", "Se recuperaron los tipos de línea con letra al exportar: ─ W ─, ─ SS ─, ─ G ─, y la variante abandonada ─/─ W ─. El modo fiel del plano base ya no borra el estilo de las utilidades marcadas."),
    ]),
    ("0.4.0", [
        ("added", "Interfaz reorganizada: herramientas a la izquierda, inventario a la derecha, barra de acción arriba y barra de estado con coordenadas y escala."),
        ("fixed", "Multileader horizontal: el texto sale al lado del extremo de la línea."),
        ("fixed", "Multileader vertical: el texto se pega más a la línea sin tocarla."),
        ("fixed", "Al editar un Multileader, el salto de línea queda más cerca de la línea."),
        ("added", "Colocar Multileader en cadena: tras poner uno queda listo para el siguiente (Esc para salir)."),
        ("fixed", "El cambio de página al escribir un número y dar Enter (o salir del campo) ya funciona."),
        ("added", "Al extender un vértice puedes elegir: misma utilidad o utilidad nueva."),
        ("removed", "Se quitaron las flechas ▲▼ del campo de altura de texto (quedan los botones − / +)."),
    ]),
    ("0.3.1", [
        ("added", "En el cambiador de páginas se puede escribir un número de página y dar Enter."),
        ("fixed", "El editor de texto es una caja flotante: Enter aplica y clic fuera también dibuja el texto."),
        ("fixed", "El botón de aumentar la altura del texto ya funciona (se agregaron botones − / +)."),
        ("fixed", "Multileader vertical y horizontal ahora son perfectamente rectos y la punta de flecha sale recta."),
        ("fixed", "La landing del Multileader se adapta al largo del texto."),
        ("fixed", "En Diagonal el texto ya no se cruza con la landing (queda encima)."),
        ("fixed", "Editar Multileader: Ctrl+Shift+Enter hace salto de línea al otro lado de la landing."),
        ("fixed", "Al cambiar de hoja los textos y Multileaders ya no se dibujan gigantes (tamaño acotado)."),
        ("fixed", "Extender desde un vértice (forma de F) tolera el temblor del ratón al hacer clic."),
    ]),
    ("0.3.0", [
        ("fixed", "Navegación de páginas con botones ◀ ▶ (ya no hay que escribir el número primero)."),
        ("fixed", "El texto libre se agrega con Enter; ya no depende de hacer clic en el panel."),
        ("added", "Enter aplica el texto/Multileader; Ctrl+Shift+Enter (o Shift+Enter) hace salto de línea."),
        ("added", "Textos seleccionables, editables (doble clic) y movibles, con estilo modificable después de crearlos."),
        ("added", "Multileader editable (doble clic) con saltos de línea."),
        ("fixed", "Horizontal ahora es recto (sin quiebre); la opción con landing pasó a Diagonal."),
        ("added", "Vertical recto con texto vertical pegado a la línea."),
        ("added", "Extender desde un vértice: clic en un vértice y luego en el dibujo para ramificar (forma de F)."),
        ("added", "Los Multileader se digitalizan como entidad MULTILEADER de CAD (no como grupo)."),
        ("fixed", "Los polígonos de borrado quedan detrás de las tuberías (solo tapan el plano)."),
        ("added", "El área de listas se puede agrandar arrastrando el separador."),
        ("added", "Se renombró todo a 'Multileader'."),
    ]),
    ("0.2.0", [
        ("fixed", "Error al abrir un proyecto (objetos gráficos ya eliminados y zonas en formato antiguo)."),
        ("fixed", "Enter finaliza de forma fiable la utilidad que estás dibujando."),
        ("added", "Cerrar proyecto: pregunta si deseas guardar los cambios."),
        ("added", "Aviso de cambios sin guardar al cerrar la aplicación."),
        ("added", "Historial de versiones desplegable en 'Acerca de'."),
    ]),
    ("0.1.0", [
        ("added", "Utilidad abandonada con línea ──/── W ──."),
        ("added", "Multileader vertical; edición de texto; texto libre con estilos."),
        ("added", "Panel de zonas borradas; selección por clic; guardar en el mismo archivo."),
        ("added", "Menús Archivo / Edición / Ayuda con manual y atajos."),
        ("base", "Dibujar utilidades, Multileaders, texto, borrar zonas, OCR/ICR, proyectos y export a DXF."),
    ]),
    ("0.0.0", [
        ("base", "Primera versión interna: digitalización PDF→DXF y marcado básico de utilidades."),
    ]),
]


def aci_qcolor(a): return QtGui.QColor(*ACI_RGB.get(a, (235, 235, 235)))
def layer_qcolor(l): return aci_qcolor(C.OUTPUT_LAYERS.get(l, 7))


def swatch_icon(color, size=14):
    pm = QtGui.QPixmap(size, size); pm.fill(QtCore.Qt.transparent)
    p = QtGui.QPainter(pm); p.setBrush(color); p.setPen(QtGui.QPen(QtGui.QColor(70, 70, 70)))
    p.drawRect(0, 0, size - 1, size - 1); p.end(); return QtGui.QIcon(pm)


def qimage_to_gray(qimg):
    qimg = qimg.convertToFormat(QtGui.QImage.Format_Grayscale8)
    w, h = qimg.width(), qimg.height(); ptr = qimg.constBits()
    arr = np.frombuffer(ptr, np.uint8).reshape(h, qimg.bytesPerLine())[:, :w]
    return arr.copy()


def point_in_poly(x, y, poly):
    n = len(poly); inside = False; j = n - 1
    for i in range(n):
        xi, yi = poly[i]; xj, yj = poly[j]
        if ((yi > y) != (yj > y)) and (x < (xj - xi) * (y - yi) / ((yj - yi) or 1e-9) + xi):
            inside = not inside
        j = i
    return inside


class InlineEdit(QtWidgets.QTextEdit):
    """Editor embebido. Enter = aplicar; Ctrl+Shift+Enter o Shift+Enter = salto de
    línea. El commit se difiere con un timer para no destruir el widget dentro de
    su propio evento (eso provocaba cierres inesperados)."""
    committed = QtCore.Signal(str)

    def __init__(self, text):
        super().__init__(); self.setPlainText(text); self._done = False
        self.setStyleSheet("background:#111;color:#7f7;border:1px solid #7f7;")

    def _commit(self):
        if not self._done:
            self._done = True; txt = self.toPlainText()
            QtCore.QTimer.singleShot(0, lambda: self.committed.emit(txt))

    def keyPressEvent(self, e):
        if e.key() in (QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter):
            m = e.modifiers()
            if (m & QtCore.Qt.ShiftModifier):     # Shift(+Ctrl)+Enter → salto de línea
                self.insertPlainText("\n"); return
            self._commit(); return
        super().keyPressEvent(e)

    def focusOutEvent(self, e):
        super().focusOutEvent(e); self._commit()


class PipelineWorker(QtCore.QThread):
    done = QtCore.Signal(str, str)

    def __init__(self, pdf, tmp): super().__init__(); self.pdf, self.tmp = pdf, tmp

    def run(self):
        try:
            import digitize
            digitize.main(self.pdf, self.tmp, verbose=False); self.done.emit(self.tmp, "")
        except Exception as e:
            import traceback; self.done.emit("", f"{e}\n\n{traceback.format_exc()}")


class OcrWorker(QtCore.QThread):
    done = QtCore.Signal(object, str)

    def __init__(self, gray): super().__init__(); self.gray = gray

    def run(self):
        try:
            import cv2, pytesseract
            from pytesseract import Output
            from collections import defaultdict
            g = self.gray; H, W = g.shape
            s = min(1.0, 2400.0 / max(H, W))
            gd = cv2.resize(g, None, fx=s, fy=s, interpolation=cv2.INTER_AREA) if s < 1 else g
            Hd, Wd = gd.shape
            codes = {90: cv2.ROTATE_90_CLOCKWISE, 270: cv2.ROTATE_90_COUNTERCLOCKWISE}
            def back(o, ox, oy):
                if o == 0: return ox, oy
                if o == 90: return oy, Hd - 1 - ox
                return Wd - 1 - oy, ox
            out = []
            for o in (0, 90, 270):
                img = gd if o == 0 else cv2.rotate(gd, codes[o])
                data = pytesseract.image_to_data(img, output_type=Output.DICT, config="--psm 11")
                lines = defaultdict(list)
                for i in range(len(data["text"])):
                    t = (data["text"][i] or "").strip()
                    if not t: continue
                    try: conf = float(data["conf"][i])
                    except: conf = -1
                    if conf < 25: continue
                    k = (data["block_num"][i], data["par_num"][i], data["line_num"][i])
                    lines[k].append((data["left"][i], data["top"][i], data["width"][i], data["height"][i], t))
                for ws in lines.values():
                    ws.sort(key=lambda w: w[0]); txt = " ".join(w[4] for w in ws)
                    x = min(w[0] for w in ws); y = min(w[1] for w in ws)
                    x1 = max(w[0] + w[2] for w in ws); y1 = max(w[1] + w[3] for w in ws)
                    cs = [back(o, x, y), back(o, x1, y), back(o, x, y1), back(o, x1, y1)]
                    xs = [c[0] / s for c in cs]; ys = [c[1] / s for c in cs]
                    out.append((min(xs), min(ys), max(xs) - min(xs), max(ys) - min(ys), txt))
            self.done.emit(out, "")
        except Exception as e:
            self.done.emit([], str(e))


class IcrWorker(QtCore.QThread):
    done = QtCore.Signal(object, str)

    def __init__(self, gray): super().__init__(); self.gray = gray

    def run(self):
        try:
            import easyocr, cv2
            g = self.gray; H, W = g.shape
            s = min(1.0, 2000.0 / max(H, W))
            gd = cv2.resize(g, None, fx=s, fy=s, interpolation=cv2.INTER_AREA) if s < 1 else g
            reader = easyocr.Reader(["en"], gpu=False, verbose=False)
            out = []
            for bbox, txt, conf in reader.readtext(gd, detail=1, paragraph=False):
                if not txt.strip() or conf < 0.15: continue
                xs = [p[0] / s for p in bbox]; ys = [p[1] / s for p in bbox]
                out.append((min(xs), min(ys), max(xs) - min(xs), max(ys) - min(ys), txt.strip()))
            self.done.emit(out, "")
        except ModuleNotFoundError:
            self.done.emit(None, "missing")
        except Exception as e:
            self.done.emit([], str(e))


class Canvas(QtWidgets.QGraphicsView):
    clicked = QtCore.Signal(float, float, object)
    dbl = QtCore.Signal(float, float)
    moved = QtCore.Signal(float, float)

    def __init__(self, win):
        super().__init__(); self.win = win
        self.setScene(QtWidgets.QGraphicsScene(self))
        self.setRenderHints(QtGui.QPainter.Antialiasing | QtGui.QPainter.SmoothPixmapTransform)
        self.setTransformationAnchor(QtWidgets.QGraphicsView.AnchorUnderMouse)
        self.setBackgroundBrush(QtGui.QColor(13, 19, 33)); self.setAcceptDrops(True)
        self.setFocusPolicy(QtCore.Qt.StrongFocus)
        self.setMouseTracking(True); self.viewport().setMouseTracking(True)
        self.pixmap_item = None; self._pan = False; self._pan0 = None; self._moving = False
        self.pdf_opacity = 1.0

    def set_image(self, qimg):
        self.scene().clear()
        self.pixmap_item = self.scene().addPixmap(QtGui.QPixmap.fromImage(qimg))
        self.pixmap_item.setZValue(Z_PDF)
        self.pixmap_item.setOpacity(self.pdf_opacity)
        self.setSceneRect(self.pixmap_item.boundingRect())
        self.resetTransform(); self.fitInView(self.pixmap_item, QtCore.Qt.KeepAspectRatio)

    def set_pdf_opacity(self, val):
        self.pdf_opacity = max(0.1, min(1.0, val))
        if self.pixmap_item: self.pixmap_item.setOpacity(self.pdf_opacity)

    def keyPressEvent(self, e):
        if e.key() in (QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter):
            self.win._on_enter(); e.accept(); return
        if e.key() == QtCore.Qt.Key_Escape:
            self.win._on_escape(); e.accept(); return
        super().keyPressEvent(e)

    def wheelEvent(self, e):
        if self.pixmap_item:
            f = 1.25 if e.angleDelta().y() > 0 else 0.8; self.scale(f, f)

    def mousePressEvent(self, e):
        if e.button() == QtCore.Qt.MiddleButton:
            self._pan = True; self._pan0 = e.position(); self.setCursor(QtCore.Qt.ClosedHandCursor); return
        if not self.pixmap_item: return super().mousePressEvent(e)
        self.setFocus(QtCore.Qt.MouseFocusReason)
        sp = self.mapToScene(e.position().toPoint())
        if self.win.mode == "move" and e.button() == QtCore.Qt.LeftButton:
            self._moving = True; self.win.begin_move(sp.x(), sp.y()); return
        if e.button() in (QtCore.Qt.LeftButton, QtCore.Qt.RightButton):
            self.clicked.emit(sp.x(), sp.y(), e.button()); return
        super().mousePressEvent(e)

    def mouseMoveEvent(self, e):
        if self.pixmap_item:
            sp = self.mapToScene(e.position().toPoint()); self.moved.emit(sp.x(), sp.y())
        if self._pan:
            d = e.position() - self._pan0; self._pan0 = e.position()
            self.horizontalScrollBar().setValue(int(self.horizontalScrollBar().value() - d.x()))
            self.verticalScrollBar().setValue(int(self.verticalScrollBar().value() - d.y())); return
        if self._moving:
            sp = self.mapToScene(e.position().toPoint()); self.win.do_move(sp.x(), sp.y()); return
        super().mouseMoveEvent(e)

    def mouseReleaseEvent(self, e):
        if e.button() == QtCore.Qt.MiddleButton:
            self._pan = False; self.setCursor(QtCore.Qt.ArrowCursor); return
        if self._moving and e.button() == QtCore.Qt.LeftButton:
            self._moving = False; self.win.end_move(); return
        super().mouseReleaseEvent(e)

    def mouseDoubleClickEvent(self, e):
        if self.win.mode == "pipe": self.win.finish_pipe(); return
        sp = self.mapToScene(e.position().toPoint()); self.dbl.emit(sp.x(), sp.y())

    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls(): e.acceptProposedAction()

    def dropEvent(self, e):
        for u in e.mimeData().urls():
            self.win.open_path(u.toLocalFile()); break


class Main(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"Digitalizador de planos — utilidades  (v{VERSION})"); self.resize(1480, 940)
        self.setAcceptDrops(True)
        self.canvas = Canvas(self); self.canvas.clicked.connect(self.on_click)
        self.canvas.dbl.connect(self.on_dblclick); self.setCentralWidget(self.canvas)
        self.zoom = 3.5; self.scale = 20 / 72.0; self.rot = 0; self.W = 0; self.H = 0
        self.derot = fitz.Matrix(1, 0, 0, 1, 0, 0); self.gray = None; self.page_idx = 0; self.pageH_px = 0
        self.pdf_path = None; self.doc = None; self.project_path = None; self.leader_hpx = 40
        self.cur_pts = []; self.pipes = []; self.leaders = []; self.text_marks = []
        self.erase_regions = []; self._erase_pts = []
        self.mode = "idle"; self._pending = None
        self.snap = False; self.snap_r = 14
        self.show_text_boxes = False; self.ocr_boxes = []; self._tess_boxes = []; self._icr_boxes = []
        self.sel_pipe = -1; self.sel_leader = -1; self.sel_region = -1; self.sel_text = -1; self._no_center = False
        self._move0 = None; self._drag_vertex = None; self._edit_pts = None; self._edit_closed = False; self._edit_leader = None
        self._move_kind = None; self._moved = False; self._press_xy = None; self._last_xy = None
        self._extending = False; self._ext_layer = None; self._ext_pipe = None; self._ext_at = None
        self._editor = None; self._undo, self._redo, self._overlay = [], [], []
        self._dirty = False; self._style_guard = False; self._prop_guard = False; self._clip = None
        self._build_ui(); self._apply_style(); self._shortcuts(); self._update_ui()

    # ─────────────────────────── UI ───────────────────────────
    def _build_ui(self):
        mb = self.menuBar()
        mfile = mb.addMenu("&Archivo")
        self._menu_act(mfile, "Abrir PDF…", self.open_pdf)
        self._menu_act(mfile, "Abrir Excel…", self.open_excel)
        mfile.addSeparator()
        self._menu_act(mfile, "Abrir proyecto…", self.open_project)
        self._menu_act(mfile, "Guardar proyecto", self.save_project, "Ctrl+S")
        self._menu_act(mfile, "Guardar proyecto como…", self.save_project_as, "Ctrl+Shift+S")
        mfile.addSeparator()
        self._menu_act(mfile, "Cerrar proyecto", self.close_project, "Ctrl+W")
        medit = mb.addMenu("&Edición")
        self._menu_act(medit, "Deshacer", self.undo, "Ctrl+Z")
        self._menu_act(medit, "Rehacer", self.redo, "Ctrl+Shift+Z")
        mhelp = mb.addMenu("A&yuda")
        self._menu_act(mhelp, "Acerca de…", self.show_about)
        self._menu_act(mhelp, "Manual de usuario", self.show_manual)
        self._menu_act(mhelp, "Atajos de teclado", self.show_shortcuts)

        # ── Barra de acción superior: zoom · deshacer/rehacer · imán · exportar ──
        tb = self.addToolBar("Acciones"); tb.setMovable(False)
        def tact(txt, tip, fn):
            a = QtGui.QAction(txt, self); a.setToolTip(tip); a.triggered.connect(fn); tb.addAction(a); return a
        tact("🔍＋", "Acercar", self._zoom_in); tact("🔍－", "Alejar", self._zoom_out)
        tb.addSeparator(); tact("↶", "Deshacer (Ctrl+Z)", self.undo); tact("↷", "Rehacer (Ctrl+Shift+Z)", self.redo)
        tb.addSeparator()
        self.chk_snap = QtWidgets.QCheckBox("Imán al trazo"); self.chk_snap.toggled.connect(self._toggle_snap); tb.addWidget(self.chk_snap)
        spacer = QtWidgets.QWidget(); spacer.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Preferred); tb.addWidget(spacer)
        self.btn_export = QtWidgets.QToolButton()
        self.btn_export.setText("⭳  Exportar DXF  ▾")
        self.btn_export.setPopupMode(QtWidgets.QToolButton.InstantPopup)
        self.btn_export.setStyleSheet("QToolButton{background:#4d8eff;color:#00285d;font-weight:bold;padding:5px 14px;border-radius:4px;}")
        exp_menu = QtWidgets.QMenu(self.btn_export)
        exp_menu.addAction("PDF + anotaciones (todo)", lambda: self.run_pipeline("todo"))
        exp_menu.addAction("Solo el PDF digitalizado", lambda: self.run_pipeline("pdf"))
        exp_menu.addAction("Solo las anotaciones", lambda: self.run_pipeline("anot"))
        self.btn_export.setMenu(exp_menu); tb.addWidget(self.btn_export)

        # ── DOCK IZQUIERDO: herramientas y flujo de trabajo ──
        ldock = QtWidgets.QDockWidget("Herramientas", self); ldock.setFeatures(QtWidgets.QDockWidget.NoDockWidgetFeatures)
        lscroll = QtWidgets.QScrollArea(); lscroll.setWidgetResizable(True)
        left = QtWidgets.QWidget(); v = QtWidgets.QVBoxLayout(left); v.setSpacing(8)

        def section(title):
            lbl = QtWidgets.QLabel(title.upper()); lbl.setStyleSheet("color:#8c909f;font-size:10px;font-weight:600;letter-spacing:1px;margin-top:4px;")
            v.addWidget(lbl)

        # 1) Páginas
        section("Páginas")
        gp = QtWidgets.QWidget(); lp = QtWidgets.QHBoxLayout(gp); lp.setContentsMargins(0, 0, 0, 0)
        self.btn_prev = QtWidgets.QPushButton("◀"); self.btn_prev.setFixedWidth(34); self.btn_prev.clicked.connect(self._prev_page)
        self.btn_next = QtWidgets.QPushButton("▶"); self.btn_next.setFixedWidth(34); self.btn_next.clicked.connect(self._next_page)
        self.page_edit = QtWidgets.QLineEdit(); self.page_edit.setAlignment(QtCore.Qt.AlignCenter)
        self.page_edit.setToolTip("Escribe un número de página y pulsa Enter")
        self.page_edit.returnPressed.connect(self._goto_page_edit)
        self.page_edit.editingFinished.connect(self._goto_page_edit)
        self.lbl_page = QtWidgets.QLabel(" / — ")
        lp.addWidget(self.btn_prev); lp.addWidget(self.page_edit, 1); lp.addWidget(self.lbl_page); lp.addWidget(self.btn_next)
        v.addWidget(gp)

        # 1b) Transparencia del PDF de fondo
        section("Transparencia del PDF")
        gtr = QtWidgets.QWidget(); ltr = QtWidgets.QHBoxLayout(gtr); ltr.setContentsMargins(0, 0, 0, 0)
        tb_l = QtWidgets.QPushButton("−"); tb_l.setFixedWidth(30); tb_l.setToolTip("Más translúcido"); tb_l.clicked.connect(lambda: self._bump_opacity(-10))
        self.lbl_opacity = QtWidgets.QLabel("100%"); self.lbl_opacity.setAlignment(QtCore.Qt.AlignCenter)
        tb_r = QtWidgets.QPushButton("+"); tb_r.setFixedWidth(30); tb_r.setToolTip("Más opaco"); tb_r.clicked.connect(lambda: self._bump_opacity(10))
        ltr.addWidget(tb_l); ltr.addWidget(self.lbl_opacity, 1); ltr.addWidget(tb_r)
        v.addWidget(gtr)

        # 2) Acciones principales (qué quieres hacer)
        section("Acción / herramienta")
        self.btn_pipe = QtWidgets.QPushButton("✏  Dibujar utilidad"); self.btn_pipe.clicked.connect(self.toggle_pipe)
        self.btn_leader = QtWidgets.QPushButton("↳  Colocar Multileader"); self.btn_leader.clicked.connect(lambda: self.start_leader(False))
        self.btn_leader_simple = QtWidgets.QPushButton("↘  Colocar Leader"); self.btn_leader_simple.clicked.connect(lambda: self.start_leader(True))
        self.btn_text = QtWidgets.QPushButton("T  Texto libre"); self.btn_text.clicked.connect(self.toggle_text_mode)
        self.btn_erase = QtWidgets.QPushButton("▭  Borrar zona"); self.btn_erase.clicked.connect(self.toggle_erase)
        for b in (self.btn_pipe, self.btn_leader, self.btn_leader_simple, self.btn_text, self.btn_erase): v.addWidget(b)

        # 3) Tipo de utilidad
        section("Tipo de utilidad")
        self.gt = QtWidgets.QWidget(); lgt = QtWidgets.QVBoxLayout(self.gt); lgt.setContentsMargins(0, 0, 0, 0)
        self.type_combo = QtWidgets.QComboBox()
        for label, layer in TIPOS:
            self.type_combo.addItem(swatch_icon(layer_qcolor(layer)), label, layer)
        self.type_combo.setCurrentIndex(0); self.type_combo.currentIndexChanged.connect(lambda _: self._redraw())
        self.chk_ab = QtWidgets.QCheckBox("Abandonado (línea ──/── W ──)")
        self.chk_ext_same = QtWidgets.QCheckBox("Al extender un extremo: continuar la misma utilidad")
        self.chk_ext_same.setChecked(True)
        lgt.addWidget(self.type_combo); lgt.addWidget(self.chk_ab); lgt.addWidget(self.chk_ext_same); v.addWidget(self.gt)

        # 4) Panel de Multileader
        self.ga = QtWidgets.QGroupBox("Multileader"); lga = QtWidgets.QVBoxLayout(self.ga)
        self.orient_combo = QtWidgets.QComboBox()
        for oid, lbl in LEADER_ORIENT: self.orient_combo.addItem(lbl, oid)
        self.orient_combo.currentIndexChanged.connect(lambda _: self._update_ui())
        lga.addWidget(QtWidgets.QLabel("Orientación:")); lga.addWidget(self.orient_combo)
        self.chk_custom = QtWidgets.QCheckBox("Usar texto personalizado"); self.chk_custom.toggled.connect(self._toggle_custom)
        lga.addWidget(self.chk_custom)
        self.txt_edit = QtWidgets.QLineEdit(); self.txt_edit.setPlaceholderText("texto personalizado…"); self.txt_edit.setEnabled(False)
        lga.addWidget(self.txt_edit)
        self.lbl_textos = QtWidgets.QLabel("Textos (columna TEXTO del Excel):"); lga.addWidget(self.lbl_textos)
        self.text_list = QtWidgets.QListWidget(); self.text_list.setMaximumHeight(140); lga.addWidget(self.text_list)
        self.lbl_lead_hint = QtWidgets.QLabel("<i>Colocas varios seguidos; Esc para salir.</i>"); lga.addWidget(self.lbl_lead_hint)
        v.addWidget(self.ga)

        # 5) Estilo de texto
        self.gtxt = QtWidgets.QGroupBox("Estilo de texto"); lgx = QtWidgets.QVBoxLayout(self.gtxt)
        self.font_combo = QtWidgets.QFontComboBox(); self.font_combo.setCurrentFont(QtGui.QFont(C.TEXT_FONT))
        self.font_combo.currentFontChanged.connect(lambda _: self._style_changed())
        r = QtWidgets.QHBoxLayout(); r.addWidget(QtWidgets.QLabel("Altura (pies):"))
        b_minus = QtWidgets.QPushButton("−"); b_minus.setFixedWidth(30); b_minus.clicked.connect(lambda: self._bump_size(-0.5))
        self.size_spin = QtWidgets.QDoubleSpinBox(); self.size_spin.setRange(0.5, 200); self.size_spin.setValue(3.0)
        self.size_spin.setSingleStep(0.5); self.size_spin.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)
        self.size_spin.valueChanged.connect(lambda _: self._style_changed())
        b_plus = QtWidgets.QPushButton("+"); b_plus.setFixedWidth(30); b_plus.clicked.connect(lambda: self._bump_size(0.5))
        r.addWidget(b_minus); r.addWidget(self.size_spin); r.addWidget(b_plus)
        self.chk_bold = QtWidgets.QCheckBox("Negrita"); self.chk_bold.toggled.connect(lambda _: self._style_changed())
        lgx.addWidget(self.font_combo); lgx.addLayout(r); lgx.addWidget(self.chk_bold)
        # Rotación (solo textos): giro libre 0–360, botones de 1 en 1
        self.rot_row = QtWidgets.QWidget(); rr2 = QtWidgets.QHBoxLayout(self.rot_row); rr2.setContentsMargins(0, 0, 0, 0)
        rr2.addWidget(QtWidgets.QLabel("Rotación (°):"))
        rb_l = QtWidgets.QPushButton("⟲"); rb_l.setFixedWidth(30); rb_l.clicked.connect(lambda: self._bump_rot(-1))
        self.rot_spin = QtWidgets.QSpinBox(); self.rot_spin.setRange(0, 360); self.rot_spin.setSingleStep(1); self.rot_spin.setWrapping(True)
        self.rot_spin.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons); self.rot_spin.valueChanged.connect(lambda _: self._style_changed())
        rb_r = QtWidgets.QPushButton("⟳"); rb_r.setFixedWidth(30); rb_r.clicked.connect(lambda: self._bump_rot(1))
        rr2.addWidget(rb_l); rr2.addWidget(self.rot_spin); rr2.addWidget(rb_r)
        lgx.addWidget(self.rot_row)
        lgx.addWidget(QtWidgets.QLabel("<i>Enter aplica · Ctrl+Shift+Enter salta de línea</i>"))
        v.addWidget(self.gtxt)

        # 6) En curso
        self.gcur = QtWidgets.QGroupBox("En curso"); lc = QtWidgets.QHBoxLayout(self.gcur)
        self.btn_fin = QtWidgets.QPushButton("Finalizar (Enter)"); self.btn_fin.clicked.connect(self._on_enter)
        b_up = QtWidgets.QPushButton("Deshacer punto"); b_up.clicked.connect(self.undo)
        lc.addWidget(self.btn_fin); lc.addWidget(b_up); v.addWidget(self.gcur)

        # 7) Reconocimiento
        section("Reconocimiento de texto")
        self.chk_txt = QtWidgets.QCheckBox("Textos impresos (OCR)"); self.chk_txt.toggled.connect(self.toggle_text_boxes)
        self.chk_icr = QtWidgets.QCheckBox("Manuscrita (ICR, offline)"); self.chk_icr.toggled.connect(self.toggle_icr)
        v.addWidget(self.chk_txt); v.addWidget(self.chk_icr)
        v.addStretch(1)

        lscroll.setWidget(left); ldock.setWidget(lscroll); ldock.setMinimumWidth(250)
        self.addDockWidget(QtCore.Qt.LeftDockWidgetArea, ldock)

        # ── DOCK DERECHO: inventario y selección ──
        rdock = QtWidgets.QDockWidget("Inventario", self); rdock.setFeatures(QtWidgets.QDockWidget.NoDockWidgetFeatures)
        right = QtWidgets.QWidget(); rv = QtWidgets.QVBoxLayout(right)
        self.tabs = QtWidgets.QTabWidget()
        self.pipe_list = QtWidgets.QListWidget(); self.pipe_list.currentRowChanged.connect(self._sel_pipe)
        self.lead_list = QtWidgets.QListWidget(); self.lead_list.currentRowChanged.connect(self._sel_leader)
        self.sleader_list = QtWidgets.QListWidget(); self.sleader_list.currentRowChanged.connect(self._sel_sleader)
        self.txt_marks_list = QtWidgets.QListWidget(); self.txt_marks_list.currentRowChanged.connect(self._sel_text)
        self.region_list = QtWidgets.QListWidget(); self.region_list.currentRowChanged.connect(self._sel_region)
        self.region_list.itemChanged.connect(self._region_toggled)
        self.tabs.addTab(self.pipe_list, "Utilidades"); self.tabs.addTab(self.lead_list, "Multileaders")
        self.tabs.addTab(self.sleader_list, "Leaders")
        self.tabs.addTab(self.txt_marks_list, "Textos"); self.tabs.addTab(self.region_list, "Zonas")
        # Menú contextual (clic derecho) en cada lista del inventario
        for listw, tab_idx in ((self.pipe_list, TAB_PIPE), (self.lead_list, TAB_ML),
                               (self.sleader_list, TAB_LEADER), (self.txt_marks_list, TAB_TEXT),
                               (self.region_list, TAB_REGION)):
            listw.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
            listw.customContextMenuRequested.connect(
                lambda pos, lw=listw, ti=tab_idx: self._list_context_menu(lw, ti, pos))
        self.tabs.currentChanged.connect(self._tab_changed); rv.addWidget(self.tabs, 1)
        # Propiedades de la utilidad seleccionada (nombre, diámetro, unidad) → XDATA en el DXF
        self.gprop = QtWidgets.QGroupBox("Propiedades de la utilidad"); fpr = QtWidgets.QFormLayout(self.gprop)
        self.prop_name = QtWidgets.QLineEdit(); self.prop_name.editingFinished.connect(self._prop_changed)
        self.prop_diam = QtWidgets.QDoubleSpinBox(); self.prop_diam.setRange(0, 100000); self.prop_diam.setDecimals(2)
        self.prop_diam.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons); self.prop_diam.valueChanged.connect(lambda _: self._prop_changed())
        self.prop_unit = QtWidgets.QComboBox(); self.prop_unit.addItems(["pulg", "pies"]); self.prop_unit.currentTextChanged.connect(lambda _: self._prop_changed())
        fpr.addRow("Nombre:", self.prop_name); fpr.addRow("Diámetro:", self.prop_diam); fpr.addRow("Unidad:", self.prop_unit)
        rv.addWidget(self.gprop)
        rr = QtWidgets.QGridLayout()
        self.btn_ct = QtWidgets.QPushButton("Cambiar tipo"); self.btn_ct.clicked.connect(self.change_pipe_type)
        self.btn_mv = QtWidgets.QPushButton("Editar/mover"); self.btn_mv.clicked.connect(self.enter_move)
        self.btn_edit = QtWidgets.QPushButton("Editar texto"); self.btn_edit.clicked.connect(self.edit_selected_text)
        self.btn_del = QtWidgets.QPushButton("Eliminar"); self.btn_del.clicked.connect(self.delete_selected)
        rr.addWidget(self.btn_ct, 0, 0); rr.addWidget(self.btn_mv, 0, 1)
        rr.addWidget(self.btn_edit, 1, 0); rr.addWidget(self.btn_del, 1, 1)
        rv.addLayout(rr)
        rdock.setWidget(right); rdock.setMinimumWidth(250)
        self.addDockWidget(QtCore.Qt.RightDockWidgetArea, rdock)

        # ── Barra de estado: modo · coordenadas · escala · imán ──
        self.status = self.statusBar(); self.status.setSizeGripEnabled(False)
        self.lbl_mode = QtWidgets.QLabel("Modo: inactivo"); self.lbl_mode.setStyleSheet("color:#adc6ff;")
        self.status.addWidget(self.lbl_mode)
        self.status.addWidget(QtWidgets.QLabel("│"))
        self.lbl_info = QtWidgets.QLabel(""); self.lbl_info.setStyleSheet("color:#8c909f;"); self.status.addWidget(self.lbl_info, 1)
        self.lbl_snap = QtWidgets.QLabel("Imán: OFF"); self.lbl_coords = QtWidgets.QLabel("X —  Y —")
        self.lbl_scale = QtWidgets.QLabel("Escala —")
        for w in (self.lbl_snap, self.lbl_coords, self.lbl_scale):
            w.setStyleSheet("color:#c2c6d6;"); self.status.addPermanentWidget(w)
        self.canvas.moved.connect(self._update_coords)
        self._info("Abre o arrastra un PDF/proyecto.")

    def _menu_act(self, menu, text, fn, sc=None):
        a = QtGui.QAction(text, self); a.triggered.connect(fn)
        if sc: a.setShortcut(sc)
        menu.addAction(a); return a

    def _apply_style(self):
        self.setStyleSheet("""
            QMainWindow,QDockWidget,QGroupBox,QScrollArea,QMenuBar,QMenu{background:#2b2b2b;color:#eee;}
            QMenuBar::item:selected,QMenu::item:selected{background:#3c5a99;}
            QGroupBox{border:1px solid #444;margin-top:8px;padding:6px;border-radius:4px;font-weight:bold;}
            QGroupBox::title{subcontrol-origin:margin;left:8px;color:#bbb;}
            QListWidget,QLineEdit,QComboBox,QSpinBox,QDoubleSpinBox,QFontComboBox,QTabWidget::pane{background:#333;color:#eee;border:1px solid #555;}
            QTabBar::tab{background:#333;color:#ccc;padding:5px;} QTabBar::tab:selected{background:#3c5a99;color:white;}
            QPushButton{background:#3c5a99;color:white;border:none;padding:6px;border-radius:4px;}
            QPushButton:hover{background:#4a6fbf;} QLabel,QCheckBox{color:#ddd;font-weight:normal;}
            QSplitter::handle{background:#555;height:6px;}""")
        self.chk_snap.setChecked(False)

    def _shortcuts(self):
        QtGui.QShortcut(QtGui.QKeySequence("Ctrl+T"), self, self.enter_move)
        for k in (QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter):
            QtGui.QShortcut(QtGui.QKeySequence(k), self, self._on_enter)
        QtGui.QShortcut(QtGui.QKeySequence(QtCore.Qt.Key_Escape), self, self._on_escape)
        QtGui.QShortcut(QtGui.QKeySequence("Ctrl+C"), self, self._copy_sel)
        QtGui.QShortcut(QtGui.QKeySequence("Ctrl+V"), self, self._paste_sel)
        QtGui.QShortcut(QtGui.QKeySequence(QtCore.Qt.Key_Delete), self, self._delete_shortcut)

    def _delete_shortcut(self):
        # Suprimir borra el elemento seleccionado, salvo mientras se escribe texto
        if self._editor is not None: return
        fw = QtWidgets.QApplication.focusWidget()
        if isinstance(fw, (QtWidgets.QLineEdit, QtWidgets.QTextEdit, QtWidgets.QAbstractSpinBox)): return
        self.delete_selected()

    # ─────────────────────────── páginas ───────────────────────────
    def _update_page_label(self):
        if self.doc:
            self.page_edit.setText(str(self.page_idx + 1)); self.lbl_page.setText(f" / {self.doc.page_count} ")
            self.page_edit.setEnabled(True)
            self.btn_prev.setEnabled(self.page_idx > 0)
            self.btn_next.setEnabled(self.page_idx < self.doc.page_count - 1)
        else:
            self.page_edit.setText(""); self.page_edit.setEnabled(False); self.lbl_page.setText(" / — ")
            self.btn_prev.setEnabled(False); self.btn_next.setEnabled(False)

    def _goto_page_edit(self):
        if not self.doc: return
        try: n = int(self.page_edit.text()) - 1
        except ValueError: self._update_page_label(); return
        n = max(0, min(n, self.doc.page_count - 1))
        if n != self.page_idx:
            if not self._confirm_discard(): self._update_page_label(); return
            self.page_idx = n; self._load_page(n)
        else:
            self._update_page_label()

    def _prev_page(self):
        if self.doc and self.page_idx > 0:
            if not self._confirm_discard(): return
            self.page_idx -= 1; self._load_page(self.page_idx)

    def _next_page(self):
        if self.doc and self.page_idx < self.doc.page_count - 1:
            if not self._confirm_discard(): return
            self.page_idx += 1; self._load_page(self.page_idx)

    # ─────────────────────────── Estado / modos ───────────────────────────
    def _on_enter(self):
        if self.mode == "pipe": self.finish_pipe()
        elif self.mode == "erase": self.finish_erase()
        elif self.mode in ("leader1", "leader2", "leader3"):
            self.set_mode("idle"); self._info("Comando Leader finalizado (Enter)")
        elif self.mode == "move":
            self.set_mode("idle"); self._info("Edición terminada (Enter)")

    def _on_escape(self):
        if self.mode == "pipe" and self.cur_pts:
            self._push(); self.cur_pts = []; self._extending = False; self._ext_pipe = None; self._ext_at = None
            self._update_ui(); self._redraw(); self._info("Puntos cancelados")
        elif self.mode == "erase" and self._erase_pts:
            self._erase_pts = []; self._redraw(); self._info("Zona cancelada")
        elif self.sel_pipe >= 0 or self.sel_leader >= 0 or self.sel_region >= 0 or self.sel_text >= 0:
            self._deselect_all(); self._info("Selección quitada")
        else:
            self.set_mode("idle"); self._info("Salió del modo")

    def _deselect_all(self):
        self.sel_pipe = self.sel_leader = self.sel_region = self.sel_text = -1
        for lst in (self.pipe_list, self.lead_list, self.sleader_list, self.txt_marks_list, self.region_list):
            lst.blockSignals(True); lst.setCurrentRow(-1); lst.clearSelection(); lst.blockSignals(False)
        if self.mode == "move": self.set_mode("idle")
        self._update_ui(); self._redraw()

    def _info(self, m): self.lbl_info.setText(m)

    def _zoom_in(self): self.canvas.scale(1.25, 1.25)
    def _zoom_out(self): self.canvas.scale(0.8, 0.8)

    def _toggle_snap(self, v):
        self.snap = v; self.lbl_snap.setText(f"Imán: {'ON' if v else 'OFF'}")
        self.lbl_snap.setStyleSheet("color:%s;" % ("#5fd35f" if v else "#c2c6d6"))

    def _update_coords(self, x, y):
        if self.canvas.pixmap_item is None: return
        cx, cy = self._to_cad(x, y); self.lbl_coords.setText(f"X {cx:,.1f}  Y {cy:,.1f}")

    def _update_ui(self):
        m = self.mode
        def st(btn, on): btn.setStyleSheet(BTN_ON if on else BTN_OFF)
        in_leader = m in ("leader1", "leader2", "leader3"); simple = bool(self._pending and self._pending.get("simple"))
        st(self.btn_pipe, m == "pipe"); st(self.btn_leader, in_leader and not simple)
        st(self.btn_leader_simple, in_leader and simple)
        st(self.btn_text, m == "text"); st(self.btn_erase, m == "erase")
        self.btn_pipe.setText("■  Salir de dibujar utilidad" if m == "pipe" else "✏  Dibujar utilidad")
        self.btn_leader.setText("●  Coloque Multileader…" if (in_leader and not simple) else "↳  Colocar Multileader")
        self.btn_leader_simple.setText("●  Coloque Leader…" if (in_leader and simple) else "↘  Colocar Leader")
        self.btn_erase.setText("■  Terminar zona (Enter)" if m == "erase" else "▭  Borrar zona (polígono)")
        ti = self.tabs.currentIndex()
        self.gt.setVisible(m in ("pipe", "move"))
        # Leader simple (solo flecha): el panel muestra únicamente la orientación y se titula "Leader"
        sel_ld_simple = ti == TAB_LEADER and 0 <= self.sel_leader < len(self.leaders) and self.leaders[self.sel_leader].get("simple")
        lead_simple = (in_leader and simple) or bool(sel_ld_simple)
        self.ga.setTitle("Leader" if lead_simple else "Multileader")
        for w in (self.chk_custom, self.txt_edit, self.lbl_textos, self.text_list, self.lbl_lead_hint):
            w.setVisible(not lead_simple)
        # el panel de estilo sirve para texto libre y para el Multileader; el Leader simple no lo usa
        txt_ctx = (m == "text") or (ti == TAB_TEXT and self.sel_text >= 0)
        lead_ctx = (m in ("leader1", "leader2", "leader3")) or (ti in (TAB_ML, TAB_LEADER) and self.sel_leader >= 0)
        self.gtxt.setVisible((txt_ctx or lead_ctx) and not lead_simple)
        self.gtxt.setTitle("Estilo del Multileader" if (lead_ctx and not txt_ctx) else "Estilo de texto")
        self.rot_row.setVisible(txt_ctx)                 # la rotación solo aplica a textos libres
        self.gprop.setVisible(ti == TAB_PIPE and self.sel_pipe >= 0)
        self.gcur.setVisible((m == "pipe" and len(self.cur_pts) >= 1) or (m == "erase" and len(self._erase_pts) >= 1))
        self.btn_fin.setEnabled((m == "pipe" and len(self.cur_pts) >= 2) or (m == "erase" and len(self._erase_pts) >= 3))
        ti = self.tabs.currentIndex()
        self.btn_ct.setVisible(ti == TAB_PIPE)
        self.btn_mv.setVisible(ti in (TAB_PIPE, TAB_LEADER, TAB_TEXT, TAB_REGION))
        self.btn_mv.setText("Mover" if ti == TAB_TEXT else "Editar/mover")
        self.btn_edit.setVisible(ti in (TAB_ML, TAB_TEXT))
        if simple:                                       # Leader (solo flecha)
            diag = self.orient_combo.currentData() == "d"
            lead1 = "Modo: Leader — clic en la cabeza de flecha (dónde señala)"
            lead2 = ("Modo: Leader — clic en el inicio del landing (bisagra)" if diag
                     else "Modo: Leader — clic en el final del cuerpo")
            lead3 = "Modo: Leader — clic en el final del cuerpo"
        else:
            lead1 = "Modo: Multileader — clic en la PUNTA (a qué señala)"
            lead2 = "Modo: Multileader — clic dónde va el TEXTO"
            lead3 = ""
        self.lbl_mode.setText({"idle": "Modo: inactivo  ·  clic en el dibujo para seleccionar",
                               "pipe": ("Modo: EXTENDIENDO desde el vértice — clic agrega puntos, Enter finaliza"
                                        if self._extending else "Modo: dibujar utilidad  ·  Enter finaliza"),
                               "leader1": lead1,
                               "leader2": lead2,
                               "leader3": lead3,
                               "text": "Modo: texto libre — clic donde escribir · Enter aplica",
                               "erase": "Modo: borrar zona — clic para el polígono, Enter cierra",
                               "move": "Modo: editar — arrastra vértice · clic en tramo inserta · clic-en-vértice extiende (F) · clic derecho elimina"}.get(m, ""))

    def set_mode(self, m):
        if m not in ("leader1", "leader2", "leader3"): self._pending = None
        if m != "erase": self._erase_pts = []
        if m != "pipe": self._extending = False
        self.mode = m; self._update_ui(); self._redraw()

    def _tab_changed(self, _):
        ti = self.tabs.currentIndex()               # sel_leader se comparte entre ML y Leaders: re-sincronizar
        if ti == TAB_ML: self.sel_leader = self._leader_at_row(self.lead_list, self.lead_list.currentRow())
        elif ti == TAB_LEADER: self.sel_leader = self._leader_at_row(self.sleader_list, self.sleader_list.currentRow())
        if self.mode == "move": self.set_mode("idle")   # no seguir editando al cambiar de pestaña
        self._update_ui(); self._redraw()
    def toggle_pipe(self): self.set_mode("idle" if self.mode == "pipe" else "pipe")
    def toggle_text_mode(self): self.set_mode("idle" if self.mode == "text" else "text")
    def toggle_erase(self): self.set_mode("idle" if self.mode == "erase" else "erase")
    def active_layer(self):
        d = self.type_combo.currentData(); return d if d else "AGUA"

    def _toggle_custom(self, on):
        self.txt_edit.setEnabled(on); self.text_list.setEnabled(not on)
        if on:
            self.text_list.blockSignals(True); self.text_list.setCurrentRow(-1); self.text_list.clearSelection(); self.text_list.blockSignals(False)
        else:
            self.txt_edit.clear()

    # ─────────────────────────── undo/redo ───────────────────────────
    def _snap_state(self):
        return copy.deepcopy(dict(cur_pts=self.cur_pts, pipes=self.pipes, leaders=self.leaders,
                                  text_marks=self.text_marks, erase_regions=self.erase_regions))

    def _push(self):
        self._undo.append(self._snap_state()); self._redo.clear(); self._dirty = True
        if len(self._undo) > 400: self._undo.pop(0)

    def _restore(self, s):
        self.cur_pts, self.pipes = s["cur_pts"], s["pipes"]
        self.leaders, self.text_marks = s["leaders"], s["text_marks"]
        self.erase_regions = s.get("erase_regions", [])
        self._refresh_lists(); self._update_ui(); self._redraw()

    def undo(self):
        if self._undo: self._redo.append(self._snap_state()); self._restore(self._undo.pop()); self._info("Deshacer")

    def redo(self):
        if self._redo: self._undo.append(self._snap_state()); self._restore(self._redo.pop()); self._info("Rehacer")

    # ─────────────────────────── abrir ───────────────────────────
    def _busy(self, m="Procesando…"):
        QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.WaitCursor); self._info(m); QtWidgets.QApplication.processEvents()
    def _unbusy(self): QtWidgets.QApplication.restoreOverrideCursor()

    def open_path(self, path):
        low = path.lower()
        if low.endswith(".pdf"): self._open_pdf_path(path)
        elif low.endswith(".digproj"): self._open_project_path(path)
        elif low.endswith((".xlsx", ".xlsm")): self._read_excel(path)

    def open_pdf(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Abrir PDF", DOWNLOADS, "PDF (*.pdf)")
        if path: self._open_pdf_path(path)

    def _open_pdf_path(self, path):
        if not self._confirm_discard(): return
        self._busy("Abriendo PDF…")
        try:
            self.pdf_path = path; self.project_path = None; self.doc = fitz.open(path)
            self.page_idx = 0; self._load_page(0)
        finally: self._unbusy()

    def _load_page(self, idx):
        self._close_editor()
        page = self.doc[idx]; self.page_idx = idx; self.scale = VP.detect_scale(page)
        self.rot = page.rotation; mbx = page.mediabox; self.W, self.H = mbx.width, mbx.height
        self.derot = page.derotation_matrix
        pix = page.get_pixmap(matrix=fitz.Matrix(self.zoom, self.zoom), alpha=False)
        self.pageH_px = pix.height
        # tamaño de marca acotado: evita textos/Multileaders gigantes por escala mal detectada
        self.leader_hpx = max(14.0, min(LEADER_TEXT_FT / self.scale * self.zoom, self.pageH_px * 0.05))
        buf = bytes(pix.samples)
        qimg = QtGui.QImage(buf, pix.width, pix.height, pix.stride, QtGui.QImage.Format_RGB888).copy()
        arr = np.frombuffer(buf, np.uint8).reshape(pix.height, pix.stride)[:, :pix.width * 3].reshape(pix.height, pix.width, 3)
        self.gray = (0.299 * arr[:, :, 0] + 0.587 * arr[:, :, 1] + 0.114 * arr[:, :, 2]).astype(np.uint8)
        self._overlay = []
        self.canvas.set_image(qimg)
        self._reset_model(); self._update_page_label()
        self.lbl_scale.setText(f"Escala 1\"={self.scale*72:.0f}'")
        self._info(f"Página {idx + 1} cargada.")

    def _reset_model(self):
        self.cur_pts = []; self.pipes = []; self.leaders = []; self.text_marks = []
        self.erase_regions = []; self._erase_pts = []
        self.sel_pipe = self.sel_leader = self.sel_region = self.sel_text = -1
        self._overlay = []; self._close_editor(); self._dirty = False; self._extending = False
        self._undo.clear(); self._redo.clear(); self.ocr_boxes = []; self._tess_boxes = []; self._icr_boxes = []
        self.set_mode("idle"); self._refresh_lists(); self._redraw()

    # ─────────────────────────── proyecto ───────────────────────────
    def _write_project(self, path):
        self._busy("Guardando proyecto…")
        try:
            model = dict(pipes=self.pipes, leaders=self.leaders, text_marks=self.text_marks,
                         erase_regions=self.erase_regions,
                         tf=dict(scale=self.scale, zoom=self.zoom, rot=self.rot, W=self.W, H=self.H,
                                 derot=[self.derot.a, self.derot.b, self.derot.c, self.derot.d, self.derot.e, self.derot.f]),
                         pdf_name=os.path.basename(self.pdf_path or ""), version=VERSION)
            ba = QtCore.QByteArray(); buf = QtCore.QBuffer(ba); buf.open(QtCore.QIODevice.WriteOnly)
            self.canvas.pixmap_item.pixmap().save(buf, "PNG")
            with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
                z.writestr("model.json", json.dumps(model)); z.writestr("page.png", bytes(ba))
            self.project_path = path; self._dirty = False; self._info(f"Proyecto guardado: {os.path.basename(path)}")
        finally: self._unbusy()

    def save_project(self):
        if self.canvas.pixmap_item is None:
            QtWidgets.QMessageBox.information(self, "Nada que guardar", "Abre un PDF o proyecto primero."); return
        if self.project_path: self._write_project(self.project_path)
        else: self.save_project_as()

    def save_project_as(self):
        if self.canvas.pixmap_item is None:
            QtWidgets.QMessageBox.information(self, "Nada que guardar", "Abre un PDF o proyecto primero."); return
        base = self.project_path or os.path.join(DOWNLOADS, "proyecto.digproj")
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Guardar proyecto como", base, "Proyecto (*.digproj)")
        if path: self._write_project(path)

    def open_project(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Abrir proyecto", DOWNLOADS, "Proyecto (*.digproj)")
        if path: self._open_project_path(path)

    def _open_project_path(self, path):
        if not self._confirm_discard(): return
        self._busy("Abriendo proyecto…")
        try:
            with zipfile.ZipFile(path) as z:
                model = json.loads(z.read("model.json")); png = z.read("page.png")
            qimg = QtGui.QImage.fromData(png, "PNG")
            self._overlay = []; self._close_editor()
            self.canvas.set_image(qimg); self.gray = qimage_to_gray(qimg)
            tf = model["tf"]; self.scale = tf["scale"]; self.zoom = tf["zoom"]; self.rot = tf["rot"]
            self.W, self.H = tf["W"], tf["H"]; self.derot = fitz.Matrix(*tf["derot"])
            self.pageH_px = qimg.height()
            self.leader_hpx = max(14.0, min(LEADER_TEXT_FT / self.scale * self.zoom, self.pageH_px * 0.05))
            self.pdf_path = None; self.doc = None; self.project_path = path; self.page_idx = 0
            self.pipes = model.get("pipes", []); self.leaders = model.get("leaders", [])
            self.text_marks = model.get("text_marks", [])
            self.erase_regions = [r if isinstance(r, dict) else {"pts": r, "enabled": True}
                                  for r in model.get("erase_regions", [])]
            self.cur_pts = []; self._erase_pts = []; self.sel_pipe = self.sel_leader = self.sel_region = self.sel_text = -1
            self._undo.clear(); self._redo.clear(); self._dirty = False
            self.ocr_boxes = []; self._tess_boxes = []; self._icr_boxes = []
            self.set_mode("idle"); self._refresh_lists(); self._update_page_label(); self._redraw()
            self.lbl_scale.setText(f"Escala 1\"={self.scale*72:.0f}'")
            self._info(f"Proyecto abierto ({len(self.pipes)} utilidades). Ctrl+S guarda en este mismo archivo.")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", str(e))
        finally: self._unbusy()

    def _confirm_discard(self):
        if not self._dirty or self.canvas.pixmap_item is None: return True
        r = QtWidgets.QMessageBox.question(
            self, "Cambios sin guardar", "Hay cambios sin guardar. ¿Deseas guardarlos?",
            QtWidgets.QMessageBox.Save | QtWidgets.QMessageBox.Discard | QtWidgets.QMessageBox.Cancel)
        if r == QtWidgets.QMessageBox.Cancel: return False
        if r == QtWidgets.QMessageBox.Save:
            self.save_project(); return not self._dirty
        return True

    def close_project(self):
        if self.canvas.pixmap_item is None: return
        if not self._confirm_discard(): return
        self.canvas.scene().clear(); self.canvas.pixmap_item = None
        self.pdf_path = None; self.doc = None; self.project_path = None; self.gray = None
        self.pipes = []; self.leaders = []; self.text_marks = []; self.erase_regions = []
        self.cur_pts = []; self._erase_pts = []; self._overlay = []; self._close_editor()
        self.sel_pipe = self.sel_leader = self.sel_region = self.sel_text = -1
        self.ocr_boxes = []; self._tess_boxes = []; self._icr_boxes = []
        self._undo.clear(); self._redo.clear(); self._dirty = False; self.show_text_boxes = False
        for chk in (self.chk_txt, self.chk_icr):
            chk.blockSignals(True); chk.setChecked(False); chk.blockSignals(False)
        self.set_mode("idle"); self._refresh_lists(); self._update_page_label(); self._info("Proyecto cerrado.")

    def closeEvent(self, e):
        if self._confirm_discard(): e.accept()
        else: e.ignore()

    # ─────────────────────────── clics ───────────────────────────
    def on_click(self, x, y, button):
        if self.canvas.pixmap_item is None: return
        if self.show_text_boxes and button == QtCore.Qt.LeftButton and self._click_box(x, y): return
        if button == QtCore.Qt.RightButton:
            if self.mode == "pipe": self.finish_pipe()
            elif self.mode == "erase": self.finish_erase()
            elif self.mode == "move": self._delete_vertex(x, y)
            return
        if self.mode == "pipe":
            self._push(); self.cur_pts.append(self._snap(x, y)); self._update_ui(); self._redraw()
        elif self.mode == "erase":
            self._erase_pts.append((x, y)); self._update_ui(); self._redraw()
        elif self.mode == "text":
            if self._editor is not None: return   # el clic confirma el texto abierto (no abre otro)
            self._new_free_text(x, y)
        elif self.mode == "leader1":
            self._pending["arrow"] = self._snap(x, y); self.mode = "leader2"; self._update_ui()
        elif self.mode == "leader2":
            simple = bool(self._pending.get("simple"))
            if simple:
                hx, hy = self._pending["arrow"]; o = self.orient_combo.currentData()
                if o == "d":                                 # diagonal: 2º clic = inicio del landing (bisagra)
                    self._pending["landing"] = self._snap(x, y); self.mode = "leader3"; self._update_ui(); return
                tail = (x, hy) if o == "h" else (hx, y)      # h/v: 2º clic = final del cuerpo, recto al eje
                self._add_simple_leader((hx, hy), tail, o); return
            txt = self._read_leader_text()
            if not txt:
                self._info("Selecciona un texto (o activa 'texto personalizado') antes de colocar.")
                self._pending = {"arrow": None, "simple": False}
                self.mode = "leader1"; self._update_ui(); return
            self._push()
            self.leaders.append({"text": txt, "orient": self.orient_combo.currentData(),
                                 "simple": False,
                                 "arrow": self._pending["arrow"], "tp": (x, y),
                                 "font": self.font_combo.currentFont().family(),
                                 "size_ft": self.size_spin.value(), "bold": self.chk_bold.isChecked()})
            self._pending = {"arrow": None, "simple": False}; self.mode = "leader1"
            self._refresh_lists(); self._update_ui(); self._redraw()      # se ve al instante
            self._info("Multileader colocado. Clic en la PUNTA del siguiente (Esc para salir).")
        elif self.mode == "leader3":                         # Leader diagonal: 3er clic = final del cuerpo
            self._add_simple_leader(self._pending["arrow"], (x, y), "d",
                                    landing=self._pending.get("landing")); return
        elif self.mode == "idle":
            self._pick(x, y)

    def _add_simple_leader(self, head, tail, orient, landing=None):
        """Coloca un Leader simple (solo flecha, sin texto) y queda listo para el siguiente.
        orient 'h'/'v' → 2 clics (cabeza→final del cuerpo, recto al eje).
        orient 'd'     → 3 clics (cabeza → inicio del landing/bisagra → final del cuerpo)."""
        self._push()
        self.leaders.append({"text": "", "orient": orient, "simple": True,
                             "arrow": head, "tp": tail if tail else head, "landing": landing,
                             "font": self.font_combo.currentFont().family(),
                             "size_ft": self.size_spin.value(), "bold": self.chk_bold.isChecked()})
        self._pending = {"arrow": None, "simple": True}; self.mode = "leader1"
        self._refresh_lists(); self._update_ui(); self._redraw()
        self._info("Leader colocado. Clic en la cabeza de flecha del siguiente (Esc para salir).")

    def on_dblclick(self, x, y):
        if self.mode not in ("idle", "move"): return
        thr = 14.0 / max(1e-6, self.canvas.transform().m11())
        for i, ld in enumerate(self.leaders):
            if not (ld.get("arrow") and ld.get("tp")): continue
            geo = self._leader_geo(ld); lx, ly = geo["label_pos"]; H = geo["H"]
            lines = ld["text"].split("\n"); tw = max((len(s) for s in lines), default=1) * H * 0.55; th = len(lines) * H
            hit_text = (lx - 8 <= x <= lx + tw + 8 and ly - 8 <= y <= ly + th + 8)
            if hit_text or math.hypot(ld["tp"][0] - x, ld["tp"][1] - y) < thr:
                self._edit_leader_text(i); return
        for i, tm in enumerate(self.text_marks):
            if self._text_hit(tm, x, y):
                self._edit_text_mark(i); return

    def _text_hit(self, tm, x, y):
        h = self._px_for_ft(tm["size_ft"]) if "size_ft" in tm else tm.get("h", 16)
        lines = tm["text"].split("\n")
        w = max((len(s) for s in lines), default=1) * h * 0.55; th = len(lines) * h
        px, py = tm["pos"]; return px - 6 <= x <= px + w + 6 and py - 6 <= y <= py + th + 6

    def _pick(self, x, y):
        thr = 10.0 / max(1e-6, self.canvas.transform().m11())
        for i, tm in enumerate(self.text_marks):        # textos primero (blancos pequeños)
            if self._text_hit(tm, x, y):
                self._no_center = True; self.tabs.setCurrentIndex(TAB_TEXT); self.txt_marks_list.setCurrentRow(i)
                self._no_center = False; return
        for i, ld in enumerate(self.leaders):           # leaders/multileaders: por la línea o el texto
            if not (ld.get("arrow") and ld.get("tp")): continue
            geo = self._leader_geo(ld); hit = False
            for s in geo["segs"]:
                for a, b in zip(s, s[1:]):
                    if VP._pt_seg_dist(x, y, a[0], a[1], b[0], b[1]) < thr: hit = True; break
                if hit: break
            lx, ly = geo["label_pos"]; H = geo["H"]
            tw = max((len(t) for t in ld["text"].split("\n")), default=1) * H * 0.6; tt = ld["text"].count("\n") + 1
            if not hit and lx - 8 <= x <= lx + tw + 8 and ly - 8 <= y <= ly + tt * H + 8: hit = True
            if hit:
                self._select_leader(i); return
        best, bd = -1, thr
        for i, p in enumerate(self.pipes):
            for a, b in zip(p["pts"], p["pts"][1:]):
                d = VP._pt_seg_dist(x, y, a[0], a[1], b[0], b[1])
                if d < bd: bd, best = d, i
        if best >= 0:
            self._no_center = True; self.tabs.setCurrentIndex(TAB_PIPE); self.pipe_list.setCurrentRow(best)
            self._no_center = False

    def _snap(self, x, y):
        if not self.snap or self.gray is None: return (x, y)
        h, w = self.gray.shape; xi, yi, r = int(x), int(y), self.snap_r
        win = self.gray[max(0, yi - r):min(h, yi + r + 1), max(0, xi - r):min(w, xi + r + 1)] < 128
        if not win.any(): return (x, y)
        ys, xs = np.nonzero(win); return (max(0, xi - r) + xs.mean(), max(0, yi - r) + ys.mean())

    # ─────────────────────────── editar / mover ───────────────────────────
    def _current_kind(self):
        ti = self.tabs.currentIndex()
        if ti == TAB_PIPE and 0 <= self.sel_pipe < len(self.pipes): return "pipe"
        if ti == TAB_LEADER and 0 <= self.sel_leader < len(self.leaders) and self.leaders[self.sel_leader].get("simple"): return "leader"
        if ti == TAB_TEXT and 0 <= self.sel_text < len(self.text_marks): return "text"
        if ti == TAB_REGION and 0 <= self.sel_region < len(self.erase_regions): return "region"
        return None

    def enter_move(self):
        kind = self._current_kind()
        # Ctrl+T con un Leader simple seleccionado: asegura la pestaña Leaders y entra a editar sus vértices
        if not kind and 0 <= self.sel_leader < len(self.leaders) and self.leaders[self.sel_leader].get("simple"):
            self._select_leader(self.sel_leader, center=True); kind = "leader"
        if kind == "leader":
            self.set_mode("move"); self._info("Editar Leader: arrastra un vértice (posición/longitud) o el trazo para mover. Enter/Esc termina.")
        elif kind:
            self.set_mode("move"); self._info("Arrastra para mover · clic en vértice extiende (F) · clic derecho elimina")
        else:
            self._info("Selecciona primero una utilidad, Leader, texto o zona")

    def _thr(self): return 12.0 / max(1e-6, self.canvas.transform().m11())

    def _segments(self, pts, closed):
        segs = list(zip(range(len(pts) - 1), pts, pts[1:]))
        if closed and len(pts) >= 3: segs.append((len(pts) - 1, pts[-1], pts[0]))
        return segs

    def begin_move(self, x, y):
        kind = self._current_kind()
        if not kind: return
        self._push(); self._moved = False; self._move_kind = kind
        self._press_xy = (x, y); self._last_xy = (x, y); thr = self._thr()
        if kind == "text":
            self._move0 = (x, y); self._drag_vertex = None; self._edit_pts = None; return
        if kind == "leader":
            ld = self.leaders[self.sel_leader]; self._edit_leader = ld
            pts = [tuple(ld["arrow"])]
            if ld.get("landing"): pts.append(tuple(ld["landing"]))
            pts.append(tuple(ld["tp"]))
            self._edit_pts = pts; self._edit_closed = False
            vi, vd = -1, thr
            for i, (px, py) in enumerate(pts):
                d = math.hypot(px - x, py - y)
                if d < vd: vd, vi = d, i
            self._drag_vertex = vi if vi >= 0 else None       # vértice cercano → arrastra; si no, mueve todo
            self._move0 = None if vi >= 0 else (x, y); return
        pts = self.pipes[self.sel_pipe]["pts"] if kind == "pipe" else self.erase_regions[self.sel_region]["pts"]
        self._edit_pts = pts; self._edit_closed = (kind == "region")
        vi, vd = -1, thr
        for i, (px, py) in enumerate(pts):
            d = math.hypot(px - x, py - y)
            if d < vd: vd, vi = d, i
        if vi >= 0:
            self._drag_vertex = vi; self._move0 = None; return
        si, sd = -1, thr
        for idx, a, b in self._segments(pts, self._edit_closed):
            d = VP._pt_seg_dist(x, y, a[0], a[1], b[0], b[1])
            if d < sd: sd, si = d, idx
        if si >= 0:
            pts.insert(si + 1, (x, y)); self._drag_vertex = si + 1; self._move0 = None; self._moved = True
            self._refresh_lists(); return
        self._drag_vertex = None; self._move0 = (x, y)

    def do_move(self, x, y):
        self._moved = True; self._last_xy = (x, y)
        if self._move_kind == "text" and 0 <= self.sel_text < len(self.text_marks):
            if self._move0:
                dx, dy = x - self._move0[0], y - self._move0[1]; self._move0 = (x, y)
                px, py = self.text_marks[self.sel_text]["pos"]
                self.text_marks[self.sel_text]["pos"] = (px + dx, py + dy); self._redraw()
            return
        if self._move_kind == "leader":
            pts = self._edit_pts
            if pts is None: return
            if self._drag_vertex is not None:
                pts[self._drag_vertex] = (x, y)               # mueve un vértice (posición/longitud)
            elif self._move0 is not None:
                dx, dy = x - self._move0[0], y - self._move0[1]; self._move0 = (x, y)
                for i in range(len(pts)): pts[i] = (pts[i][0] + dx, pts[i][1] + dy)
            else:
                return
            self._sync_leader(); self._redraw(); return
        pts = self._edit_pts
        if pts is None: return
        if self._drag_vertex is not None:
            pts[self._drag_vertex] = (x, y); self._redraw(); return
        if self._move0 is not None:
            dx, dy = x - self._move0[0], y - self._move0[1]; self._move0 = (x, y)
            for i in range(len(pts)): pts[i] = (pts[i][0] + dx, pts[i][1] + dy)
            self._redraw()

    def end_move(self):
        # clic (casi sin arrastrar) sobre un vértice de una tubería → extender
        dist = math.hypot(self._last_xy[0] - self._press_xy[0], self._last_xy[1] - self._press_xy[1]) \
            if (self._last_xy and self._press_xy) else 0
        if (self._move_kind == "pipe" and self._drag_vertex is not None and dist < 6
                and 0 <= self.sel_pipe < len(self.pipes)):
            self._move0 = None; self._move_kind = None
            self._start_extension(self.sel_pipe, self._drag_vertex); self._drag_vertex = None; return
        self._move0 = None; self._drag_vertex = None; self._move_kind = None; self._edit_leader = None

    def _sync_leader(self):
        """Vuelca los vértices editados (self._edit_pts) al leader (arrow / landing / tp)."""
        ld = self._edit_leader; pts = self._edit_pts
        if not ld or not pts: return
        ld["arrow"] = pts[0]
        if ld.get("landing"): ld["landing"] = pts[1]; ld["tp"] = pts[2]
        else: ld["tp"] = pts[-1]

    def _start_extension(self, pi, vi):
        # Sin ventana emergente: la casilla 'continuar la misma utilidad' decide.
        pts = self.pipes[pi]["pts"]; vpos = tuple(pts[vi]); is_end = (vi == 0 or vi == len(pts) - 1)
        same = self.chk_ext_same.isChecked() and is_end
        self._extending = True; self._ext_layer = self.pipes[pi]["layer"]; self.cur_pts = [vpos]
        if same:
            self._ext_pipe = pi; self._ext_at = "start" if vi == 0 else "end"
            self._info("Continuando la MISMA utilidad: clic para agregar puntos, Enter finaliza.")
        else:
            self._ext_pipe = None; self._ext_at = None
            if self.chk_ext_same.isChecked() and not is_end:
                self._info("Solo desde un extremo se continúa; se creará una utilidad NUEVA. Clic para agregar, Enter finaliza.")
            else:
                self._info("Utilidad NUEVA (rama en F): clic para agregar puntos, Enter finaliza.")
        self.set_mode("pipe")

    def _delete_vertex(self, x, y):
        kind = self._current_kind()
        if kind not in ("pipe", "region"): return
        pts = self.pipes[self.sel_pipe]["pts"] if kind == "pipe" else self.erase_regions[self.sel_region]["pts"]
        floor = 3 if kind == "region" else 2
        if len(pts) <= floor: self._info(f"Necesita al menos {floor} puntos"); return
        thr = self._thr(); vi, vd = -1, thr
        for i, (px, py) in enumerate(pts):
            d = math.hypot(px - x, py - y)
            if d < vd: vd, vi = d, i
        if vi >= 0:
            self._push(); pts.pop(vi); self._refresh_lists(); self._redraw(); self._info("Vértice eliminado")

    # ─────────────────────────── utilidades ───────────────────────────
    def finish_pipe(self):
        if self._extending and self._ext_pipe is not None and 0 <= self._ext_pipe < len(self.pipes):
            extra = self.cur_pts[1:]                       # el 1er punto es el vértice existente
            if extra:
                self._push(); pts = self.pipes[self._ext_pipe]["pts"]
                if self._ext_at == "end": pts.extend(extra)
                else: self.pipes[self._ext_pipe]["pts"] = list(reversed(extra)) + pts
        elif len(self.cur_pts) >= 2:
            layer = self._ext_layer if self._extending else self.active_layer()
            ab = False if self._extending else self.chk_ab.isChecked()
            self._push(); self.pipes.append({"layer": layer, "pts": self.cur_pts[:], "ab": ab})
        self.cur_pts = []; self._extending = False; self._ext_layer = None; self._ext_pipe = None; self._ext_at = None
        self._refresh_lists(); self._update_ui(); self._redraw()

    def _sel_pipe(self, r):
        self.sel_pipe = r
        if 0 <= r < len(self.pipes):
            p = self.pipes[r]
            if not self._no_center:
                pts = p["pts"]; mid = pts[len(pts) // 2]; self.canvas.centerOn(mid[0], mid[1])
            self._prop_guard = True
            self.prop_name.setText(p.get("name", "")); self.prop_diam.setValue(p.get("diam", 0.0))
            self.prop_unit.setCurrentText(p.get("unit", "pulg")); self._prop_guard = False
        self._update_ui(); self._redraw()

    def _leader_at_row(self, lst, r):
        """Índice real en self.leaders del item de la fila r (o -1)."""
        it = lst.item(r) if r is not None and r >= 0 else None
        return it.data(QtCore.Qt.UserRole) if it is not None else -1

    def _sel_leader(self, r):                              # pestaña Multileaders
        i = self._leader_at_row(self.lead_list, r); self.sel_leader = i
        if 0 <= i < len(self.leaders):
            ld = self.leaders[i]
            if not self._no_center and ld.get("tp"): self.canvas.centerOn(ld["tp"][0], ld["tp"][1])
            self._style_guard = True
            self.font_combo.setCurrentFont(QtGui.QFont(ld.get("font", C.TEXT_FONT)))
            self.size_spin.setValue(ld.get("size_ft", LEADER_TEXT_FT)); self.chk_bold.setChecked(bool(ld.get("bold")))
            self._style_guard = False
        self._update_ui(); self._redraw()

    def _sel_sleader(self, r):                             # pestaña Leaders (solo flechas)
        i = self._leader_at_row(self.sleader_list, r); self.sel_leader = i
        if 0 <= i < len(self.leaders):
            ld = self.leaders[i]
            if not self._no_center and ld.get("tp"): self.canvas.centerOn(ld["tp"][0], ld["tp"][1])
        self._update_ui(); self._redraw()

    def _select_leader(self, i, center=False):
        """Selecciona el leader nº i en la pestaña que le corresponde (ML o Leaders)."""
        if not (0 <= i < len(self.leaders)): return
        simple = self.leaders[i].get("simple")
        lst = self.sleader_list if simple else self.lead_list
        row = next((r for r in range(lst.count()) if lst.item(r).data(QtCore.Qt.UserRole) == i), -1)
        self._no_center = not center
        self.tabs.setCurrentIndex(TAB_LEADER if simple else TAB_ML); lst.setCurrentRow(row)
        self._no_center = False

    def _sel_text(self, r):
        self.sel_text = r
        if 0 <= r < len(self.text_marks):
            tm = self.text_marks[r]
            if not self._no_center: self.canvas.centerOn(tm["pos"][0], tm["pos"][1])
            self._style_guard = True
            self.font_combo.setCurrentFont(QtGui.QFont(tm.get("font", C.TEXT_FONT)))
            self.size_spin.setValue(tm.get("size_ft", 3.0)); self.chk_bold.setChecked(bool(tm.get("bold")))
            self.rot_spin.setValue(int(tm.get("rot", 0)) % 360); self._style_guard = False
        self._update_ui(); self._redraw()

    def _sel_region(self, r):
        self.sel_region = r
        if not self._no_center and 0 <= r < len(self.erase_regions):
            pts = self.erase_regions[r]["pts"]
            cx = sum(p[0] for p in pts) / len(pts); cy = sum(p[1] for p in pts) / len(pts); self.canvas.centerOn(cx, cy)
        self._update_ui(); self._redraw()

    def _region_toggled(self, item):
        r = self.region_list.row(item)
        if 0 <= r < len(self.erase_regions):
            self.erase_regions[r]["enabled"] = (item.checkState() == QtCore.Qt.Checked); self._dirty = True; self._redraw()

    def _bump_size(self, delta):
        self.size_spin.setValue(round(max(0.5, self.size_spin.value() + delta), 2))

    def _bump_rot(self, delta):
        self.rot_spin.setValue((self.rot_spin.value() + delta) % 360)

    def _bump_opacity(self, delta_pct):
        val = self.canvas.pdf_opacity + delta_pct / 100.0
        self.canvas.set_pdf_opacity(val)
        self.lbl_opacity.setText(f"{round(self.canvas.pdf_opacity * 100)}%")

    def _px_for_ft(self, ft):
        raw = ft / self.scale * self.zoom if self.scale else ft * self.zoom
        cap = self.pageH_px * 0.06 if self.pageH_px else 200
        return max(8.0, min(raw, cap))

    def _style_changed(self):
        if self._style_guard: return
        ti = self.tabs.currentIndex()
        if ti == TAB_TEXT and 0 <= self.sel_text < len(self.text_marks):
            tm = self.text_marks[self.sel_text]; self._push()
            tm["font"] = self.font_combo.currentFont().family(); tm["size_ft"] = self.size_spin.value()
            tm["bold"] = self.chk_bold.isChecked(); tm["rot"] = self.rot_spin.value() % 360
            self._redraw()
        elif ti == TAB_ML and 0 <= self.sel_leader < len(self.leaders):
            ld = self.leaders[self.sel_leader]; self._push()
            ld["font"] = self.font_combo.currentFont().family(); ld["size_ft"] = self.size_spin.value()
            ld["bold"] = self.chk_bold.isChecked()
            self._redraw()

    def _prop_changed(self):
        if self._prop_guard: return
        if self.tabs.currentIndex() == TAB_PIPE and 0 <= self.sel_pipe < len(self.pipes):
            p = self.pipes[self.sel_pipe]; self._push()
            p["name"] = self.prop_name.text().strip(); p["diam"] = self.prop_diam.value()
            p["unit"] = self.prop_unit.currentText(); self._refresh_lists()

    def change_pipe_type(self):
        if 0 <= self.sel_pipe < len(self.pipes):
            self._push(); self.pipes[self.sel_pipe]["layer"] = self.active_layer()
            self.pipes[self.sel_pipe]["ab"] = self.chk_ab.isChecked(); self._refresh_lists(); self._redraw()

    def edit_selected_text(self):
        ti = self.tabs.currentIndex()
        if ti == TAB_ML and 0 <= self.sel_leader < len(self.leaders): self._edit_leader_text(self.sel_leader)
        elif ti == TAB_TEXT and 0 <= self.sel_text < len(self.text_marks): self._edit_text_mark(self.sel_text)

    def _list_context_menu(self, listw, tab_idx, pos):
        item = listw.itemAt(pos)
        if item is None: return
        if self.tabs.currentIndex() != tab_idx: self.tabs.setCurrentIndex(tab_idx)
        listw.setCurrentRow(listw.row(item))          # selecciona la fila bajo el cursor
        menu = QtWidgets.QMenu(self)
        if tab_idx == TAB_PIPE:
            self._menu_act(menu, "Cambiar tipo", self.change_pipe_type)
            self._menu_act(menu, "Editar/mover", self.enter_move)
        elif tab_idx == TAB_ML:
            self._menu_act(menu, "Editar texto", self.edit_selected_text)
        elif tab_idx == TAB_LEADER:
            self._menu_act(menu, "Editar/mover", self.enter_move)
        elif tab_idx == TAB_TEXT:
            self._menu_act(menu, "Mover", self.enter_move)
            self._menu_act(menu, "Editar texto", self.edit_selected_text)
        elif tab_idx == TAB_REGION:
            self._menu_act(menu, "Editar/mover", self.enter_move)
        menu.addSeparator()
        self._menu_act(menu, "Eliminar", self.delete_selected)
        menu.exec(listw.viewport().mapToGlobal(pos))

    def delete_selected(self):
        ti = self.tabs.currentIndex()
        if ti == TAB_PIPE and 0 <= self.sel_pipe < len(self.pipes):
            self._push(); self.pipes.pop(self.sel_pipe); self.sel_pipe = -1
        elif ti in (TAB_ML, TAB_LEADER) and 0 <= self.sel_leader < len(self.leaders):
            self._push(); self.leaders.pop(self.sel_leader); self.sel_leader = -1
        elif ti == TAB_TEXT and 0 <= self.sel_text < len(self.text_marks):
            self._push(); self.text_marks.pop(self.sel_text); self.sel_text = -1
        elif ti == TAB_REGION and 0 <= self.sel_region < len(self.erase_regions):
            self._push(); self.erase_regions.pop(self.sel_region); self.sel_region = -1
        self._refresh_lists(); self._redraw()

    def _copy_sel(self):
        ti = self.tabs.currentIndex()
        if ti in (TAB_ML, TAB_LEADER) and 0 <= self.sel_leader < len(self.leaders): self._clip = ("leader", copy.deepcopy(self.leaders[self.sel_leader]))
        elif ti == TAB_PIPE and 0 <= self.sel_pipe < len(self.pipes): self._clip = ("pipe", copy.deepcopy(self.pipes[self.sel_pipe]))
        elif ti == TAB_TEXT and 0 <= self.sel_text < len(self.text_marks): self._clip = ("text", copy.deepcopy(self.text_marks[self.sel_text]))
        else: self._info("Selecciona algo para copiar."); return
        self._info("Copiado. Ctrl+V para pegar una copia.")

    def _paste_sel(self):
        if not self._clip: return
        kind, obj = self._clip; o = copy.deepcopy(obj); d = 30; self._push()
        if kind == "leader":
            if o.get("arrow"): o["arrow"] = (o["arrow"][0] + d, o["arrow"][1] + d)
            if o.get("landing"): o["landing"] = (o["landing"][0] + d, o["landing"][1] + d)
            if o.get("tp"): o["tp"] = (o["tp"][0] + d, o["tp"][1] + d)
            self.leaders.append(o); self._refresh_lists(); self._select_leader(len(self.leaders) - 1)
        elif kind == "pipe":
            o["pts"] = [(x + d, y + d) for (x, y) in o["pts"]]; self.pipes.append(o)
            self.tabs.setCurrentIndex(TAB_PIPE); self._refresh_lists(); self.pipe_list.setCurrentRow(len(self.pipes) - 1)
        elif kind == "text":
            o["pos"] = (o["pos"][0] + d, o["pos"][1] + d); self.text_marks.append(o)
            self.tabs.setCurrentIndex(TAB_TEXT); self._refresh_lists(); self.txt_marks_list.setCurrentRow(len(self.text_marks) - 1)
        self._redraw(); self._info("Pegado (copia desplazada).")

    def _refresh_lists(self):
        self.pipe_list.blockSignals(True); self.pipe_list.clear()
        for i, p in enumerate(self.pipes, 1):
            tag = " (AB)" if p.get("ab") else ""
            nm = f" · {p['name']}" if p.get("name") else ""
            it = QtWidgets.QListWidgetItem(swatch_icon(layer_qcolor(p["layer"])), f"{i}. {p['layer']}{tag}{nm} ({len(p['pts'])})")
            it.setForeground(QtGui.QColor("white")); self.pipe_list.addItem(it)
        self.pipe_list.blockSignals(False)
        self.lead_list.blockSignals(True); self.lead_list.clear()
        self.sleader_list.blockSignals(True); self.sleader_list.clear()
        nml = ns = 0
        for i, ld in enumerate(self.leaders):
            if ld.get("simple"):
                ns += 1; o = {"h": "horizontal", "v": "vertical", "d": "diagonal"}.get(ld.get("orient", "d"), "")
                it = QtWidgets.QListWidgetItem(f"{ns}. Leader {o}".rstrip())
                it.setData(QtCore.Qt.UserRole, i); self.sleader_list.addItem(it)
            else:
                nml += 1; it = QtWidgets.QListWidgetItem(f"{nml}. {ld['text'][:24].replace(chr(10), ' / ')}")
                it.setData(QtCore.Qt.UserRole, i); self.lead_list.addItem(it)
        self.lead_list.blockSignals(False); self.sleader_list.blockSignals(False)
        self.txt_marks_list.blockSignals(True); self.txt_marks_list.clear()
        for i, tm in enumerate(self.text_marks, 1): self.txt_marks_list.addItem(f"{i}. {tm['text'][:28].replace(chr(10), ' / ')}")
        self.txt_marks_list.blockSignals(False)
        self.region_list.blockSignals(True); self.region_list.clear()
        for i, rg in enumerate(self.erase_regions, 1):
            it = QtWidgets.QListWidgetItem(f"Zona {i} ({len(rg['pts'])} vértices)")
            it.setFlags(it.flags() | QtCore.Qt.ItemIsUserCheckable)
            it.setCheckState(QtCore.Qt.Checked if rg.get("enabled", True) else QtCore.Qt.Unchecked)
            it.setForeground(QtGui.QColor("white")); self.region_list.addItem(it)
        self.region_list.blockSignals(False)

    # ─────────────────────────── borrar zona ───────────────────────────
    def finish_erase(self):
        if len(self._erase_pts) < 3: return
        self._push(); poly = self._erase_pts[:]; self.erase_regions.append({"pts": poly, "enabled": True})
        self.pipes = [p for p in self.pipes if not all(point_in_poly(px, py, poly) for (px, py) in p["pts"])]
        self.leaders = [l for l in self.leaders if not (l.get("tp") and point_in_poly(l["tp"][0], l["tp"][1], poly))]
        self.text_marks = [t for t in self.text_marks if not point_in_poly(t["pos"][0], t["pos"][1], poly)]
        self._erase_pts = []; self.set_mode("idle"); self._refresh_lists()
        self._info("Zona agregada: al exportar también se borra la geometría base del plano dentro de ella.")

    # ─────────────────────────── Multileader ───────────────────────────
    def _read_leader_text(self):
        if self.chk_custom.isChecked(): return self.txt_edit.text().strip()
        it = self.text_list.currentItem()
        return it.text() if it and self.text_list.currentRow() >= 0 else ""

    def start_leader(self, simple=False):
        # No se captura nada al entrar: la orientación y el texto se leen al COLOCAR.
        # simple=True → Leader (solo flecha, sin texto); False → Multileader.
        self._pending = {"arrow": None, "simple": bool(simple)}
        self.set_mode("leader1")
        if simple:
            if self.orient_combo.currentData() == "d":
                self._info("Leader diagonal: cabeza → inicio del landing (bisagra) → final del cuerpo. Enter/Esc para salir.")
            else:
                self._info("Leader: cabeza de flecha → final del cuerpo. Enter/Esc para salir.")
        else:
            self._info("Multileader: elige el texto; clic en la PUNTA y luego dónde va el TEXTO. Esc para salir.")

    def _edit_leader_text(self, idx):
        ld = self.leaders[idx]; tp = ld["tp"]
        if ld.get("simple"):                                 # el Leader simple no tiene texto que editar
            self._info("El Leader simple no lleva texto."); return
        def commit(val):
            self._close_editor()
            if val.strip(): self._push(); ld["text"] = val.rstrip("\n"); self._refresh_lists()
            self._redraw()
        self._open_editor(tp[0], tp[1] - self.leader_hpx, ld["text"], commit)

    def _leader_geo(self, ld):
        """Geometría del Multileader (px). La 'cola' (parte de la línea junto al texto)
        se adapta al largo del texto. La punta se orienta con segs[0][1]."""
        ax, ay = ld["arrow"]; tx, ty = ld["tp"]
        ftsize = ld.get("size_ft", LEADER_TEXT_FT)
        H = self._px_for_ft(ftsize)
        lines = ld["text"].split("\n"); maxlen = max((len(s) for s in lines), default=1); nlines = len(lines)
        tw = max(maxlen * H * 0.6, H * 2); th = nlines * H; gap = H * 0.5; near = H * 0.22
        if ld.get("simple"):                               # LEADER simple: solo flecha, sin texto
            lp = ld.get("landing")
            if lp:                                         # diagonal con landing: cabeza → bisagra → final
                segs = [[(ax, ay), (lp[0], lp[1]), (tx, ty)]]
            else:                                          # recto h/v: cabeza → final del cuerpo
                segs = [[(ax, ay), (tx, ty)]]
            end = segs[0][-1]
            return dict(segs=segs, label_pos=end, rot=0, side="right", verts_px=segs[0], insert_px=end,
                        dogleg=0.0, H=H, cad_h=ftsize, tcenter_px=end, cad_rot=0)
        orient = ld.get("orient", "h")
        if orient == "v":                                  # recto vertical; texto vertical junto a la cola
            signY = -1 if ty < ay else 1
            L = max(abs(ty - ay), tw + gap); ey = ay + signY * L; my = (ay + ey) / 2
            side = "top" if signY < 0 else "bottom"
            lbl = (ax + H * 0.08, my + tw / 2)              # rot -90, centrado a lo largo, pegado a la línea
            segs = [[(ax, ay), (ax, ey)]]
            return dict(segs=segs, label_pos=lbl, rot=-90, side=side, verts_px=segs[0], insert_px=(ax, ey),
                        dogleg=0.0, H=H, cad_h=ftsize, tcenter_px=(ax + th / 2 + H * 0.08, my), cad_rot=90)
        if orient == "h":                                  # recto horizontal; texto encima de la cola
            signX = 1 if tx >= ax else -1
            L = max(abs(tx - ax), tw + gap); ex = ax + signX * L
            lblx = ex - tw if signX > 0 else ex
            side = "right" if signX > 0 else "left"
            segs = [[(ax, ay), (ex, ay)]]
            return dict(segs=segs, label_pos=(lblx, ay - th - near), rot=0, side=side, verts_px=segs[0], insert_px=(ex, ay),
                        dogleg=0.0, H=H, cad_h=ftsize, tcenter_px=(lblx + tw / 2, ay - near - th / 2), cad_rot=0)
        # diagonal: flecha → 2º clic → landing horizontal → texto encima del landing
        right = tx >= ax; sgn = 1 if right else -1
        lx = tx + sgn * (tw + gap); text_x = min(tx, lx) + gap
        side = "right" if right else "left"
        lbl = (text_x, ty - H - near)                      # 1ª línea encima; extras al otro lado
        segs = [[(ax, ay), (tx, ty), (lx, ty)]]
        return dict(segs=segs, label_pos=lbl, rot=0, side=side, verts_px=segs[0], insert_px=(lx, ty),
                    dogleg=0.0, H=H, cad_h=ftsize, tcenter_px=(text_x + tw / 2, ty - H - near + th / 2), cad_rot=0)

    # ─────────────────────────── texto libre ───────────────────────────
    def _new_free_text(self, x, y):
        def commit(val):
            self._close_editor()
            if val.strip():
                self._push()
                self.text_marks.append({"pos": (x, y), "text": val.rstrip("\n"),
                                        "size_ft": self.size_spin.value(), "font": self.font_combo.currentFont().family(),
                                        "bold": self.chk_bold.isChecked(), "rot": self.rot_spin.value() % 360, "free": True})
                self._refresh_lists()
            self._redraw()
        self._open_editor(x, y, "", commit)

    def _edit_text_mark(self, idx):
        tm = self.text_marks[idx]
        def commit(val):
            self._close_editor()
            if val.strip(): self._push(); tm["text"] = val.rstrip("\n"); self._refresh_lists()
            self._redraw()
        self._open_editor(tm["pos"][0], tm["pos"][1], tm["text"], commit)

    def _open_editor(self, x, y, initial, on_commit, w=220):
        # Caja flotante hija del viewport (no escala con el zoom y recibe el teclado
        # de forma fiable: Enter aplica, Ctrl+Shift+Enter salta de línea, clic fuera aplica).
        self._close_editor()
        ed = InlineEdit(initial); ed.setParent(self.canvas.viewport())
        ed.setFixedWidth(w); ed.setFixedHeight(64)
        vp = self.canvas.mapFromScene(QtCore.QPointF(x, y))
        ed.move(vp); ed.show(); ed.raise_()
        ed.committed.connect(on_commit); self._editor = ed
        ed.setFocus(QtCore.Qt.OtherFocusReason); ed.selectAll()

    def _close_editor(self):
        if self._editor:
            ed = self._editor; self._editor = None
            try: ed.hide(); ed.deleteLater()
            except Exception: pass

    # ─────────────────────────── OCR / ICR ───────────────────────────
    def _sync_boxes(self):
        self.ocr_boxes = self._tess_boxes + self._icr_boxes
        self.show_text_boxes = self.chk_txt.isChecked() or self.chk_icr.isChecked()
        self._redraw()

    def toggle_text_boxes(self, on):
        if on and not self._tess_boxes and self.gray is not None:
            self._prog = QtWidgets.QProgressDialog("Detectando texto impreso (0/90/270°)…", None, 0, 0, self)
            self._prog.setWindowTitle("OCR"); self._prog.setWindowModality(QtCore.Qt.WindowModal)
            self._prog.setCancelButton(None); self._prog.show()
            self._ocr = OcrWorker(self.gray); self._ocr.done.connect(self._ocr_done); self._ocr.start()
        else: self._sync_boxes()

    def _ocr_done(self, boxes, err):
        if getattr(self, "_prog", None): self._prog.close()
        if err: self._info(f"OCR no disponible: {err}"); return
        self._tess_boxes = [(QtCore.QRectF(x, y, w, h), t) for (x, y, w, h, t) in boxes]
        self._info(f"{len(self._tess_boxes)} zonas de texto impreso. Clic en una para corregirla."); self._sync_boxes()

    def toggle_icr(self, on):
        if on and not self._icr_boxes and self.gray is not None:
            self._prog = QtWidgets.QProgressDialog("Reconociendo manuscrita (EasyOCR)…\n"
                                                   "La 1ª vez descarga el modelo, puede tardar.", None, 0, 0, self)
            self._prog.setWindowTitle("ICR"); self._prog.setWindowModality(QtCore.Qt.WindowModal)
            self._prog.setCancelButton(None); self._prog.show()
            self._icr = IcrWorker(self.gray); self._icr.done.connect(self._icr_done); self._icr.start()
        else: self._sync_boxes()

    def _icr_done(self, boxes, err):
        if getattr(self, "_prog", None): self._prog.close()
        if err == "missing":
            self.chk_icr.blockSignals(True); self.chk_icr.setChecked(False); self.chk_icr.blockSignals(False)
            QtWidgets.QMessageBox.information(self, "EasyOCR no instalado",
                "Para la ICR manuscrita (offline) instala EasyOCR con tu Python 3.12:\n\n"
                r'C:\Users\Deyvy\AppData\Local\Programs\Python\Python312\python.exe -m pip install easyocr'
                "\n\nLa primera vez descargará el modelo (~100 MB) automáticamente.")
            return
        if err:
            self.chk_icr.blockSignals(True); self.chk_icr.setChecked(False); self.chk_icr.blockSignals(False)
            self._info(f"ICR no disponible: {err}"); return
        self._icr_boxes = [(QtCore.QRectF(x, y, w, h), t) for (x, y, w, h, t) in boxes]
        self._info(f"{len(self._icr_boxes)} zonas manuscritas. Clic en una para corregirla."); self._sync_boxes()

    def _click_box(self, x, y):
        for rect, txt in self.ocr_boxes:
            if rect.contains(x, y):
                def commit(val):
                    self._close_editor()
                    if val.strip():
                        self._push()
                        self.text_marks.append({"pos": (rect.center().x(), rect.center().y()), "text": val.rstrip("\n"),
                                                "h": rect.height(), "box": (rect.x(), rect.y(), rect.width(), rect.height())})
                        self._refresh_lists()
                    self._redraw()
                self._open_editor(rect.x(), rect.y(), txt, commit, max(160, int(rect.width())))
                return True
        return False

    # ─────────────────────────── dibujo ───────────────────────────
    def _redraw(self):
        sc = self.canvas.scene()
        for it in self._overlay:
            try: sc.removeItem(it)
            except (RuntimeError, ValueError): pass
        self._overlay = []
        if self.canvas.pixmap_item is None: return
        # zonas de borrado — DETRÁS de todo (solo tapan el PDF)
        if self._erase_pts:
            self._poly(self._erase_pts, QtGui.QColor(255, 220, 0), 1.8, dots=True, z=Z_MARK)
        for i, rg in enumerate(self.erase_regions):
            qp = QtGui.QPolygonF([QtCore.QPointF(px, py) for (px, py) in rg["pts"]])
            enabled = rg.get("enabled", True); sel = (i == self.sel_region)
            if not enabled:
                pen = QtGui.QPen(QtGui.QColor(150, 150, 150), 1.2, QtCore.Qt.DashLine); brush = QtGui.QBrush(QtGui.QColor(200, 200, 200, 30))
            elif sel:
                pen = QtGui.QPen(QtGui.QColor(255, 40, 40), 2.0); brush = QtGui.QBrush(QtGui.QColor(255, 255, 255, 128))
            else:
                pen = QtGui.QPen(QtGui.QColor(255, 255, 255), 1.6); brush = QtGui.QBrush(QtGui.QColor(255, 255, 255, 255))
            pen.setCosmetic(True); it = sc.addPolygon(qp, pen, brush); it.setZValue(Z_ERASE); self._overlay.append(it)
            if sel and self.mode == "move": self._handles(rg["pts"])
        # utilidades
        for i, p in enumerate(self.pipes):
            sel = (i == self.sel_pipe)
            self._poly(p["pts"], layer_qcolor(p["layer"]), 4.0 if sel else 2.0, z=Z_MARK)
            if sel and self.mode == "move": self._handles(p["pts"])
        self._poly(self.cur_pts, layer_qcolor(self._ext_layer or self.active_layer()), 2.0, dots=True, z=Z_MARK)
        # multileaders
        anno = aci_qcolor(8)
        for i, ld in enumerate(self.leaders):
            if not (ld.get("arrow") and ld.get("tp")): continue
            col = QtGui.QColor(120, 220, 120) if i == self.sel_leader else anno
            geo = self._leader_geo(ld)
            for s in geo["segs"]: self._poly(s, col, 1.6, z=Z_MARK)
            self._arrow(ld["arrow"], geo["segs"][0][1], col)   # punta orientada a lo largo de la línea
            if i == self.sel_leader and self.mode == "move" and ld.get("simple"):
                self._handles(geo["segs"][0])                  # vértices editables (cabeza / bisagra / final)
            if ld["text"]:                                     # Leader simple no lleva texto
                t = sc.addText(ld["text"]); t.setDefaultTextColor(col); t.document().setDocumentMargin(0)
                f = t.font(); f.setPixelSize(int(geo["H"])); t.setFont(f)
                if geo["rot"]: t.setRotation(geo["rot"])
                t.setPos(geo["label_pos"][0], geo["label_pos"][1]); t.setZValue(Z_MARK); self._overlay.append(t)
        # cajas OCR/ICR
        if self.show_text_boxes:
            pen = QtGui.QPen(QtGui.QColor(255, 200, 0), 1); pen.setCosmetic(True)
            for rect, _ in self.ocr_boxes:
                it = sc.addRect(rect, pen); it.setZValue(Z_MARK); self._overlay.append(it)
        # textos
        for i, tm in enumerate(self.text_marks):
            t = sc.addText(tm["text"]); t.setDefaultTextColor(QtGui.QColor(120, 220, 120)); t.document().setDocumentMargin(0)
            hpx = self._px_for_ft(tm["size_ft"]) if "size_ft" in tm else tm.get("h", 16)
            f = t.font(); f.setPixelSize(max(6, int(hpx)))
            if tm.get("font"): f.setFamily(tm["font"])
            f.setBold(bool(tm.get("bold"))); t.setFont(f)
            t.setPos(tm["pos"][0], tm["pos"][1])
            if tm.get("rot"): t.setRotation(-tm["rot"])       # rot en grados CCW; Qt gira en sentido horario
            t.setZValue(Z_MARK); self._overlay.append(t)
            if i == self.sel_text and self.tabs.currentIndex() == TAB_TEXT:
                br = t.boundingRect(); pen = QtGui.QPen(QtGui.QColor(255, 180, 40)); pen.setCosmetic(True)
                rit = sc.addRect(tm["pos"][0], tm["pos"][1], br.width(), br.height(), pen); rit.setZValue(Z_MARK); self._overlay.append(rit)

    def _handles(self, pts):
        sc = self.canvas.scene(); pen = QtGui.QPen(QtGui.QColor(255, 255, 255)); pen.setCosmetic(True)
        for (vx, vy) in pts:
            it = sc.addRect(vx - 5, vy - 5, 10, 10, pen, QtGui.QBrush(QtGui.QColor(255, 180, 40)))
            it.setZValue(Z_HANDLE); self._overlay.append(it)

    def _poly(self, pts, color, width, dots=False, z=Z_MARK):
        sc = self.canvas.scene(); pen = QtGui.QPen(color, width); pen.setCosmetic(True)
        for a, b in zip(pts, pts[1:]):
            it = sc.addLine(a[0], a[1], b[0], b[1], pen); it.setZValue(z); self._overlay.append(it)
        if dots:
            for (x, y) in pts:
                it = sc.addEllipse(x - 3, y - 3, 6, 6, pen, QtGui.QBrush(color)); it.setZValue(z); self._overlay.append(it)

    def _arrow(self, a, b, color):
        ang = math.atan2(a[1] - b[1], a[0] - b[0]); L = self.leader_hpx * 0.8
        p1 = (a[0] - L * math.cos(ang - 0.4), a[1] - L * math.sin(ang - 0.4))
        p2 = (a[0] - L * math.cos(ang + 0.4), a[1] - L * math.sin(ang + 0.4))
        poly = QtGui.QPolygonF([QtCore.QPointF(*a), QtCore.QPointF(*p1), QtCore.QPointF(*p2)])
        it = self.canvas.scene().addPolygon(poly, QtGui.QPen(color), QtGui.QBrush(color)); it.setZValue(Z_MARK); self._overlay.append(it)

    # ─────────────────────────── coords ───────────────────────────
    def _to_cad(self, x, y):
        mp = fitz.Point(x / self.zoom, y / self.zoom) * self.derot
        px, py = mp.x, mp.y; s, r, W, H = self.scale, self.rot, self.W, self.H
        if r == 0: return (px * s, (H - py) * s)
        if r == 90: return (py * s, px * s)
        if r == 180: return ((W - px) * s, py * s)
        return (py * s, px * s)

    # ─────────────────────────── exportar ───────────────────────────
    def run_pipeline(self, mode="todo"):
        """mode: 'todo' = PDF digitalizado + anotaciones · 'pdf' = solo el PDF ·
        'anot' = solo las anotaciones dibujadas en el programa."""
        if self.canvas.pixmap_item is None:
            QtWidgets.QMessageBox.information(self, "Nada", "Abre un PDF o proyecto."); return
        need_pdf = mode in ("todo", "pdf")
        if need_pdf and not self.pdf_path:
            QtWidgets.QMessageBox.information(self, "Sin PDF",
                "No hay un PDF cargado para digitalizar. Se exportarán solo las anotaciones."); mode = "anot"; need_pdf = False
        base = os.path.splitext(os.path.basename(self.pdf_path))[0] if self.pdf_path else "proyecto"
        suffix = {"todo": "_completo", "pdf": "_plano", "anot": "_anotaciones"}[mode]
        out, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Guardar DXF", os.path.join(DOWNLOADS, base + suffix + ".dxf"), "DXF (*.dxf)")
        if not out: return
        self._out = out; self._mode = mode
        if not need_pdf:                                   # solo anotaciones: sin pipeline
            try:
                doc = ezdxf.new("R2010", setup=True); doc.header["$INSUNITS"] = C.INSUNITS
                self._merge_into(doc, marks=True); doc.saveas(out)
                QtWidgets.QMessageBox.information(self, "Listo", f"Exportado (solo anotaciones):\n{out}")
                self._info("DXF de anotaciones exportado.")
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, "Error", str(e))
            return
        self._tmp = out + ".base.tmp.dxf"
        self._prog = QtWidgets.QProgressDialog("Digitalizando el plano…", None, 0, 0, self)
        self._prog.setWindowTitle("Procesando"); self._prog.setWindowModality(QtCore.Qt.WindowModal)
        self._prog.setCancelButton(None); self._prog.show()
        self._worker = PipelineWorker(self.pdf_path, self._tmp); self._worker.done.connect(self._pipeline_done); self._worker.start()

    def _pipeline_done(self, tmp, err):
        if getattr(self, "_prog", None): self._prog.close()
        if err: QtWidgets.QMessageBox.critical(self, "Error al digitalizar", err); return
        try:
            marks = (self._mode == "todo")               # 'pdf' = solo el plano, sin anotaciones
            doc = ezdxf.readfile(tmp); self._merge_into(doc, marks=marks); doc.saveas(self._out)
            if os.path.exists(tmp): os.remove(tmp)
            if marks:
                nreg = sum(1 for r in self.erase_regions if r.get("enabled", True))
                msg = (f"Exportado (PDF + anotaciones):\n{self._out}\n\n{len(self.pipes)} utilidades, "
                       f"{len(self.leaders)} leaders/multileaders, {len(self.text_marks)} textos, {nreg} zonas borradas.")
            else:
                msg = f"Exportado (solo el PDF digitalizado):\n{self._out}"
            QtWidgets.QMessageBox.information(self, "Listo", msg); self._info("DXF exportado.")
        except Exception as e:
            import traceback; QtWidgets.QMessageBox.critical(self, "Error al guardar", f"{e}\n{traceback.format_exc()}")

    def _text_style(self, doc, font, bold):
        name = f"TXT_{font}_{'B' if bold else 'N'}".replace(" ", "_")[:60]
        if name not in doc.styles:
            try: doc.styles.add(name, font=font)
            except Exception: return "CAD_TEXT"
        return name

    def _merge_into(self, doc, marks=True):
        VP.setup_linetypes(doc); msp = doc.modelspace()
        self._apply_erase(msp)                            # las zonas de borrado recortan el plano base
        if not marks:                                     # 'solo PDF': no agregar utilidades/leaders/textos
            return
        if "PDFCAD" not in doc.appids: doc.appids.add("PDFCAD")   # para XDATA de propiedades
        for p in self.pipes:
            layer = p["layer"]; VP.ensure_layer(doc, layer)
            # El linetype con letra (─ W ─, ─ SS ─…) se aplica SOLO a la entidad que dibujas,
            # NO a la capa: así el contenido del plano base (en la misma capa) no se restilea.
            lt = C.LAYER_LINETYPE_AB.get(layer) if p.get("ab") else C.LAYER_LINETYPE.get(layer)
            att = {"layer": layer}
            if lt and lt in doc.linetypes: att["linetype"] = lt
            poly = msp.add_lwpolyline([self._to_cad(x, y) for (x, y) in p["pts"]], dxfattribs=att)
            # Propiedades de la utilidad como dato (XDATA)
            if p.get("name") or p.get("diam"):
                poly.set_xdata("PDFCAD", [(1000, f"NOMBRE={p.get('name', '')}"),
                                          (1000, f"DIAMETRO={p.get('diam', 0)}"),
                                          (1000, f"UNIDAD={p.get('unit', '')}")])
        VP.ensure_layer(doc, "ANOTACION")
        if "CAD_TEXT" not in doc.styles: doc.styles.add("CAD_TEXT", font=C.TEXT_FONT)
        for ld in self.leaders:
            if ld.get("arrow") and ld.get("tp"): self._add_leader(doc, msp, ld)
        for tm in self.text_marks:
            if tm.get("free"): self._add_free_text(doc, msp, tm)
            else: self._replace_text(doc, msp, tm)

    def _apply_erase(self, msp):
        regions = [r for r in self.erase_regions if r.get("enabled", True)]
        if not regions: return
        polys = [[self._to_cad(px, py) for (px, py) in r["pts"]] for r in regions]
        def inside(pt): return any(point_in_poly(pt[0], pt[1], poly) for poly in polys)
        for e in list(msp):
            t = e.dxftype()
            try:
                if t == "LWPOLYLINE":
                    pts = [(p[0], p[1]) for p in e.get_points()]
                    hit = inside((sum(p[0] for p in pts) / len(pts), sum(p[1] for p in pts) / len(pts)))
                elif t == "LINE":
                    a, b = e.dxf.start, e.dxf.end; hit = inside(((a.x + b.x) / 2, (a.y + b.y) / 2))
                elif t in ("TEXT", "MTEXT"):
                    ins = e.dxf.insert; hit = inside((ins.x, ins.y))
                elif t in ("CIRCLE", "ARC"):
                    c = e.dxf.center; hit = inside((c.x, c.y))
                else: hit = False
                if hit: msp.delete_entity(e)
            except Exception: continue

    def _add_leader(self, doc, msp, ld):
        geo = self._leader_geo(ld)
        if ld.get("simple"):                                 # Leader simple → entidad LEADER nativa
            if self._add_simple_leader_dxf(doc, msp, ld, geo): return
        else:                                                # Multileader → entidad MULTILEADER nativa
            if self._add_multileader_dxf(doc, msp, ld, geo): return
        # si el visor/plantilla no soporta la entidad nativa, se dibuja explícito como respaldo
        self._add_leader_explicit(doc, msp, ld, geo)

    def _mleader_style(self, doc, arrow_ft, char_ft):
        """Estilo MLEADER propio con el tamaño de flecha/altura de texto dados (pies)."""
        name = f"PDFCAD_ML_{int(round(arrow_ft * 10))}_{int(round(char_ft * 10))}"
        try:
            if name not in doc.mleader_styles:
                doc.mleader_styles.duplicate_entry("Standard", name)
            st = doc.mleader_styles.get(name)
            st.dxf.arrow_head_size = arrow_ft; st.dxf.char_height = char_ft
            return name
        except Exception:
            return "Standard"

    def _add_multileader_dxf(self, doc, msp, ld, geo):
        """Exporta el Multileader como entidad MULTILEADER nativa (flecha + directriz + texto).
        Devuelve True si se creó."""
        if not ld.get("text"): return False
        try:
            from ezdxf.math import Vec2
            from ezdxf.render.mleader import ConnectionSide, TextAlignment
        except Exception:
            return False
        ch = max(0.1, geo.get("cad_h", LEADER_TEXT_FT))
        asz = max(0.05, ch * 0.6)
        style = self._mleader_style(doc, asz, ch)
        # segs[0] va [punta(flecha) … landing]; el MULTILEADER quiere [insert(landing) … punta]
        cad_v = [self._to_cad(x, y) for (x, y) in geo["segs"][0]][::-1]
        if len(cad_v) < 2: return False
        insert, tip = cad_v[0], cad_v[-1]
        side = ConnectionSide.left if tip[0] < insert[0] else ConnectionSide.right
        align = TextAlignment.left if side == ConnectionSide.right else TextAlignment.right
        leader_pts = [Vec2(x, y) for (x, y) in cad_v[1:]]
        try:
            mb = msp.add_multileader_mtext(style)
            mb.set_content(ld["text"], char_height=ch, alignment=align)
            mb.add_leader_line(side, leader_pts)
            mb.build(insert=Vec2(insert[0], insert[1]))
            try: mb.multileader.dxf.layer = "ANOTACION"
            except Exception: pass
            return True
        except Exception:
            return False

    def _add_simple_leader_dxf(self, doc, msp, ld, geo):
        """Exporta el Leader simple como entidad LEADER (con punta de flecha) siguiendo
        sus vértices: cabeza → (bisagra) → final. Devuelve True si se creó."""
        verts = [self._to_cad(x, y) for (x, y) in geo["segs"][0]]   # el 1er vértice lleva la flecha
        asz = max(0.05, geo.get("cad_h", LEADER_TEXT_FT) * 0.6)     # tamaño de la punta (unidades CAD)
        dimstyle = "EZDXF" if "EZDXF" in doc.dimstyles else ("Standard" if "Standard" in doc.dimstyles else None)
        if dimstyle is None: return False
        try:
            msp.add_leader(verts, dimstyle=dimstyle, override={"dimasz": asz, "dimscale": 1.0},
                           dxfattribs={"layer": "ANOTACION"})
            return True
        except Exception:
            return False

    def _add_leader_explicit(self, doc, msp, ld, geo):
        """Multileader (o respaldo del Leader simple): geometría exacta (línea + punta + texto)
        agrupada en un GROUP, igual que en la vista previa."""
        ents = []
        for s in geo["segs"]:
            ents.append(msp.add_lwpolyline([self._to_cad(x, y) for (x, y) in s], dxfattribs={"layer": "ANOTACION"}))
        a = self._to_cad(*ld["arrow"]); b = self._to_cad(*geo["segs"][0][1])
        ang = math.atan2(a[1] - b[1], a[0] - b[0]); L = geo.get("cad_h", LEADER_TEXT_FT) * 0.9
        p1 = (a[0] - L * math.cos(ang - 0.28), a[1] - L * math.sin(ang - 0.28))
        p2 = (a[0] - L * math.cos(ang + 0.28), a[1] - L * math.sin(ang + 0.28))
        ents.append(msp.add_solid([a, p1, p2, a], dxfattribs={"layer": "ANOTACION"}))
        if ld["text"]:                                       # Leader simple: solo línea + punta, sin texto
            font = ld.get("font", C.TEXT_FONT); bold = bool(ld.get("bold")); ch = geo.get("cad_h", LEADER_TEXT_FT)
            style = self._text_style(doc, font, bold)
            content = ld["text"].replace("\n", "\\P")
            if bold: content = f"{{\\f{font}|b1;{content}}}"
            m = msp.add_mtext(content, dxfattribs={"layer": "ANOTACION", "style": style, "char_height": ch})
            m.set_location(self._to_cad(*geo["tcenter_px"]), rotation=float(geo.get("cad_rot", 0)), attachment_point=5)
            ents.append(m)
        try:
            g = doc.groups.new()
            with g.edit_data() as data: data.extend(ents)
        except Exception: pass

    def _add_free_text(self, doc, msp, tm):
        font = tm.get("font", C.TEXT_FONT); bold = bool(tm.get("bold")); h = tm.get("size_ft", LEADER_TEXT_FT)
        style = self._text_style(doc, font, bold)
        body = tm["text"].replace("\n", "\\P")
        if bold: body = f"{{\\f{font}|b1;{body}}}"
        m = msp.add_mtext(body, dxfattribs={"layer": "ANOTACION", "style": style, "char_height": h})
        m.set_location(self._to_cad(*tm["pos"]), rotation=float(tm.get("rot", 0) or 0), attachment_point=1)

    def _replace_text(self, doc, msp, tm):
        bx, by, bw, bh = tm.get("box", (tm["pos"][0] - 30, tm["pos"][1] - 10, 60, 20))
        cs = [self._to_cad(bx, by), self._to_cad(bx + bw, by), self._to_cad(bx, by + bh), self._to_cad(bx + bw, by + bh)]
        xs = [c[0] for c in cs]; ys = [c[1] for c in cs]; pad = LEADER_TEXT_FT * 2
        x0, x1, y0, y1 = min(xs) - pad, max(xs) + pad, min(ys) - pad, max(ys) + pad
        for e in list(msp.query("TEXT MTEXT")):
            if e.dxf.layer in ("ANOTACION", "TEXTO"):
                ins = e.dxf.insert
                if x0 <= ins.x <= x1 and y0 <= ins.y <= y1: msp.delete_entity(e)
        t = msp.add_text(tm["text"], height=LEADER_TEXT_FT, dxfattribs={"layer": "ANOTACION", "style": "CAD_TEXT"})
        t.set_placement(self._to_cad(tm["pos"][0], tm["pos"][1]), align=TextEntityAlignment.MIDDLE_CENTER)

    # ─────────────────────────── Excel ───────────────────────────
    def open_excel(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Abrir Excel", DOWNLOADS, "Excel (*.xlsx *.xlsm)")
        if path: self._read_excel(path)

    def _read_excel(self, path):
        self._busy("Leyendo Excel…")
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True); texts = []
            for ws in wb.worksheets:
                try:
                    hc = hr = None
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value is not None and str(cell.value).strip().upper() in ("TEXTO", "TEXTOS"):
                                hc, hr = cell.column, cell.row; break
                        if hc: break
                    if hc is None: continue
                    for row in ws.iter_rows(min_row=hr + 1, min_col=hc, max_col=hc):
                        val = row[0].value
                        if val is not None and str(val).strip(): texts.append(str(val).strip())
                except Exception: continue
            self.text_list.clear(); self.text_list.addItems(texts)
            self._info(f"Excel: {len(texts)} textos de la columna TEXTO." if texts else "No hallé columna TEXTO/TEXTOS.")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error al leer Excel", str(e))
        finally: self._unbusy()

    # ─────────────────────────── Ayuda ───────────────────────────
    def _show_html(self, title, html, w=780, h=660):
        dlg = QtWidgets.QDialog(self); dlg.setWindowTitle(title); dlg.resize(w, h)
        lay = QtWidgets.QVBoxLayout(dlg); tb = QtWidgets.QTextBrowser(); tb.setOpenExternalLinks(True)
        tb.setStyleSheet("background:#1e1e1e;color:#e8e8e8;font-size:14px;"); tb.setHtml(html)
        btn = QtWidgets.QPushButton("Cerrar"); btn.clicked.connect(dlg.accept)
        lay.addWidget(tb); lay.addWidget(btn); dlg.exec()

    def show_about(self):
        dlg = QtWidgets.QDialog(self); dlg.setWindowTitle("Acerca de"); dlg.resize(760, 680)
        lay = QtWidgets.QVBoxLayout(dlg)
        head = QtWidgets.QLabel(
            f"<h2>Digitalizador de planos — utilidades</h2>"
            f"<p><b>Versión {VERSION}</b> · para ingeniería civil (agua, alcantarillado, gas, "
            f"eléctrico, telefonía, drenaje).</p>"
            f"<p>Convierte un PDF de plano a DXF y te deja marcar utilidades, Multileaders y notas "
            f"sobre la imagen, exportando todo en las mismas coordenadas.</p>"
            f"<p style='color:#888;'>Historial de versiones (la más reciente arriba, la más antigua "
            f"abajo). Haz clic en cada versión para desplegarla.</p>")
        head.setWordWrap(True); head.setStyleSheet("color:#e8e8e8;"); lay.addWidget(head)
        box = QtWidgets.QToolBox()
        box.setStyleSheet("QToolBox::tab{background:#333;color:#ddd;border:1px solid #555;}"
                          "QToolBox::tab:selected{background:#3c5a99;color:white;font-weight:bold;}")
        icon = {"added": ("#5fd35f", "✚ nueva"), "removed": ("#e06060", "✖ quitada"),
                "fixed": ("#6cc5e0", "✎ corregida"), "changed": ("#e0c060", "↻ cambiada"),
                "base": ("#cfcfcf", "•")}
        default = ("#cfcfcf", "•")
        for ver, items in CHANGELOG:
            tb = QtWidgets.QTextBrowser(); tb.setStyleSheet("background:#1e1e1e;color:#e8e8e8;border:none;")
            lis = "".join(f'<li style="color:{icon.get(s, default)[0]};margin-bottom:4px;">'
                          f'<b>[{icon.get(s, default)[1]}]</b> {t}</li>' for s, t in items)
            tb.setHtml(f"<ul>{lis}</ul>"); box.addItem(tb, f"v{ver}")
        lay.addWidget(box, 1)
        btn = QtWidgets.QPushButton("Cerrar"); btn.clicked.connect(dlg.accept); lay.addWidget(btn)
        dlg.exec()

    def show_manual(self):
        html = """
        <h2>Manual de usuario</h2>
        <h3>Disposición de la ventana</h3>
        <p>A la <b>izquierda</b> están las herramientas y el flujo de trabajo (páginas, acción,
        tipo de utilidad, Multileader, estilo de texto, reconocimiento). En el <b>centro</b> el plano.
        A la <b>derecha</b> el inventario de lo que has marcado (Utilidades, Multileaders, Textos, Zonas)
        con los botones para editar/eliminar. Arriba una barra de acción (zoom, deshacer/rehacer, imán,
        Exportar DXF) y abajo la barra de estado con el modo, las coordenadas y la escala.</p>
        <h3>1. Abrir el plano</h3>
        <p><b>Archivo → Abrir PDF…</b> (o arrastra el PDF). Navega entre páginas con los botones
        <b>◀ ▶</b> de la barra superior. Puedes arrastrar un <b>Excel</b> para cargar sus textos.</p>
        <h3>2. Dibujar una utilidad</h3>
        <ol>
          <li>Elige el <b>Tipo de utilidad</b> (color por servicio).</li>
          <li>Pulsa <b>✏ Dibujar utilidad</b> (se pone verde) y haz clic punto por punto.</li>
          <li>Pulsa <b>Enter</b> (o clic derecho / doble clic) para finalizar.</li>
          <li>Marca <b>Abandonado</b> para la línea ──/── W ──.</li>
        </ol>
        <h3>3. Colocar un Multileader</h3>
        <ol>
          <li>Elige la orientación: <b>Horizontal</b> (recto), <b>Vertical</b> (recto, texto vertical
              pegado a la línea) o <b>Diagonal</b> (con landing/quiebre).</li>
          <li>Marca <b>Usar texto personalizado</b> y escríbelo, o selecciona un texto del Excel.</li>
          <li>Pulsa <b>↳ Colocar Multileader</b>, clic en la <b>punta</b> y luego dónde va el <b>texto</b>.
              El lado (izq/der o arriba/abajo) lo decide el segundo clic. Queda listo para colocar
              <b>otro</b> sin volver a pulsar el botón; pulsa <b>Esc</b> para salir.</li>
          <li>Para editarlo: <b>doble clic</b> sobre el texto. <b>Enter</b> aplica, <b>Ctrl+Shift+Enter</b>
              hace salto de línea.</li>
        </ol>
        <p>Al exportar, cada Multileader se digitaliza como una <b>entidad MULTILEADER</b> de CAD.</p>
        <h3>4. Texto libre</h3>
        <ol>
          <li>Pulsa <b>T Texto libre</b>, ajusta fuente, altura y negrita (puedes cambiarlos también
              mientras escribes o después, con el texto seleccionado).</li>
          <li>Haz clic donde escribir, teclea y pulsa <b>Enter</b> para agregarlo
              (<b>Ctrl+Shift+Enter</b> = salto de línea).</li>
          <li>Los textos son <b>seleccionables</b> (clic), <b>editables</b> (doble clic) y
              <b>movibles</b> (Editar/mover). Su estilo se cambia al seleccionarlos.</li>
        </ol>
        <h3>5. OCR e ICR</h3>
        <p>Marca <b>Ayudar con textos impresos (OCR)</b> o <b>ICR manuscrita</b>. Aparecen recuadros:
        clic en uno, corrige y confirma con Enter.</p>
        <h3>6. Borrar una zona</h3>
        <p>Pulsa <b>▭ Borrar zona</b>, marca el polígono y pulsa <b>Enter</b>. Se rellena de blanco y
        queda <b>detrás de las tuberías</b> (solo tapa el plano). Al exportar borra la geometría base
        dentro. En la pestaña <b>Zonas borradas</b> puedes activarla/desactivarla, editarla o eliminarla.</p>
        <h3>7. Editar, mover y extender</h3>
        <ul>
          <li>Clic en el dibujo selecciona (sin mover la vista); desde las listas la vista se centra.</li>
          <li><b>Editar/mover</b> (Ctrl+T): arrastra un vértice, clic en un tramo inserta un vértice,
              clic derecho sobre un vértice lo elimina, y arrastrar lejos mueve todo.</li>
          <li><b>Extender un vértice</b>: en editar, haz un clic sobre un vértice (sin arrastrar). Te
              preguntará si la extensión es de la <b>misma utilidad</b> (continúa la línea desde un
              extremo) o una <b>utilidad nueva</b> (rama en forma de F). Luego clic en el dibujo para
              agregar puntos y Enter para finalizar.</li>
          <li><b>Escape</b> quita la selección; de nuevo sale del modo.</li>
        </ul>
        <h3>8. Guardar y exportar</h3>
        <ul>
          <li><b>Guardar proyecto</b> (Ctrl+S): guarda todo en un <b>.digproj</b> (se comparte y abre
              sin el PDF). Si abriste un proyecto, actualiza ese mismo archivo.</li>
          <li><b>Digitalizar plano + exportar DXF</b>: genera el DXF final (plano base + tu marcado).</li>
          <li><b>Cerrar proyecto</b> (Ctrl+W): pregunta si hay cambios sin guardar.</li>
        </ul>
        <h3>9. Historial de versiones</h3>
        <p><b>Ayuda → Acerca de</b>: versión actual e historial desplegable por versión (verde=nueva,
        celeste=corregida, rojo=quitada).</p>
        """
        self._show_html("Manual de usuario", html, 840, 740)

    def show_shortcuts(self):
        rows = [("Ctrl+Z / Ctrl+Shift+Z", "Deshacer / Rehacer"),
                ("Enter", "Aplicar: finaliza utilidad/zona, o agrega texto/edición"),
                ("Ctrl+Shift+Enter", "Salto de línea dentro de un texto o Multileader"),
                ("Escape", "Quitar la selección; si no hay, salir del modo"),
                ("Ctrl+T", "Editar/mover lo seleccionado"),
                ("Ctrl+S / Ctrl+Shift+S", "Guardar proyecto / Guardar como…"),
                ("Ctrl+W", "Cerrar proyecto (pregunta si hay cambios)"),
                ("Doble clic", "Sobre un Multileader o texto: editarlo"),
                ("Clic derecho", "Finaliza línea/zona; en editar, elimina el vértice"),
                ("◀ ▶", "Página anterior / siguiente"),
                ("Rueda", "Zoom · Botón central + arrastrar: desplazar")]
        body = "".join(f'<tr><td style="padding:4px 14px;color:#8bd;"><b>{k}</b></td>'
                       f'<td style="padding:4px;">{d}</td></tr>' for k, d in rows)
        self._show_html("Atajos de teclado", f"<h2>Atajos de teclado</h2><table>{body}</table>", 640, 500)

    # ─────────────────────────── drag & drop ───────────────────────────
    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls(): e.acceptProposedAction()

    def dropEvent(self, e):
        for u in e.mimeData().urls(): self.open_path(u.toLocalFile()); break


def main():
    app = QtWidgets.QApplication(sys.argv)
    win = Main(); win.show()
    if len(sys.argv) > 1:
        win.open_path(sys.argv[1])
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
